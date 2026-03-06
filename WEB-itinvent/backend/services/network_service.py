from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import threading
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Optional

from local_store import get_local_store

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

try:
    import fitz  # type: ignore
except Exception:  # pragma: no cover
    fitz = None


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _h(v: Any) -> str:
    raw = _s(v).lower().replace("-", " ").replace("_", " ").replace("/", " ")
    return re.sub(r"\s+", " ", raw).strip()



def _socket_key(v: Any) -> str:
    # Compare sockets in a tolerant way: ignore case and extra spaces.
    return re.sub(r"\s+", "", _s(v).lower())


def _point_label_from_port(port_name: Any, patch_panel_port: Any) -> str:
    port = _s(port_name)
    socket = _s(patch_panel_port)
    if port and socket:
        return f"PORT {port} · Розетка {socket}"
    if socket:
        return f"Розетка {socket}"
    if port:
        return f"PORT {port}"
    return "Точка"


def _split(v: str) -> list[str]:
    t = _s(v)
    if not t:
        return []
    return [x.strip() for x in re.split(r"\n+|\s{2,}", t) if x.strip()]


def _ips(v: str) -> list[str]:
    return re.findall(r"\b\d{1,3}(?:\.\d{1,3}){3}\b", _s(v))


def _macs(v: str) -> list[str]:
    return re.findall(r"\b[0-9A-Fa-f]{2}(?:[:-][0-9A-Fa-f]{2}){5}\b", _s(v))


def _mac_normalized(v: Any) -> str:
    raw = re.sub(r"[^0-9A-Fa-f]", "", _s(v))
    if len(raw) != 12:
        return ""
    raw = raw.upper()
    return ":".join(raw[i:i + 2] for i in range(0, 12, 2))


def _extract_mac_candidates(v: Any) -> list[str]:
    text = _s(v)
    if not text:
        return []
    matches = re.findall(r"(?:[0-9A-Fa-f]{2}(?:[:-])){5}[0-9A-Fa-f]{2}|[0-9A-Fa-f]{12}", text)
    out: list[str] = []
    for raw in matches:
        normalized = _mac_normalized(raw)
        if normalized and normalized not in out:
            out.append(normalized)
    return out


def _normalize_mac_multiline(value: Any) -> str:
    """Normalize MAC list into canonical multi-line format."""
    normalized = _extract_mac_candidates(value)
    if normalized:
        return "\n".join(normalized)
    return _s(value)


def _append_unique(target: list[str], values: Iterable[Any]) -> None:
    for value in values:
        item = _s(value)
        if item and item not in target:
            target.append(item)


def _extract_ip_candidates(value: Any) -> list[str]:
    out: list[str] = []
    for ip in _ips(_s(value)):
        if ip not in out:
            out.append(ip)
    return out


def _extract_text_candidates(value: Any) -> list[str]:
    out: list[str] = []
    for token in _split(_s(value)):
        if token and token not in out:
            out.append(token)
    return out


def _join_lines(values: list[str]) -> Optional[str]:
    if not values:
        return None
    return "\n".join(values)


def _people_word(count: int) -> str:
    n = int(count)
    n10 = n % 10
    n100 = n % 100
    if n10 == 1 and n100 != 11:
        return "человек"
    if 2 <= n10 <= 4 and not (12 <= n100 <= 14):
        return "человека"
    return "человек"


def _parse_socket_parts(socket_code: Any) -> tuple[Optional[int], Optional[int]]:
    value = _s(socket_code)
    if not value:
        return None, None
    
    # Try to extract numbers for sorting purposes if possible
    # We don't want to enforce strict integer types here if they aren't numbers
    match = re.search(r"(\d+)\s*[/\\\-_.:]\s*(\d+)", value)
    if match:
        return int(match.group(1)), int(match.group(2))
        
    digits = re.findall(r"\d+", value)
    if len(digits) >= 2:
        return int(digits[0]), int(digits[1])
        
    # If no numbers found, just return None. The socket_code itself will hold the raw string.
    return None, None


def _socket_code(panel_no: Optional[int], port_no: Optional[int], fallback: Any = "") -> str:
    if panel_no and port_no:
        return f"{int(panel_no)}/{int(port_no)}"
    return _s(fallback)


def _transliterate(text: str) -> str:
    """
    Transliterate Cyrillic text to Latin characters.
    "Первомайская 19/21" -> "pervomayskaya-19-21"
    """
    # Cyrillic to Latin mapping
    cyrillic_to_latin = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'E',
        'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
        'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
        'Ф': 'F', 'Х': 'Kh', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch',
        'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya',
    }

    result = []
    for char in text:
        if char in cyrillic_to_latin:
            result.append(cyrillic_to_latin[char])
        elif char.isalnum():
            result.append(char)
        elif char in ' /\\-_.:':
            result.append('-')  # Replace spaces and slashes with hyphens
        # Skip other special characters

    # Clean up: remove consecutive hyphens, trim
    result_str = ''.join(result)
    result_str = re.sub(r'-+', '-', result_str)
    result_str = result_str.strip('-').lower()

    return result_str


def _generate_branch_code(name: str, city_code: str, conn: sqlite3.Connection) -> str:
    """
    Generate a unique branch code from the branch name.
    Uses transliteration and adds suffix if code already exists.
    """
    base_code = _transliterate(name)
    if not base_code:
        # Fallback to a generic code if transliteration fails
        base_code = "branch"

    # Combine with city code for uniqueness
    full_code = f"{city_code}-{base_code}"

    # Check if code already exists
    suffix = 1
    candidate_code = full_code
    while True:
        existing = conn.execute(
            "SELECT id FROM network_branches WHERE city_code=? AND branch_code=?",
            (_s(city_code), _s(candidate_code))
        ).fetchone()
        if existing is None:
            return candidate_code
        suffix += 1
        candidate_code = f"{full_code}-{suffix}"


def _vlans(raw: str) -> list[str]:
    out: list[str] = []
    for token in re.split(r"[,;/\s]+", _s(raw).lower()):
        token = token.strip()
        if token and token not in out:
            out.append(token)
    return out


def _occupied(name_raw: str, ip_raw: str, mac_raw: str) -> tuple[int, int]:
    has_any = bool(_s(name_raw) or _s(ip_raw) or _s(mac_raw))
    count = max(len(_split(name_raw)), len(_ips(ip_raw)), len(_macs(mac_raw)), 1 if has_any else 0)
    return count, 1 if has_any else 0


def _site_code(label: str) -> str:
    t = _s(label).lower()
    if "21" in t or "2324" in t or "c9300" in t:
        return "p21"
    return "p19"


def _json_safe(value: Any) -> Any:
    """Convert non-JSON-safe values (e.g. bytes) into compact serializable form."""
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(v) for v in value]
    if isinstance(value, (bytes, bytearray, memoryview)):
        try:
            size = len(value)
        except Exception:
            size = 0
        return {"__type__": "bytes", "size": int(size)}
    return value


class NetworkConflictError(ValueError):
    """Domain conflict (e.g. duplicate socket on the same map)."""


@dataclass
class ImportSummary:
    sheets_total: int = 0
    devices_created: int = 0
    devices_updated: int = 0
    ports_total: int = 0
    ports_created: int = 0
    ports_updated: int = 0
    maps_created: int = 0
    maps_updated: int = 0

    def as_dict(self) -> dict[str, int]:
        return {
            "sheets_total": self.sheets_total,
            "devices_created": self.devices_created,
            "devices_updated": self.devices_updated,
            "ports_total": self.ports_total,
            "ports_created": self.ports_created,
            "ports_updated": self.ports_updated,
            "maps_created": self.maps_created,
            "maps_updated": self.maps_updated,
        }


class NetworkService:
    def __init__(self) -> None:
        self.db_path = Path(get_local_store().db_path)
        self._lock = threading.RLock()
        self._ensure_schema()

    def _connect(self) -> sqlite3.Connection:
        c = sqlite3.connect(self.db_path, timeout=30, check_same_thread=False)
        c.row_factory = sqlite3.Row
        c.execute("PRAGMA foreign_keys=ON;")
        return c

    @staticmethod
    def _has_column(conn: sqlite3.Connection, table_name: str, column_name: str) -> bool:
        rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
        return any(_s(row["name"]).lower() == _s(column_name).lower() for row in rows)

    def _ensure_column(self, conn: sqlite3.Connection, table_name: str, column_name: str, ddl_type: str) -> None:
        if self._has_column(conn, table_name, column_name):
            return
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {ddl_type}")

    def _ensure_schema(self) -> None:
        with self._lock, self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS network_branches(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    city_code TEXT NOT NULL,
                    branch_code TEXT NOT NULL,
                    name TEXT NOT NULL,
                    is_active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(city_code, branch_code)
                );
                CREATE TABLE IF NOT EXISTS network_sites(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_id INTEGER NOT NULL,
                    site_code TEXT NOT NULL,
                    name TEXT NOT NULL,
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(branch_id, site_code),
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS network_devices(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_id INTEGER NOT NULL,
                    site_id INTEGER,
                    device_code TEXT NOT NULL,
                    device_type TEXT NOT NULL DEFAULT 'switch',
                    vendor TEXT,
                    model TEXT,
                    sheet_name TEXT,
                    mgmt_ip TEXT,
                    notes TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(branch_id, device_code),
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE CASCADE,
                    FOREIGN KEY(site_id) REFERENCES network_sites(id) ON DELETE SET NULL
                );
                CREATE TABLE IF NOT EXISTS network_ports(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    device_id INTEGER NOT NULL,
                    port_name TEXT NOT NULL,
                    patch_panel_port TEXT,
                    location_code TEXT,
                    vlan_raw TEXT,
                    vlan_normalized_json TEXT,
                    endpoint_name_raw TEXT,
                    endpoint_ip_raw TEXT,
                    endpoint_mac_raw TEXT,
                    endpoint_count INTEGER NOT NULL DEFAULT 0,
                    is_occupied INTEGER NOT NULL DEFAULT 0,
                    row_source_hash TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(device_id, port_name),
                    FOREIGN KEY(device_id) REFERENCES network_devices(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS network_socket_profiles(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_id INTEGER NOT NULL UNIQUE,
                    panel_count INTEGER NOT NULL,
                    ports_per_panel INTEGER NOT NULL,
                    is_uniform INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS network_panels(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_id INTEGER NOT NULL,
                    panel_index INTEGER NOT NULL,
                    port_count INTEGER NOT NULL,
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(branch_id, panel_index),
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS network_sockets(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_id INTEGER NOT NULL,
                    site_id INTEGER,
                    socket_code TEXT NOT NULL,
                    panel_no INTEGER,
                    port_no INTEGER,
                    port_id INTEGER,
                    device_id INTEGER,
                    mac_address TEXT,
                    fio TEXT,
                    fio_source_db TEXT,
                    fio_resolved_at TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(branch_id, socket_code),
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE CASCADE,
                    FOREIGN KEY(site_id) REFERENCES network_sites(id) ON DELETE SET NULL,
                    FOREIGN KEY(port_id) REFERENCES network_ports(id) ON DELETE SET NULL,
                    FOREIGN KEY(device_id) REFERENCES network_devices(id) ON DELETE SET NULL
                );
                CREATE TABLE IF NOT EXISTS network_branch_db_map(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_id INTEGER NOT NULL UNIQUE,
                    db_id TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    updated_by TEXT,
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS network_maps(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_id INTEGER NOT NULL,
                    site_id INTEGER,
                    title TEXT,
                    floor_label TEXT,
                    file_name TEXT NOT NULL,
                    mime_type TEXT NOT NULL,
                    file_blob BLOB NOT NULL,
                    file_size INTEGER NOT NULL DEFAULT 0,
                    checksum_sha256 TEXT NOT NULL,
                    source_path TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(branch_id, file_name),
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE CASCADE,
                    FOREIGN KEY(site_id) REFERENCES network_sites(id) ON DELETE SET NULL
                );
                CREATE TABLE IF NOT EXISTS network_map_points(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_id INTEGER NOT NULL,
                    map_id INTEGER NOT NULL,
                    site_id INTEGER,
                    device_id INTEGER,
                    port_id INTEGER,
                    socket_id INTEGER,
                    x_ratio REAL NOT NULL,
                    y_ratio REAL NOT NULL,
                    label TEXT,
                    note TEXT,
                    color TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE CASCADE,
                    FOREIGN KEY(map_id) REFERENCES network_maps(id) ON DELETE CASCADE,
                    FOREIGN KEY(site_id) REFERENCES network_sites(id) ON DELETE SET NULL,
                    FOREIGN KEY(device_id) REFERENCES network_devices(id) ON DELETE SET NULL,
                    FOREIGN KEY(port_id) REFERENCES network_ports(id) ON DELETE SET NULL,
                    FOREIGN KEY(socket_id) REFERENCES network_sockets(id) ON DELETE SET NULL
                );
                CREATE TABLE IF NOT EXISTS network_import_jobs(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    city_code TEXT NOT NULL,
                    branch_id INTEGER,
                    status TEXT NOT NULL,
                    started_at TEXT NOT NULL,
                    finished_at TEXT,
                    summary_json TEXT,
                    error_text TEXT,
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE SET NULL
                );
                CREATE TABLE IF NOT EXISTS network_audit_log(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_id INTEGER,
                    entity_type TEXT NOT NULL,
                    entity_id TEXT,
                    action TEXT NOT NULL,
                    diff_json TEXT,
                    actor_user_id INTEGER,
                    actor_role TEXT,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(branch_id) REFERENCES network_branches(id) ON DELETE SET NULL
                );
                CREATE INDEX IF NOT EXISTS idx_network_branches_city ON network_branches(city_code);
                CREATE INDEX IF NOT EXISTS idx_network_devices_branch ON network_devices(branch_id);
                CREATE INDEX IF NOT EXISTS idx_network_ports_device ON network_ports(device_id);
                CREATE INDEX IF NOT EXISTS idx_network_sockets_branch_code ON network_sockets(branch_id, socket_code);
                CREATE INDEX IF NOT EXISTS idx_network_sockets_port ON network_sockets(port_id);
                CREATE INDEX IF NOT EXISTS idx_network_maps_branch ON network_maps(branch_id);
                CREATE INDEX IF NOT EXISTS idx_network_map_points_map ON network_map_points(map_id);
                CREATE INDEX IF NOT EXISTS idx_network_map_points_device ON network_map_points(device_id);
                CREATE INDEX IF NOT EXISTS idx_network_map_points_branch ON network_map_points(branch_id);
                CREATE INDEX IF NOT EXISTS idx_network_audit_branch ON network_audit_log(branch_id);
                """
            )
            self._ensure_column(conn, "network_map_points", "socket_id", "INTEGER")
            self._ensure_column(conn, "network_sockets", "site_id", "INTEGER")
            self._ensure_column(conn, "network_sockets", "port_id", "INTEGER")
            self._ensure_column(conn, "network_sockets", "device_id", "INTEGER")
            self._ensure_column(conn, "network_sockets", "mac_address", "TEXT")
            self._ensure_column(conn, "network_sockets", "fio", "TEXT")
            self._ensure_column(conn, "network_sockets", "fio_source_db", "TEXT")
            self._ensure_column(conn, "network_sockets", "fio_resolved_at", "TEXT")
            self._ensure_column(conn, "network_socket_profiles", "is_uniform", "INTEGER NOT NULL DEFAULT 1")
            self._ensure_column(conn, "network_branches", "default_site_code", "TEXT")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_network_map_points_socket ON network_map_points(socket_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_network_panels_branch ON network_panels(branch_id)")
            # Bootstrap migration: enforce one socket (PORT P/P) -> one point per map.
            self._repair_map_point_socket_conflicts_in_conn(
                conn,
                actor_user_id=None,
                actor_role="system_bootstrap",
            )
            conn.commit()

    @staticmethod
    def _d(row: sqlite3.Row | None) -> dict[str, Any] | None:
        if row is None:
            return None
        return {k: row[k] for k in row.keys()}

    def _audit(
        self,
        conn: sqlite3.Connection,
        *,
        branch_id: Optional[int],
        entity_type: str,
        entity_id: Optional[str],
        action: str,
        diff: dict[str, Any],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> None:
        safe_diff = _json_safe(diff)
        conn.execute(
            """
            INSERT INTO network_audit_log(branch_id, entity_type, entity_id, action, diff_json, actor_user_id, actor_role, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (branch_id, entity_type, entity_id, action, json.dumps(safe_diff, ensure_ascii=False), actor_user_id, actor_role, _now()),
        )

    @staticmethod
    def _point_conflict_detail(*, map_id: int, point_id: int, socket_value: str) -> str:
        socket_text = _s(socket_value) or "не указано"
        return f'Розетка "{socket_text}" уже привязана к другой точке (map_id={int(map_id)}, point_id={int(point_id)})'

    def _map_point_row_with_links(self, conn: sqlite3.Connection, *, point_id: int) -> sqlite3.Row | None:
        return conn.execute(
            """
            SELECT
                mp.*,
                m.title map_title,
                m.file_name map_file_name,
                d.device_code,
                d.model device_model,
                p.port_name,
                COALESCE(s.socket_code, p.patch_panel_port) patch_panel_port,
                s.socket_code,
                p.location_code port_location_code,
                p.vlan_raw port_vlan_raw,
                p.endpoint_name_raw,
                p.endpoint_ip_raw,
                COALESCE(s.mac_address, p.endpoint_mac_raw) endpoint_mac_raw,
                s.fio,
                s.fio_source_db
            FROM network_map_points mp
            LEFT JOIN network_maps m ON m.id=mp.map_id
            LEFT JOIN network_devices d ON d.id=mp.device_id
            LEFT JOIN network_ports p ON p.id=mp.port_id
            LEFT JOIN network_sockets s ON s.id=mp.socket_id
            WHERE mp.id=?
            """,
            (int(point_id),),
        ).fetchone()

    def _map_point_public_dict(self, row: sqlite3.Row | None) -> dict[str, Any]:
        item = self._d(row) or {}
        item["socket_key"] = _socket_key(item.get("patch_panel_port"))
        item["point_identity"] = _s(item.get("socket_code") or item.get("patch_panel_port")) or None
        return item

    def _map_row_for_api(self, conn: sqlite3.Connection, *, map_id: int) -> dict[str, Any]:
        row = conn.execute(
            """
            SELECT
                m.id,
                m.branch_id,
                m.site_id,
                s.site_code,
                s.name site_name,
                m.title,
                m.floor_label,
                m.file_name,
                m.mime_type,
                m.file_size,
                m.checksum_sha256,
                m.source_path,
                m.created_at,
                m.updated_at
            FROM network_maps m
            LEFT JOIN network_sites s ON s.id=m.site_id
            WHERE m.id=?
            """,
            (int(map_id),),
        ).fetchone()
        return self._d(row) or {}

    @staticmethod
    def _canonical_socket_code(value: Any) -> str:
        # Just clean up the string to allow letters and symbols like "6/1.1" or "A/12"
        raw = _s(value)
        if not raw:
            return ""
        # Remove any excess whitespace and return the exact string entered by the user
        cleaned = re.sub(r"\s+", "", raw).upper()
        return cleaned

    def _socket_row_with_links(self, conn: sqlite3.Connection, *, socket_id: int) -> sqlite3.Row | None:
        return conn.execute(
            """
            SELECT
                s.*,
                d.device_code,
                d.model device_model,
                p.port_name,
                p.patch_panel_port,
                p.location_code,
                p.vlan_raw,
                p.endpoint_name_raw,
                p.endpoint_ip_raw,
                p.endpoint_mac_raw,
                mp.id map_point_id,
                mp.map_id,
                mp.x_ratio,
                mp.y_ratio
            FROM network_sockets s
            LEFT JOIN network_devices d ON d.id=s.device_id
            LEFT JOIN network_ports p ON p.id=s.port_id
            LEFT JOIN network_map_points mp ON mp.socket_id=s.id
            WHERE s.id=?
            """,
            (int(socket_id),),
        ).fetchone()

    def get_branch_db_mapping(self, branch_id: int) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            branch = conn.execute("SELECT id FROM network_branches WHERE id=?", (int(branch_id),)).fetchone()
            if branch is None:
                raise ValueError("Branch not found")
            row = conn.execute(
                "SELECT branch_id, db_id, updated_at, updated_by FROM network_branch_db_map WHERE branch_id=?",
                (int(branch_id),),
            ).fetchone()
            if row is None:
                return {"branch_id": int(branch_id), "db_id": None, "updated_at": None, "updated_by": None}
            return self._d(row) or {"branch_id": int(branch_id), "db_id": None}

    def update_branch_db_mapping(
        self,
        *,
        branch_id: int,
        db_id: str,
        updated_by: Optional[str],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        db_text = _s(db_id)
        if not db_text:
            raise ValueError("db_id is required")
        now = _now()
        with self._lock, self._connect() as conn:
            branch = conn.execute("SELECT id FROM network_branches WHERE id=?", (int(branch_id),)).fetchone()
            if branch is None:
                raise ValueError("Branch not found")
            before = conn.execute("SELECT * FROM network_branch_db_map WHERE branch_id=?", (int(branch_id),)).fetchone()
            conn.execute(
                """
                INSERT INTO network_branch_db_map(branch_id, db_id, updated_at, updated_by)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(branch_id) DO UPDATE SET
                    db_id=excluded.db_id,
                    updated_at=excluded.updated_at,
                    updated_by=excluded.updated_by
                """,
                (int(branch_id), db_text, now, _s(updated_by) or None),
            )
            after = conn.execute("SELECT * FROM network_branch_db_map WHERE branch_id=?", (int(branch_id),)).fetchone()
            self._audit(
                conn,
                branch_id=int(branch_id),
                entity_type="branch_db_mapping",
                entity_id=str(branch_id),
                action="update",
                diff={"before": self._d(before), "after": self._d(after)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._d(after) or {}

    def update_branch_default_site_code(
        self,
        conn: sqlite3.Connection,
        *,
        branch_id: int,
        default_site_code: Optional[str],
    ) -> bool:
        """Update the default site code for a branch."""
        now = _now()
        cursor = conn.execute(
            """
            UPDATE network_branches
            SET default_site_code=?, updated_at=?
            WHERE id=?
            """,
            (_s(default_site_code) or None, now, int(branch_id)),
        )
        return cursor.rowcount > 0

    def update_branch(
        self,
        *,
        branch_id: int,
        branch_name: Optional[str],
        default_site_code: Optional[str],
        db_id: Optional[str],
        updated_by: Optional[str],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        now = _now()
        with self._lock, self._connect() as conn:
            before_branch = conn.execute("SELECT * FROM network_branches WHERE id=?", (int(branch_id),)).fetchone()
            if before_branch is None:
                raise ValueError("Branch not found")
            before_mapping = conn.execute(
                "SELECT * FROM network_branch_db_map WHERE branch_id=?",
                (int(branch_id),),
            ).fetchone()

            branch_fields: list[str] = []
            branch_params: list[Any] = []
            if branch_name is not None:
                branch_name_text = _s(branch_name)
                if not branch_name_text:
                    raise ValueError("branch_name is required")
                branch_fields.append("name=?")
                branch_params.append(branch_name_text)
            if default_site_code is not None:
                branch_fields.append("default_site_code=?")
                branch_params.append(_s(default_site_code) or None)

            if not branch_fields and db_id is None:
                raise ValueError("No fields to update")

            if branch_fields:
                branch_fields.append("updated_at=?")
                branch_params.append(now)
                branch_params.append(int(branch_id))
                conn.execute(
                    f"UPDATE network_branches SET {', '.join(branch_fields)} WHERE id=?",
                    tuple(branch_params),
                )

            if db_id is not None:
                db_text = _s(db_id)
                if db_text:
                    conn.execute(
                        """
                        INSERT INTO network_branch_db_map(branch_id, db_id, updated_at, updated_by)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(branch_id) DO UPDATE SET
                            db_id=excluded.db_id,
                            updated_at=excluded.updated_at,
                            updated_by=excluded.updated_by
                        """,
                        (int(branch_id), db_text, now, _s(updated_by) or None),
                    )
                else:
                    conn.execute(
                        "DELETE FROM network_branch_db_map WHERE branch_id=?",
                        (int(branch_id),),
                    )

            after_branch = conn.execute("SELECT * FROM network_branches WHERE id=?", (int(branch_id),)).fetchone()
            after_mapping = conn.execute(
                "SELECT * FROM network_branch_db_map WHERE branch_id=?",
                (int(branch_id),),
            ).fetchone()

            self._audit(
                conn,
                branch_id=int(branch_id),
                entity_type="branch",
                entity_id=str(branch_id),
                action="update",
                diff={
                    "before": self._d(before_branch),
                    "after": self._d(after_branch),
                    "db_mapping_before": self._d(before_mapping),
                    "db_mapping_after": self._d(after_mapping),
                },
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()

            out = self._d(after_branch) or {}
            out["db_id"] = _s(after_mapping["db_id"]) if after_mapping is not None else None
            return out

    def _upsert_socket_profile_in_conn(
        self,
        conn: sqlite3.Connection,
        *,
        branch_id: int,
        panel_count: int,
        ports_per_panel: int,
        is_uniform: bool = True,
    ) -> dict[str, Any]:
        if int(panel_count) <= 0 or int(ports_per_panel) <= 0:
            raise ValueError("panel_count and ports_per_panel must be positive")
        now = _now()
        conn.execute(
            """
            INSERT INTO network_socket_profiles(branch_id, panel_count, ports_per_panel, is_uniform, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(branch_id) DO UPDATE SET
                panel_count=excluded.panel_count,
                ports_per_panel=excluded.ports_per_panel,
                is_uniform=excluded.is_uniform,
                updated_at=excluded.updated_at
            """,
            (int(branch_id), int(panel_count), int(ports_per_panel), 1 if is_uniform else 0, now, now),
        )
        row = conn.execute("SELECT * FROM network_socket_profiles WHERE branch_id=?", (int(branch_id),)).fetchone()
        return self._d(row) or {}

    def _upsert_panels_in_conn(
        self,
        conn: sqlite3.Connection,
        *,
        branch_id: int,
        panels: list[dict[str, int]],
    ) -> dict[str, Any]:
        """Upsert individual panel configurations for heterogeneous mode.

        Args:
            conn: Database connection
            branch_id: Branch ID
            panels: List of dicts with 'panel_index' (1-based) and 'port_count' keys

        Returns:
            Dict with 'created', 'updated', 'deleted' counts
        """
        if not panels:
            raise ValueError("panels list cannot be empty")

        now = _now()
        # Get existing panel indices
        existing = {
            row["panel_index"]: int(row["id"])
            for row in conn.execute(
                "SELECT id, panel_index FROM network_panels WHERE branch_id=?",
                (int(branch_id),),
            ).fetchall()
        }

        # Track stats
        created = 0
        updated = 0
        deleted = 0

        # Upsert panels
        for panel in panels:
            panel_index = int(panel.get("panel_index", 0))
            port_count = int(panel.get("port_count", 0))
            if panel_index <= 0 or port_count <= 0:
                raise ValueError("panel_index and port_count must be positive")

            if panel_index in existing:
                conn.execute(
                    """
                    UPDATE network_panels
                    SET port_count=?, updated_at=?
                    WHERE id=?
                    """,
                    (port_count, now, existing[panel_index]),
                )
                updated += 1
            else:
                conn.execute(
                    """
                    INSERT INTO network_panels(branch_id, panel_index, port_count, sort_order, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (int(branch_id), panel_index, port_count, panel_index, now, now),
                )
                created += 1

        # Delete panels not in the new list
        new_indices = {int(p.get("panel_index", 0)) for p in panels}
        for panel_index, panel_id in existing.items():
            if panel_index not in new_indices:
                conn.execute("DELETE FROM network_panels WHERE id=?", (panel_id,))
                deleted += 1

        return {"created": created, "updated": updated, "deleted": deleted}

    def get_socket_profile(self, branch_id: int) -> dict[str, Any] | None:
        with self._lock, self._connect() as conn:
            row = conn.execute("SELECT * FROM network_socket_profiles WHERE branch_id=?", (int(branch_id),)).fetchone()
            return self._d(row)

    def _bootstrap_sockets_for_branch_in_conn(self, conn: sqlite3.Connection, *, branch_id: int) -> dict[str, int]:
        profile = conn.execute("SELECT * FROM network_socket_profiles WHERE branch_id=?", (int(branch_id),)).fetchone()
        if profile is None:
            return {"created": 0, "updated": 0, "total": 0}

        # Check if using heterogeneous panels
        is_uniform_val = profile["is_uniform"]
        is_uniform = bool(int(is_uniform_val)) if is_uniform_val is not None else True
        panels = None
        if not is_uniform:
            panels = [
                {"panel_index": int(row["panel_index"]), "port_count": int(row["port_count"])}
                for row in conn.execute(
                    "SELECT panel_index, port_count FROM network_panels WHERE branch_id=? ORDER BY sort_order, panel_index",
                    (int(branch_id),),
                ).fetchall()
            ]

        # Fallback to uniform mode if no panels defined
        if not is_uniform and not panels:
            is_uniform = True

        if is_uniform:
            # Uniform mode: use panel_count and ports_per_panel from profile
            panel_count = int(profile["panel_count"] or 0)
            ports_per_panel = int(profile["ports_per_panel"] or 0)
            if panel_count <= 0 or ports_per_panel <= 0:
                return {"created": 0, "updated": 0, "total": 0}
            panels = [{"panel_index": i, "port_count": ports_per_panel} for i in range(1, panel_count + 1)]

        if not panels:
            return {"created": 0, "updated": 0, "total": 0}

        now = _now()
        site_row = conn.execute(
            "SELECT id FROM network_sites WHERE branch_id=? ORDER BY sort_order, id LIMIT 1",
            (int(branch_id),),
        ).fetchone()
        site_id = int(site_row["id"]) if site_row else None
        created = 0
        updated = 0
        total = 0
        for panel in panels:
            panel_no = int(panel["panel_index"])
            port_count = int(panel["port_count"])
            for port_no in range(1, port_count + 1):
                total += 1
                socket_code = _socket_code(panel_no, port_no)
                exists = conn.execute(
                    "SELECT id FROM network_sockets WHERE branch_id=? AND socket_code=?",
                    (int(branch_id), socket_code),
                ).fetchone()
                if exists is None:
                    conn.execute(
                        """
                        INSERT INTO network_sockets(
                            branch_id, site_id, socket_code, panel_no, port_no,
                            created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            int(branch_id),
                            site_id,
                            socket_code,
                            panel_no,
                            port_no,
                            now,
                            now,
                        ),
                    )
                    created += 1
                else:
                    conn.execute(
                        """
                        UPDATE network_sockets
                        SET panel_no=?, port_no=?, updated_at=?
                        WHERE id=?
                        """,
                        (panel_no, port_no, now, int(exists["id"])),
                    )
                    updated += 1
        return {"created": int(created), "updated": int(updated), "total": int(total)}

    def bootstrap_branch_sockets(
        self,
        *,
        branch_id: int,
        panel_count: Optional[int],
        ports_per_panel: Optional[int],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            branch = conn.execute("SELECT * FROM network_branches WHERE id=?", (int(branch_id),)).fetchone()
            if branch is None:
                raise ValueError("Branch not found")
            if panel_count is not None or ports_per_panel is not None:
                current = conn.execute("SELECT * FROM network_socket_profiles WHERE branch_id=?", (int(branch_id),)).fetchone()
                resolved_panel_count = int(panel_count or (current["panel_count"] if current else 0) or 0)
                resolved_ports_per_panel = int(ports_per_panel or (current["ports_per_panel"] if current else 0) or 0)
                self._upsert_socket_profile_in_conn(
                    conn,
                    branch_id=int(branch_id),
                    panel_count=resolved_panel_count,
                    ports_per_panel=resolved_ports_per_panel,
                )

            summary = self._bootstrap_sockets_for_branch_in_conn(conn, branch_id=int(branch_id))
            self._sync_all_sockets_in_conn(conn, branch_id=int(branch_id))
            self._link_map_points_to_sockets_in_conn(conn, branch_id=int(branch_id))
            self._audit(
                conn,
                branch_id=int(branch_id),
                entity_type="socket_profile",
                entity_id=str(branch_id),
                action="bootstrap",
                diff={"summary": summary},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            profile = conn.execute("SELECT * FROM network_socket_profiles WHERE branch_id=?", (int(branch_id),)).fetchone()
            return {
                "branch_id": int(branch_id),
                "profile": self._d(profile),
                "summary": summary,
            }

    def _sync_socket_for_port_in_conn(self, conn: sqlite3.Connection, *, port_id: int) -> Optional[int]:
        affected_socket_ids = [
            int(item["id"])
            for item in conn.execute("SELECT id FROM network_sockets WHERE port_id=?", (int(port_id),)).fetchall()
            if item is not None and item["id"] is not None
        ]
        row = conn.execute(
            """
            SELECT
                p.id,
                p.patch_panel_port,
                p.endpoint_mac_raw,
                d.id device_id,
                d.branch_id,
                d.site_id
            FROM network_ports p
            JOIN network_devices d ON d.id=p.device_id
            WHERE p.id=?
            """,
            (int(port_id),),
        ).fetchone()
        if row is None:
            conn.execute("UPDATE network_sockets SET port_id=NULL, device_id=NULL, mac_address=NULL, updated_at=? WHERE port_id=?", (_now(), int(port_id)))
            self._sync_map_points_by_socket_ids_in_conn(conn, socket_ids=affected_socket_ids)
            return None

        socket_code = self._canonical_socket_code(row["patch_panel_port"])
        if not socket_code:
            conn.execute("UPDATE network_sockets SET port_id=NULL, device_id=NULL, mac_address=NULL, updated_at=? WHERE port_id=?", (_now(), int(port_id)))
            self._sync_map_points_by_socket_ids_in_conn(conn, socket_ids=affected_socket_ids)
            return None

        panel_no, port_no = _parse_socket_parts(socket_code)
        now = _now()
        mac_raw = _normalize_mac_multiline(row["endpoint_mac_raw"])
        existing = conn.execute(
            "SELECT id FROM network_sockets WHERE branch_id=? AND socket_code=?",
            (int(row["branch_id"]), socket_code),
        ).fetchone()

        if existing is None:
            conn.execute(
                """
                INSERT INTO network_sockets(
                    branch_id, site_id, socket_code, panel_no, port_no,
                    port_id, device_id, mac_address, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(row["branch_id"]),
                    row["site_id"],
                    socket_code,
                    panel_no,
                    port_no,
                    int(port_id),
                    int(row["device_id"]),
                    mac_raw or None,
                    now,
                    now,
                ),
            )
            socket_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        else:
            socket_id = int(existing["id"])
            conn.execute(
                """
                UPDATE network_sockets
                SET site_id=?, panel_no=?, port_no=?, port_id=?, device_id=?, mac_address=?, updated_at=?
                WHERE id=?
                """,
                (
                    row["site_id"],
                    panel_no,
                    port_no,
                    int(port_id),
                    int(row["device_id"]),
                    mac_raw or None,
                    now,
                    socket_id,
                ),
            )

        if socket_id not in affected_socket_ids:
            affected_socket_ids.append(socket_id)
        conn.execute("UPDATE network_sockets SET port_id=NULL, device_id=NULL, mac_address=NULL, updated_at=? WHERE port_id=? AND id<>?", (now, int(port_id), socket_id))
        self._sync_map_points_by_socket_ids_in_conn(conn, socket_ids=affected_socket_ids)
        return socket_id

    def _sync_map_points_by_socket_ids_in_conn(self, conn: sqlite3.Connection, *, socket_ids: Iterable[int]) -> None:
        normalized_ids: list[int] = []
        seen: set[int] = set()
        for raw_id in socket_ids:
            sid = int(raw_id or 0)
            if sid <= 0 or sid in seen:
                continue
            seen.add(sid)
            normalized_ids.append(sid)
        if not normalized_ids:
            return
        now = _now()
        for sid in normalized_ids:
            socket_row = conn.execute(
                "SELECT id, port_id, device_id FROM network_sockets WHERE id=?",
                (int(sid),),
            ).fetchone()
            if socket_row is None:
                continue
            conn.execute(
                """
                UPDATE network_map_points
                SET port_id=?, device_id=?, updated_at=?
                WHERE socket_id=?
                """,
                (
                    socket_row["port_id"],
                    socket_row["device_id"],
                    now,
                    int(sid),
                ),
            )

    def _sync_all_sockets_in_conn(self, conn: sqlite3.Connection, branch_id: Optional[int] = None) -> None:
        params: list[Any] = []
        where = ""
        if branch_id is not None:
            where = "WHERE d.branch_id=?"
            params.append(int(branch_id))
        rows = conn.execute(
            f"""
            SELECT p.id
            FROM network_ports p
            JOIN network_devices d ON d.id=p.device_id
            {where}
            """,
            params,
        ).fetchall()
        for row in rows:
            self._sync_socket_for_port_in_conn(conn, port_id=int(row["id"]))

    def _link_map_points_to_sockets_in_conn(self, conn: sqlite3.Connection, branch_id: Optional[int] = None) -> None:
        params: list[Any] = []
        where = ""
        if branch_id is not None:
            where = "WHERE mp.branch_id=?"
            params.append(int(branch_id))
        points = conn.execute(
            f"""
            SELECT mp.id point_id, mp.port_id, mp.socket_id, mp.branch_id, p.patch_panel_port
            FROM network_map_points mp
            LEFT JOIN network_ports p ON p.id=mp.port_id
            {where}
            """,
            params,
        ).fetchall()
        for point in points:
            point_id = int(point["point_id"])
            if point["socket_id"] is not None:
                self._sync_map_points_by_socket_ids_in_conn(conn, socket_ids=[int(point["socket_id"])])
                continue
            if point["port_id"] is not None:
                socket_id = self._sync_socket_for_port_in_conn(conn, port_id=int(point["port_id"]))
                if socket_id:
                    conn.execute(
                        "UPDATE network_map_points SET socket_id=?, updated_at=? WHERE id=?",
                        (int(socket_id), _now(), point_id),
                    )
                    self._sync_map_points_by_socket_ids_in_conn(conn, socket_ids=[int(socket_id)])
                    continue
            socket_code = self._canonical_socket_code(point["patch_panel_port"])
            if not socket_code:
                continue
            socket_row = conn.execute(
                "SELECT id FROM network_sockets WHERE branch_id=? AND socket_code=?",
                (int(point["branch_id"]), socket_code),
            ).fetchone()
            if socket_row is not None:
                socket_id = int(socket_row["id"])
                conn.execute(
                    "UPDATE network_map_points SET socket_id=?, updated_at=? WHERE id=?",
                    (socket_id, _now(), point_id),
                )
                self._sync_map_points_by_socket_ids_in_conn(conn, socket_ids=[socket_id])

    def list_sockets(self, branch_id: int, *, search: str = "", limit: int = 5000) -> list[dict[str, Any]]:
        where = ["s.branch_id=?"]
        params: list[Any] = [int(branch_id)]
        if _s(search):
            where.append(
                "("
                "COALESCE(s.socket_code,'') LIKE ? OR "
                "COALESCE(d.device_code,'') LIKE ? OR "
                "COALESCE(p.port_name,'') LIKE ? OR "
                "COALESCE(p.location_code,'') LIKE ? OR "
                "COALESCE(p.vlan_raw,'') LIKE ? OR "
                "COALESCE(p.endpoint_ip_raw,'') LIKE ? OR "
                "COALESCE(s.mac_address,'') LIKE ? OR "
                "COALESCE(s.fio,'') LIKE ?"
                ")"
            )
            v = f"%{_s(search)}%"
            params.extend([v, v, v, v, v, v, v, v])
        params.append(int(max(1, min(int(limit), 10000))))
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                f"""
                SELECT
                    s.*,
                    d.device_code,
                    d.model device_model,
                    p.port_name,
                    p.location_code,
                    p.vlan_raw,
                    p.endpoint_name_raw,
                    p.endpoint_ip_raw,
                    p.endpoint_mac_raw,
                    mp.id map_point_id,
                    mp.map_id,
                    mp.x_ratio,
                    mp.y_ratio
                FROM network_sockets s
                LEFT JOIN network_devices d ON d.id=s.device_id
                LEFT JOIN network_ports p ON p.id=s.port_id
                LEFT JOIN network_map_points mp ON mp.socket_id=s.id
                WHERE {' AND '.join(where)}
                ORDER BY s.panel_no, s.port_no, s.socket_code COLLATE NOCASE
                LIMIT ?
                """,
                params,
            ).fetchall()
            return [self._d(row) for row in rows if row is not None]

    def create_socket(
        self,
        *,
        branch_id: int,
        payload: dict[str, Any],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        socket_code = self._canonical_socket_code(payload.get("socket_code"))
        if not socket_code:
            raise ValueError("socket_code is required")

        panel_no, port_no = _parse_socket_parts(socket_code)
        mac_address = _normalize_mac_multiline(payload.get("mac_address"))
        now = _now()

        with self._lock, self._connect() as conn:
            branch_row = conn.execute(
                "SELECT id, default_site_code FROM network_branches WHERE id=?",
                (int(branch_id),),
            ).fetchone()
            if branch_row is None:
                raise ValueError("Branch not found")

            conflict = conn.execute(
                "SELECT id FROM network_sockets WHERE branch_id=? AND socket_code=?",
                (int(branch_id), socket_code),
            ).fetchone()
            if conflict is not None:
                raise NetworkConflictError(f'Розетка "{socket_code}" уже существует в филиале')

            default_site_code = _s(branch_row["default_site_code"])
            site_id = None
            if default_site_code:
                site_id = self._ensure_site(
                    conn,
                    branch_id=int(branch_id),
                    site_code=default_site_code,
                    site_name=default_site_code,
                )

            conn.execute(
                """
                INSERT INTO network_sockets(
                    branch_id, site_id, socket_code, panel_no, port_no,
                    port_id, device_id, mac_address, fio, fio_source_db, fio_resolved_at,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, NULL, NULL, ?, NULL, NULL, NULL, ?, ?)
                """,
                (
                    int(branch_id),
                    site_id,
                    socket_code,
                    panel_no,
                    port_no,
                    mac_address or None,
                    now,
                    now,
                ),
            )

            new_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            row = self._socket_row_with_links(conn, socket_id=new_id)
            self._audit(
                conn,
                branch_id=int(branch_id),
                entity_type="socket",
                entity_id=str(new_id),
                action="create",
                diff={"after": self._d(row)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._d(row) or {}

    def update_socket(
        self,
        *,
        socket_id: int,
        payload: dict[str, Any],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        patch = {k: v for k, v in payload.items() if k in {"socket_code", "port_id", "mac_address"}}
        if not patch:
            raise ValueError("No fields to update")
        with self._lock, self._connect() as conn:
            before = conn.execute("SELECT * FROM network_sockets WHERE id=?", (int(socket_id),)).fetchone()
            if before is None:
                raise ValueError("Socket not found")
            fields: list[str] = []
            params: list[Any] = []

            if "socket_code" in patch:
                next_code = self._canonical_socket_code(patch.get("socket_code"))
                if not next_code:
                    raise ValueError("socket_code is required")
                conflict = conn.execute(
                    "SELECT id FROM network_sockets WHERE branch_id=? AND socket_code=? AND id<>?",
                    (int(before["branch_id"]), next_code, int(socket_id)),
                ).fetchone()
                if conflict is not None:
                    raise NetworkConflictError(f'Розетка "{next_code}" уже существует в филиале')
                panel_no, port_no = _parse_socket_parts(next_code)
                fields.extend(["socket_code=?", "panel_no=?", "port_no=?"])
                params.extend([next_code, panel_no, port_no])

            if "port_id" in patch:
                port_id = patch.get("port_id")
                if port_id in ("", 0, "0", None):
                    fields.extend(["port_id=?", "device_id=?"])
                    params.extend([None, None])
                else:
                    port_row = self._validate_port_for_map_point(
                        conn,
                        branch_id=int(before["branch_id"]),
                        port_id=int(port_id),
                        expected_device_id=None,
                    )
                    socket_code = self._canonical_socket_code(port_row["patch_panel_port"])
                    if socket_code:
                        conflict = conn.execute(
                            "SELECT id FROM network_sockets WHERE branch_id=? AND socket_code=? AND id<>?",
                            (int(before["branch_id"]), socket_code, int(socket_id)),
                        ).fetchone()
                        if conflict is not None:
                            raise NetworkConflictError(f'Розетка "{socket_code}" уже существует в филиале')
                        panel_no, port_no = _parse_socket_parts(socket_code)
                        fields.extend(["socket_code=?", "panel_no=?", "port_no=?"])
                        params.extend([socket_code, panel_no, port_no])
                    fields.extend(["port_id=?", "device_id=?", "mac_address=?"])
                    params.extend([int(port_id), int(port_row["device_id"]), _normalize_mac_multiline(port_row["endpoint_mac_raw"]) or None])
                    conn.execute("UPDATE network_map_points SET port_id=?, device_id=? WHERE socket_id=?", (int(port_id), int(port_row["device_id"]), int(socket_id)))

            if "mac_address" in patch:
                fields.append("mac_address=?")
                params.append(_normalize_mac_multiline(patch.get("mac_address")) or None)

            fields.append("updated_at=?")
            params.append(_now())
            params.append(int(socket_id))
            conn.execute(f"UPDATE network_sockets SET {', '.join(fields)} WHERE id=?", params)
            after = self._socket_row_with_links(conn, socket_id=int(socket_id))
            self._audit(
                conn,
                branch_id=int(before["branch_id"]),
                entity_type="socket",
                entity_id=str(socket_id),
                action="update",
                diff={"before": self._d(before), "after": self._d(after)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._d(after) or {}

    def delete_socket(self, *, socket_id: int, actor_user_id: Optional[int], actor_role: Optional[str]) -> bool:
        with self._lock, self._connect() as conn:
            before = conn.execute("SELECT * FROM network_sockets WHERE id=?", (int(socket_id),)).fetchone()
            if before is None:
                return False

            socket_id_int = int(before["id"])
            branch_id = int(before["branch_id"])
            now = _now()

            port_id = int(before["port_id"]) if before["port_id"] is not None else None
            if port_id:
                # If socket was tied to a port, clear PORT P/P so auto-sync does not recreate it.
                conn.execute(
                    "UPDATE network_ports SET patch_panel_port=NULL, row_source_hash=NULL, updated_at=? WHERE id=?",
                    (now, port_id),
                )
                self._sync_socket_for_port_in_conn(conn, port_id=port_id)

            conn.execute(
                """
                UPDATE network_map_points
                SET socket_id=NULL, port_id=NULL, device_id=NULL, updated_at=?
                WHERE socket_id=?
                """,
                (now, socket_id_int),
            )
            conn.execute("DELETE FROM network_sockets WHERE id=?", (socket_id_int,))
            self._audit(
                conn,
                branch_id=branch_id,
                entity_type="socket",
                entity_id=str(socket_id_int),
                action="delete",
                diff={"before": self._d(before)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return True

    @staticmethod
    def _resolve_fio_by_mac_candidates(db_id: str, mac_addresses: list[str]) -> dict[str, Any] | None:
        if not db_id:
            return None
        try:
            from backend.database.connection import get_db  # local import to avoid hard dependency in bot runtime
        except Exception:
            return None
        normalized_macs = [_mac_normalized(v) for v in mac_addresses or []]
        normalized_macs = [v for v in normalized_macs if v]
        if not normalized_macs:
            return None

        normalized_raw = list(dict.fromkeys(value.replace(":", "") for value in normalized_macs))
        placeholders = ", ".join(["?"] * len(normalized_raw))
        query = """
            SELECT
                UPPER(REPLACE(REPLACE(COALESCE(i.MAC_ADDRESS,''),':',''),'-','')) as mac_norm,
                o.OWNER_DISPLAY_NAME as fio,
                o.OWNER_NO as owner_no,
                i.MAC_ADDRESS as mac_address,
                i.CH_DATE as changed_at
            FROM ITEMS i
            LEFT JOIN OWNERS o ON o.OWNER_NO = i.EMPL_NO
            WHERE i.CI_TYPE = 1
              AND UPPER(REPLACE(REPLACE(COALESCE(i.MAC_ADDRESS,''),':',''),'-','')) IN ({placeholders})
              AND COALESCE(o.OWNER_DISPLAY_NAME,'') <> ''
            ORDER BY i.CH_DATE DESC, o.OWNER_DISPLAY_NAME
        """.format(placeholders=placeholders)
        try:
            db = get_db(db_id)
            rows = db.execute_query(query, tuple(normalized_raw))
        except Exception:
            return None
        if not rows:
            return None
        latest_owner_by_mac: dict[str, dict[str, Any]] = {}
        for row in rows:
            rec = row or {}
            mac_norm = _s(rec.get("mac_norm") or rec.get("MAC_NORM")).upper()
            fio = _s(rec.get("fio") or rec.get("OWNER_DISPLAY_NAME"))
            if not mac_norm or not fio:
                continue
            if mac_norm in latest_owner_by_mac:
                continue
            latest_owner_by_mac[mac_norm] = rec

        if not latest_owner_by_mac:
            return None

        fio_list: list[str] = []
        for rec in latest_owner_by_mac.values():
            fio = _s(rec.get("fio") or rec.get("OWNER_DISPLAY_NAME"))
            if fio and fio not in fio_list:
                fio_list.append(fio)
        if not fio_list:
            return None
        fio_count = len(fio_list)
        fio_value = fio_list[0] if fio_count == 1 else f"{fio_count} {_people_word(fio_count)}: {'; '.join(fio_list)}"
        first_row = next(iter(latest_owner_by_mac.values()))
        return {
            "fio": fio_value,
            "fio_count": fio_count,
            "fio_list": fio_list,
            "owner_no": first_row.get("owner_no"),
            "mac_address": first_row.get("mac_address"),
            "changed_at": first_row.get("changed_at"),
            "db_id": db_id,
        }

    def sync_socket_host_context(
        self,
        *,
        branch_id: int,
        socket_ids: Optional[list[int]],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        try:
            from backend.database import queries  # local import to avoid hard dependency in bot runtime
        except Exception as exc:  # pragma: no cover
            raise ValueError("Не удалось загрузить SQL-резолвер контекста ПК") from exc

        with self._lock, self._connect() as conn:
            mapping = conn.execute("SELECT db_id FROM network_branch_db_map WHERE branch_id=?", (int(branch_id),)).fetchone()
            if mapping is None or not _s(mapping["db_id"]):
                raise ValueError("Для филиала не настроена БД для синхронизации по MAC")
            db_id = _s(mapping["db_id"])
            where = ["branch_id=?"]
            params: list[Any] = [int(branch_id)]
            if socket_ids:
                normalized_ids = [int(x) for x in socket_ids if int(x) > 0]
                if normalized_ids:
                    placeholders = ",".join(["?"] * len(normalized_ids))
                    where.append(f"id IN ({placeholders})")
                    params.extend(normalized_ids)
            rows = conn.execute(
                f"""
                SELECT
                    s.*,
                    p.id as linked_port_id,
                    p.endpoint_mac_raw port_mac_raw,
                    p.endpoint_ip_raw port_ip_raw,
                    p.endpoint_name_raw port_name_raw
                FROM network_sockets s
                LEFT JOIN network_ports p ON p.id=s.port_id
                WHERE {' AND '.join(where)}
                """,
                params,
            ).fetchall()
            resolved = 0
            checked = 0
            not_found = 0
            skipped_no_mac = 0
            ports_updated = 0
            sockets_updated = 0
            now = _now()
            for row in rows:
                checked += 1
                final_macs: list[str] = []
                _append_unique(final_macs, _extract_mac_candidates(row["mac_address"]))
                _append_unique(final_macs, _extract_mac_candidates(row["port_mac_raw"]))
                if not final_macs:
                    skipped_no_mac += 1
                    continue

                ip_values: list[str] = []
                endpoint_names: list[str] = []
                fio_values: list[str] = []
                row_resolved = False

                for mac_candidate in list(final_macs):
                    payload: Optional[dict[str, Any]] = None
                    try:
                        payload = queries.resolve_pc_context_by_mac_or_hostname(
                            mac_address=mac_candidate,
                            hostname=None,
                            db_id=db_id,
                        )
                    except Exception:
                        payload = None
                    if not isinstance(payload, dict):
                        continue

                    row_resolved = True
                    _append_unique(
                        final_macs,
                        [_mac_normalized(payload.get("mac_address")) or mac_candidate],
                    )
                    _append_unique(ip_values, _extract_ip_candidates(payload.get("ip_address")))
                    _append_unique(endpoint_names, _extract_text_candidates(payload.get("network_name")))
                    _append_unique(fio_values, [_s(payload.get("employee_name"))])

                if row_resolved:
                    resolved += 1
                else:
                    not_found += 1

                mac_value = _join_lines(final_macs)
                ip_value = _join_lines(ip_values)
                endpoint_name_value = _join_lines(endpoint_names)
                if row_resolved and not ip_value:
                    ip_value = _s(row["port_ip_raw"]) or None
                if row_resolved and not endpoint_name_value:
                    endpoint_name_value = _s(row["port_name_raw"]) or None

                if len(fio_values) == 1:
                    fio_value = fio_values[0]
                elif len(fio_values) > 1:
                    fio_count = len(fio_values)
                    fio_value = f"{fio_count} {_people_word(fio_count)}: {'; '.join(fio_values)}"
                else:
                    fio_value = None

                linked_port_id = int(row["linked_port_id"]) if row["linked_port_id"] is not None else None
                if linked_port_id is not None:
                    if row_resolved:
                        conn.execute(
                            """
                            UPDATE network_ports
                            SET endpoint_name_raw=?, endpoint_ip_raw=?, endpoint_mac_raw=?, updated_at=?
                            WHERE id=?
                            """,
                            (
                                endpoint_name_value,
                                ip_value,
                                mac_value,
                                now,
                                linked_port_id,
                            ),
                        )
                    else:
                        conn.execute(
                            """
                            UPDATE network_ports
                            SET endpoint_mac_raw=?, updated_at=?
                            WHERE id=?
                            """,
                            (
                                mac_value,
                                now,
                                linked_port_id,
                            ),
                        )
                    ports_updated += 1

                if row_resolved:
                    conn.execute(
                        """
                        UPDATE network_sockets
                        SET mac_address=?, fio=?, fio_source_db=?, fio_resolved_at=?, updated_at=?
                        WHERE id=?
                        """,
                        (
                            mac_value,
                            fio_value,
                            db_id if fio_value else None,
                            now if fio_value else None,
                            now,
                            int(row["id"]),
                        ),
                    )
                else:
                    conn.execute(
                        """
                        UPDATE network_sockets
                        SET mac_address=?, updated_at=?
                        WHERE id=?
                        """,
                        (
                            mac_value,
                            now,
                            int(row["id"]),
                        ),
                    )
                sockets_updated += 1

            self._audit(
                conn,
                branch_id=int(branch_id),
                entity_type="socket",
                entity_id=None,
                action="host_context_sync",
                diff={
                    "mode": "mac",
                    "checked": checked,
                    "updated": resolved,
                    "resolved": resolved,
                    "not_found": not_found,
                    "skipped_no_mac": skipped_no_mac,
                    "ports_updated": ports_updated,
                    "sockets_updated": sockets_updated,
                    "db_id": db_id,
                },
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return {
                "branch_id": int(branch_id),
                "checked": int(checked),
                "updated": int(resolved),
                "resolved": int(resolved),  # backwards compatibility with old frontend
                "not_found": int(not_found),
                "skipped_no_mac": int(skipped_no_mac),
                "ports_updated": int(ports_updated),
                "sockets_updated": int(sockets_updated),
                "db_id": db_id,
            }

    def resolve_socket_fio(
        self,
        *,
        branch_id: int,
        socket_ids: Optional[list[int]],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        # Backward-compatibility alias for legacy clients.
        return self.sync_socket_host_context(
            branch_id=branch_id,
            socket_ids=socket_ids,
            actor_user_id=actor_user_id,
            actor_role=actor_role,
        )

    def import_sockets_template(
        self,
        *,
        branch_id: int,
        file_name: str,
        file_bytes: bytes,
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        if openpyxl is None:
            raise RuntimeError("openpyxl is not available")
        if not file_bytes:
            raise ValueError("Template file is empty")

        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

        created = 0
        updated = 0
        linked_ports = 0

        with self._lock, self._connect() as conn:
            branch = conn.execute("SELECT id FROM network_branches WHERE id=?", (int(branch_id),)).fetchone()
            if branch is None:
                raise ValueError("Branch not found")

            # Iterate over ALL sheets in the workbook
            for worksheet in workbook.worksheets:
                headers = {_h(worksheet.cell(1, column).value): column for column in range(1, worksheet.max_column + 1)}
                socket_col = headers.get("socket") or headers.get("socket code") or headers.get("port p p") or headers.get("розетка")
                if socket_col is None:
                    # Skip sheets without a socket/port column
                    continue

                mac_col = headers.get("mac") or headers.get("mac address") or headers.get("mac address")
                fio_col = headers.get("fio") or headers.get("owner") or headers.get("employee") or headers.get("employee name") or headers.get("фио") or headers.get("name")
                device_col = headers.get("asw") or headers.get("device") or headers.get("device code") or headers.get("switch") or headers.get("swich")
                port_col = headers.get("port") or headers.get("port name")

                for row_no in range(2, worksheet.max_row + 1):
                    socket_raw = _s(worksheet.cell(row_no, socket_col).value)
                    socket_code = self._canonical_socket_code(socket_raw)
                    if not socket_code:
                        continue
                    panel_no, port_no = _parse_socket_parts(socket_code)
                    mac_raw_text = _s(worksheet.cell(row_no, mac_col).value) if mac_col else ""
                    mac_candidates = _extract_mac_candidates(mac_raw_text)
                    mac_address = mac_raw_text or (mac_candidates[0] if mac_candidates else "")
                    fio = _s(worksheet.cell(row_no, fio_col).value) if fio_col else ""
                    device_code = _s(worksheet.cell(row_no, device_col).value) if device_col else ""
                    port_name = _s(worksheet.cell(row_no, port_col).value) if port_col else ""

                    port_row = None
                    if device_code and port_name:
                        port_row = conn.execute(
                            """
                            SELECT p.id, p.device_id, p.endpoint_mac_raw
                            FROM network_ports p
                            JOIN network_devices d ON d.id=p.device_id
                            WHERE d.branch_id=? AND d.device_code=? AND p.port_name=?
                            LIMIT 1
                            """,
                            (int(branch_id), device_code, port_name),
                        ).fetchone()

                    existing = conn.execute(
                        "SELECT * FROM network_sockets WHERE branch_id=? AND socket_code=?",
                        (int(branch_id), socket_code),
                    ).fetchone()
                    if existing is None:
                        conn.execute(
                            """
                            INSERT INTO network_sockets(
                                branch_id, socket_code, panel_no, port_no, port_id, device_id,
                                mac_address, fio, fio_source_db, fio_resolved_at, created_at, updated_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                int(branch_id),
                                socket_code,
                                panel_no,
                                port_no,
                                int(port_row["id"]) if port_row is not None else None,
                                int(port_row["device_id"]) if port_row is not None else None,
                                (mac_address or (_s(port_row["endpoint_mac_raw"]) if port_row is not None else "")) or None,
                                fio or None,
                                "template" if fio else None,
                                _now() if fio else None,
                                _now(),
                                _now(),
                            ),
                        )
                        created += 1
                        socket_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
                    else:
                        socket_id = int(existing["id"])
                        conn.execute(
                            """
                            UPDATE network_sockets
                            SET panel_no=?,
                                port_no=?,
                                port_id=COALESCE(?, port_id),
                                device_id=COALESCE(?, device_id),
                                mac_address=COALESCE(?, mac_address),
                                fio=COALESCE(?, fio),
                                fio_source_db=CASE WHEN ? IS NOT NULL AND ? <> '' THEN 'template' ELSE fio_source_db END,
                                fio_resolved_at=CASE WHEN ? IS NOT NULL AND ? <> '' THEN ? ELSE fio_resolved_at END,
                                updated_at=?
                            WHERE id=?
                            """,
                            (
                                panel_no,
                                port_no,
                                int(port_row["id"]) if port_row is not None else None,
                                int(port_row["device_id"]) if port_row is not None else None,
                                mac_address or (_s(port_row["endpoint_mac_raw"]) if port_row is not None else None),
                                fio or None,
                                fio or None,
                                fio or None,
                                fio or None,
                                fio or None,
                                _now(),
                                _now(),
                                socket_id,
                            ),
                        )
                        updated += 1

                    if port_row is not None:
                        linked_ports += 1
                        conn.execute(
                            "UPDATE network_map_points SET socket_id=?, updated_at=? WHERE port_id=? AND socket_id IS NULL",
                            (socket_id, _now(), int(port_row["id"])),
                        )

            self._sync_all_sockets_in_conn(conn, branch_id=int(branch_id))
            self._link_map_points_to_sockets_in_conn(conn, branch_id=int(branch_id))
            self._audit(
                conn,
                branch_id=int(branch_id),
                entity_type="socket_template",
                entity_id=str(branch_id),
                action="import_merge",
                diff={
                    "file_name": _s(file_name),
                    "created": int(created),
                    "updated": int(updated),
                    "linked_ports": int(linked_ports),
                },
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return {
                "branch_id": int(branch_id),
                "file_name": _s(file_name),
                "created": int(created),
                "updated": int(updated),
                "linked_ports": int(linked_ports),
            }


    def _repair_map_point_socket_conflicts_in_conn(
        self,
        conn: sqlite3.Connection,
        *,
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, int]:
        rows = conn.execute(
            """
            SELECT
                mp.id,
                mp.branch_id,
                mp.map_id,
                mp.port_id,
                mp.socket_id,
                mp.device_id,
                mp.label,
                mp.note,
                mp.color,
                mp.created_at,
                mp.updated_at,
                COALESCE(s.socket_code, p.patch_panel_port) patch_panel_port
            FROM network_map_points mp
            LEFT JOIN network_ports p ON p.id=mp.port_id
            LEFT JOIN network_sockets s ON s.id=mp.socket_id
            ORDER BY mp.branch_id, mp.id
            """
        ).fetchall()

        groups: dict[tuple[int, str], list[sqlite3.Row]] = {}
        orphans = 0
        for row in rows:
            if row["socket_id"] is not None:
                socket_norm = f"id:{int(row['socket_id'])}"
            else:
                socket_norm = _socket_key(row["patch_panel_port"])
            if not socket_norm:
                orphans += 1
                continue
            key = (int(row["branch_id"]), socket_norm)
            groups.setdefault(key, []).append(row)

        merged_groups = 0
        removed_points = 0
        total_conflicts = 0

        for (branch_id, socket_norm), group_rows in groups.items():
            if len(group_rows) <= 1:
                continue
            merged_groups += 1
            total_conflicts += len(group_rows) - 1
            ordered = sorted(
                group_rows,
                key=lambda row: (
                    _s(row["updated_at"]),
                    _s(row["created_at"]),
                    int(row["id"]),
                ),
                reverse=True,
            )
            keeper_id = int(ordered[0]["id"])
            for duplicate in ordered[1:]:
                duplicate_id = int(duplicate["id"])
                keeper_before = conn.execute("SELECT * FROM network_map_points WHERE id=?", (keeper_id,)).fetchone()
                duplicate_before = conn.execute("SELECT * FROM network_map_points WHERE id=?", (duplicate_id,)).fetchone()
                if keeper_before is None or duplicate_before is None:
                    continue

                # Preserve meaningful metadata from duplicate if keeper has empty values.
                patch_fields: list[str] = []
                patch_params: list[Any] = []
                duplicate_port_id = duplicate_before["port_id"]
                keeper_port_id = keeper_before["port_id"]
                duplicate_socket_id = duplicate_before["socket_id"] if "socket_id" in duplicate_before.keys() else None
                keeper_socket_id = keeper_before["socket_id"] if "socket_id" in keeper_before.keys() else None
                if (keeper_socket_id is None or int(keeper_socket_id or 0) == 0) and duplicate_socket_id is not None:
                    patch_fields.append("socket_id=?")
                    patch_params.append(duplicate_socket_id)
                if (keeper_port_id is None or int(keeper_port_id or 0) == 0) and duplicate_port_id is not None:
                    patch_fields.append("port_id=?")
                    patch_params.append(duplicate_port_id)
                    if keeper_before["device_id"] is None and duplicate_before["device_id"] is not None:
                        patch_fields.append("device_id=?")
                        patch_params.append(duplicate_before["device_id"])
                if not _s(keeper_before["label"]) and _s(duplicate_before["label"]):
                    patch_fields.append("label=?")
                    patch_params.append(_s(duplicate_before["label"]))
                if not _s(keeper_before["note"]) and _s(duplicate_before["note"]):
                    patch_fields.append("note=?")
                    patch_params.append(_s(duplicate_before["note"]))
                if not _s(keeper_before["color"]) and _s(duplicate_before["color"]):
                    patch_fields.append("color=?")
                    patch_params.append(_s(duplicate_before["color"]))

                if patch_fields:
                    patch_fields.append("updated_at=?")
                    patch_params.append(_now())
                    patch_params.append(keeper_id)
                    conn.execute(
                        f"UPDATE network_map_points SET {', '.join(patch_fields)} WHERE id=?",
                        patch_params,
                    )

                conn.execute("DELETE FROM network_map_points WHERE id=?", (duplicate_id,))
                removed_points += 1

                keeper_after = conn.execute("SELECT * FROM network_map_points WHERE id=?", (keeper_id,)).fetchone()
                self._audit(
                    conn,
                    branch_id=int(keeper_before["branch_id"]),
                    entity_type="map_point",
                    entity_id=str(keeper_id),
                    action="socket_conflict_resolved",
                    diff={
                        "branch_id": branch_id,
                        "socket_key": socket_norm,
                        "socket": _s(duplicate["patch_panel_port"]),
                        "kept_point_id": keeper_id,
                        "removed_point_id": duplicate_id,
                        "keeper_before": self._d(keeper_before),
                        "keeper_after": self._d(keeper_after),
                        "removed_before": self._d(duplicate_before),
                    },
                    actor_user_id=actor_user_id,
                    actor_role=actor_role,
                )

        return {
            "groups_with_conflicts": int(merged_groups),
            "conflicts_found": int(total_conflicts),
            "removed_points": int(removed_points),
            "orphan_points_without_socket": int(orphans),
        }

    def repair_map_point_socket_conflicts(
        self,
        *,
        actor_user_id: Optional[int] = None,
        actor_role: Optional[str] = "system_maintenance",
    ) -> dict[str, int]:
        with self._lock, self._connect() as conn:
            summary = self._repair_map_point_socket_conflicts_in_conn(
                conn,
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return summary

    def ensure_branch(self, *, city_code: str, branch_code: str, name: str) -> dict[str, Any]:
        now = _now()
        with self._lock, self._connect() as conn:
            conn.execute(
                """
                INSERT INTO network_branches(city_code, branch_code, name, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(city_code, branch_code) DO UPDATE SET name=excluded.name, updated_at=excluded.updated_at
                """,
                (_s(city_code), _s(branch_code), _s(name), now, now),
            )
            row = conn.execute(
                "SELECT * FROM network_branches WHERE city_code=? AND branch_code=?",
                (_s(city_code), _s(branch_code)),
            ).fetchone()
            conn.commit()
            return self._d(row) or {}

    def create_branch_with_profile(
        self,
        *,
        city_code: str,
        branch_code: Optional[str] = None,
        name: str,
        panel_count: Optional[int] = None,
        ports_per_panel: Optional[int] = None,
        panels: Optional[list[dict[str, int]]] = None,
        default_site_code: Optional[str] = None,
        db_id: Optional[str] = None,
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        now = _now()

        has_profile = False
        is_uniform = panels is None or len(panels) == 0

        if is_uniform:
            if panel_count is not None and ports_per_panel is not None and int(panel_count) > 0 and int(ports_per_panel) > 0:
                has_profile = True
        else:
            if not panels or len(panels) == 0:
                raise ValueError("panels list cannot be empty for heterogeneous mode")
            for panel in panels:
                panel_index = int(panel.get("panel_index", 0))
                port_count = int(panel.get("port_count", 0))
                if panel_index <= 0 or port_count <= 0:
                    raise ValueError("panel_index and port_count must be positive")

            panel_count = len(panels)
            ports_per_panel = max((int(p.get("port_count", 0)) for p in panels), default=1)
            has_profile = True

        with self._lock, self._connect() as conn:
            # Auto-generate branch_code if not provided
            final_branch_code = _s(branch_code)
            if not final_branch_code:
                final_branch_code = _generate_branch_code(name, city_code, conn)

            conn.execute(
                """
                INSERT INTO network_branches(city_code, branch_code, name, default_site_code, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(city_code, branch_code) DO UPDATE SET name=excluded.name, default_site_code=excluded.default_site_code, updated_at=excluded.updated_at
                """,
                (_s(city_code), _s(final_branch_code), _s(name), _s(default_site_code) or None, now, now),
            )
            branch = conn.execute(
                "SELECT * FROM network_branches WHERE city_code=? AND branch_code=?",
                (_s(city_code), _s(final_branch_code)),
            ).fetchone()
            if branch is None:
                raise ValueError("Failed to create branch")
            branch_id = int(branch["id"])

            # Create db_mapping if db_id is provided
            if db_id:
                conn.execute(
                    """
                    INSERT INTO network_branch_db_map(branch_id, db_id, updated_at)
                    VALUES (?, ?, ?)
                    ON CONFLICT(branch_id) DO UPDATE SET db_id=excluded.db_id, updated_at=excluded.updated_at
                    """,
                    (branch_id, _s(db_id), now),
                )

            if has_profile:
                # Upsert socket profile
                self._upsert_socket_profile_in_conn(
                    conn,
                    branch_id=branch_id,
                    panel_count=int(panel_count),
                    ports_per_panel=int(ports_per_panel),
                    is_uniform=is_uniform,
                )

                # Upsert individual panels for heterogeneous mode
                if not is_uniform:
                    self._upsert_panels_in_conn(
                        conn,
                        branch_id=branch_id,
                        panels=panels,
                    )

            summary = self._bootstrap_sockets_for_branch_in_conn(conn, branch_id=branch_id)
            self._audit(
                conn,
                branch_id=branch_id,
                entity_type="branch",
                entity_id=str(branch_id),
                action="create_or_update_with_profile",
                diff={"branch": self._d(branch), "socket_summary": summary, "is_uniform": is_uniform},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            profile = conn.execute("SELECT * FROM network_socket_profiles WHERE branch_id=?", (branch_id,)).fetchone()
            return {"branch": self._d(branch) or {}, "profile": self._d(profile) or {}, "socket_summary": summary}

    def _ensure_site(self, conn: sqlite3.Connection, *, branch_id: int, site_code: str, site_name: str) -> Optional[int]:
        code = _s(site_code).lower()
        if not code:
            return None
        now = _now()
        conn.execute(
            """
            INSERT INTO network_sites(branch_id, site_code, name, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(branch_id, site_code) DO UPDATE SET name=excluded.name, updated_at=excluded.updated_at
            """,
            (int(branch_id), code, _s(site_name) or code.upper(), now, now),
        )
        row = conn.execute("SELECT id FROM network_sites WHERE branch_id=? AND site_code=?", (int(branch_id), code)).fetchone()
        return int(row["id"]) if row else None

    def list_branches(self, city_code: str) -> list[dict[str, Any]]:
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                """
                SELECT b.*,
                    m.db_id,
                    (SELECT COUNT(*) FROM network_devices d WHERE d.branch_id=b.id) devices_count,
                    (SELECT COUNT(*) FROM network_ports p JOIN network_devices d ON d.id=p.device_id WHERE d.branch_id=b.id) ports_count,
                    (SELECT COUNT(*) FROM network_ports p JOIN network_devices d ON d.id=p.device_id WHERE d.branch_id=b.id AND p.is_occupied=1) occupied_ports,
                    (SELECT COUNT(*) FROM network_sockets s WHERE s.branch_id=b.id) sockets_count,
                    (SELECT COUNT(*) FROM network_maps m2 WHERE m2.branch_id=b.id) maps_count,
                    (SELECT COUNT(*) FROM network_map_points mp WHERE mp.branch_id=b.id) map_points_count
                FROM network_branches b
                LEFT JOIN network_branch_db_map m ON b.id = m.branch_id
                WHERE b.city_code=? AND b.is_active=1
                ORDER BY b.name COLLATE NOCASE
                """,
                (_s(city_code),),
            ).fetchall()
            result = [self._d(r) for r in rows if r is not None]
            if result:
                return result
        if _s(city_code).lower() == "tmn":
            self.ensure_branch(city_code="tmn", branch_code="tmn-p19-21", name="Первомайская 19/21")
            with self._lock, self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT b.*,
                        m.db_id,
                        (SELECT COUNT(*) FROM network_devices d WHERE d.branch_id=b.id) devices_count,
                        (SELECT COUNT(*) FROM network_ports p JOIN network_devices d ON d.id=p.device_id WHERE d.branch_id=b.id) ports_count,
                        (SELECT COUNT(*) FROM network_ports p JOIN network_devices d ON d.id=p.device_id WHERE d.branch_id=b.id AND p.is_occupied=1) occupied_ports,
                        (SELECT COUNT(*) FROM network_sockets s WHERE s.branch_id=b.id) sockets_count,
                        (SELECT COUNT(*) FROM network_maps m2 WHERE m2.branch_id=b.id) maps_count,
                        (SELECT COUNT(*) FROM network_map_points mp WHERE mp.branch_id=b.id) map_points_count
                    FROM network_branches b
                    LEFT JOIN network_branch_db_map m ON b.id = m.branch_id
                    WHERE b.city_code=? AND b.is_active=1
                    ORDER BY b.name COLLATE NOCASE
                    """,
                    (_s(city_code),),
                ).fetchall()
                return [self._d(r) for r in rows if r is not None]
        return []

    def get_branches_with_db_mapping(self, conn: sqlite3.Connection) -> list[dict[str, Any]]:
        """Get all branches with their db_mapping."""
        rows = conn.execute(
            """
            SELECT b.*, m.db_id
            FROM network_branches b
            LEFT JOIN network_branch_db_map m ON b.id = m.branch_id
            WHERE b.is_active=1
            ORDER BY b.city_code, b.name COLLATE NOCASE
            """
        ).fetchall()
        return [self._d(r) for r in rows]

    def delete_branch(
        self,
        *,
        branch_id: int,
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> bool:
        """Delete a branch and all related data (CASCADE)."""
        with self._lock, self._connect() as conn:
            branch = conn.execute("SELECT * FROM network_branches WHERE id=?", (int(branch_id),)).fetchone()
            if branch is None:
                raise ValueError("Branch not found")

            # Log audit before deletion
            self._audit(
                conn,
                branch_id=int(branch_id),
                entity_type="branch",
                entity_id=str(branch_id),
                action="delete",
                diff={"branch": self._d(branch)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )

            # Delete branch (CASCADE will handle related records)
            conn.execute("DELETE FROM network_branches WHERE id=?", (int(branch_id),))
            conn.commit()
            return True

    def get_branch_overview(self, branch_id: int) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            branch = conn.execute("SELECT * FROM network_branches WHERE id=?", (int(branch_id),)).fetchone()
            if branch is None:
                raise ValueError("Branch not found")
            metrics = conn.execute(
                """
                SELECT
                    (SELECT COUNT(*) FROM network_devices WHERE branch_id=?) devices_count,
                    (SELECT COUNT(*) FROM network_ports p JOIN network_devices d ON d.id=p.device_id WHERE d.branch_id=?) ports_count,
                    (SELECT COUNT(*) FROM network_ports p JOIN network_devices d ON d.id=p.device_id WHERE d.branch_id=? AND p.is_occupied=1) occupied_ports,
                    (SELECT COUNT(*) FROM network_sockets WHERE branch_id=?) sockets_count,
                    (SELECT COUNT(*) FROM network_maps WHERE branch_id=?) maps_count,
                    (SELECT COUNT(*) FROM network_map_points WHERE branch_id=?) map_points_count
                """,
                (int(branch_id), int(branch_id), int(branch_id), int(branch_id), int(branch_id), int(branch_id)),
            ).fetchone()
            sites = conn.execute(
                """
                SELECT s.*,
                    (SELECT COUNT(*) FROM network_devices d WHERE d.site_id=s.id) devices_count,
                    (SELECT COUNT(*) FROM network_ports p JOIN network_devices d ON d.id=p.device_id WHERE d.site_id=s.id) ports_count
                FROM network_sites s
                WHERE s.branch_id=?
                ORDER BY s.sort_order, s.name COLLATE NOCASE
                """,
                (int(branch_id),),
            ).fetchall()
            return {"branch": self._d(branch), "metrics": self._d(metrics) or {}, "sites": [self._d(r) for r in sites if r is not None]}

    def list_devices(self, branch_id: int) -> list[dict[str, Any]]:
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                """
                SELECT d.*, s.site_code, s.name site_name,
                    (SELECT COUNT(*) FROM network_ports p WHERE p.device_id=d.id) ports_count,
                    (SELECT COUNT(*) FROM network_ports p WHERE p.device_id=d.id AND p.is_occupied=1) occupied_ports
                FROM network_devices d
                LEFT JOIN network_sites s ON s.id=d.site_id
                WHERE d.branch_id=?
                ORDER BY COALESCE(s.sort_order,9999), d.device_code COLLATE NOCASE
                """,
                (int(branch_id),),
            ).fetchall()
            return [self._d(r) for r in rows if r is not None]

    def list_ports(self, device_id: int, *, search: str = "", vlan: str = "", occupied: Optional[bool] = None, location: str = "") -> list[dict[str, Any]]:
        where = ["p.device_id=?"]
        params: list[Any] = [int(device_id)]
        if _s(search):
            where.append(
                "("
                "p.port_name LIKE ? OR "
                "COALESCE(ns.socket_code, p.patch_panel_port, '') LIKE ? OR "
                "COALESCE(p.endpoint_name_raw,'') LIKE ? OR "
                "COALESCE(p.endpoint_ip_raw,'') LIKE ? OR "
                "COALESCE(p.endpoint_mac_raw,'') LIKE ? OR "
                "COALESCE(s.fio,'') LIKE ?"
                ")"
            )
            v = f"%{_s(search)}%"
            params.extend([v, v, v, v, v, v])
        if _s(vlan):
            where.append("COALESCE(p.vlan_raw,'') LIKE ?")
            params.append(f"%{_s(vlan)}%")
        if _s(location):
            where.append("COALESCE(p.location_code,'') LIKE ?")
            params.append(f"%{_s(location)}%")
        if occupied is not None:
            where.append("p.is_occupied=?")
            params.append(1 if occupied else 0)
        q = f"""
            SELECT
                p.*,
                s.id socket_id,
                s.socket_code,
                s.fio,
                s.fio_source_db
            FROM network_ports p
            LEFT JOIN network_sockets s ON s.port_id=p.id
            WHERE {' AND '.join(where)}
            ORDER BY p.port_name COLLATE NOCASE
        """
        with self._lock, self._connect() as conn:
            rows = conn.execute(q, params).fetchall()
            return [self._d(r) for r in rows if r is not None]

    def list_ports_by_branch(
        self,
        branch_id: int,
        *,
        search: str = "",
        vlan: str = "",
        occupied: Optional[bool] = None,
        location: str = "",
        limit: int = 5000,
    ) -> list[dict[str, Any]]:
        where = ["d.branch_id=?"]
        params: list[Any] = [int(branch_id)]
        if _s(search):
            where.append(
                "("
                "COALESCE(d.device_code,'') LIKE ? OR "
                "COALESCE(d.model,'') LIKE ? OR "
                "COALESCE(p.port_name,'') LIKE ? OR "
                "COALESCE(ns.socket_code, p.patch_panel_port, '') LIKE ? OR "
                "COALESCE(p.location_code,'') LIKE ? OR "
                "COALESCE(p.vlan_raw,'') LIKE ? OR "
                "COALESCE(p.endpoint_name_raw,'') LIKE ? OR "
                "COALESCE(p.endpoint_ip_raw,'') LIKE ? OR "
                "COALESCE(p.endpoint_mac_raw,'') LIKE ? OR "
                "COALESCE(ns.fio,'') LIKE ?"
                ")"
            )
            v = f"%{_s(search)}%"
            params.extend([v, v, v, v, v, v, v, v, v, v])
        if _s(vlan):
            where.append("COALESCE(p.vlan_raw,'') LIKE ?")
            params.append(f"%{_s(vlan)}%")
        if _s(location):
            where.append("COALESCE(p.location_code,'') LIKE ?")
            params.append(f"%{_s(location)}%")
        if occupied is not None:
            where.append("p.is_occupied=?")
            params.append(1 if occupied else 0)

        q = f"""
            SELECT
                p.*,
                d.device_code,
                d.model device_model,
                d.branch_id,
                s.site_code,
                s.name site_name,
                ns.id socket_id,
                ns.socket_code,
                ns.fio,
                ns.fio_source_db
            FROM network_ports p
            JOIN network_devices d ON d.id=p.device_id
            LEFT JOIN network_sites s ON s.id=d.site_id
            LEFT JOIN network_sockets ns ON ns.port_id=p.id
            WHERE {' AND '.join(where)}
            ORDER BY d.device_code COLLATE NOCASE, p.port_name COLLATE NOCASE
            LIMIT ?
        """
        params.append(int(limit))
        with self._lock, self._connect() as conn:
            rows = conn.execute(q, params).fetchall()
            return [self._d(r) for r in rows if r is not None]

    def list_maps(self, branch_id: int) -> list[dict[str, Any]]:
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                """
                SELECT m.id,m.branch_id,m.site_id,s.site_code,s.name site_name,m.title,m.floor_label,m.file_name,m.mime_type,m.file_size,m.checksum_sha256,m.source_path,m.created_at,m.updated_at
                FROM network_maps m
                LEFT JOIN network_sites s ON s.id=m.site_id
                WHERE m.branch_id=?
                ORDER BY COALESCE(s.sort_order,9999), COALESCE(m.floor_label,''), m.file_name COLLATE NOCASE
                """,
                (int(branch_id),),
            ).fetchall()
            return [self._d(r) for r in rows if r is not None]

    @staticmethod
    def _validate_port_for_map_point(
        conn: sqlite3.Connection,
        *,
        branch_id: int,
        port_id: int,
        expected_device_id: Optional[int] = None,
    ) -> sqlite3.Row:
        port_row = conn.execute(
            """
            SELECT
                p.id,
                p.device_id,
                p.port_name,
                p.patch_panel_port,
                p.location_code,
                p.endpoint_name_raw,
                p.endpoint_ip_raw,
                p.endpoint_mac_raw,
                d.branch_id
            FROM network_ports p
            JOIN network_devices d ON d.id=p.device_id
            WHERE p.id=?
            """,
            (int(port_id),),
        ).fetchone()
        if port_row is None:
            raise ValueError("Port not found")
        if int(port_row["branch_id"]) != int(branch_id):
            raise ValueError("Port does not belong to branch")
        if expected_device_id is not None and int(port_row["device_id"]) != int(expected_device_id):
            raise ValueError("Port does not belong to selected device")
        if not _s(port_row["patch_panel_port"]):
            raise ValueError("У выбранного порта не заполнено поле PORT P/P")
        return port_row

    @staticmethod
    def _find_point_conflict_by_socket(
        conn: sqlite3.Connection,
        *,
        branch_id: int,
        socket_id: Optional[int] = None,
        socket_value: str = "",
        exclude_point_id: Optional[int] = None,
    ) -> Optional[sqlite3.Row]:
        socket_norm = _socket_key(socket_value)
        rows = conn.execute(
            """
            SELECT
                mp.id,
                mp.branch_id,
                mp.map_id,
                mp.port_id,
                mp.socket_id,
                COALESCE(s.socket_code, p.patch_panel_port) patch_panel_port
            FROM network_map_points mp
            LEFT JOIN network_ports p ON p.id=mp.port_id
            LEFT JOIN network_sockets s ON s.id=mp.socket_id
            WHERE mp.branch_id=?
            ORDER BY mp.id
            """,
            (int(branch_id),),
        ).fetchall()
        for row in rows:
            if exclude_point_id is not None and int(row["id"]) == int(exclude_point_id):
                continue
            if socket_id is not None and row["socket_id"] is not None and int(row["socket_id"]) == int(socket_id):
                return row
            if socket_norm and _socket_key(row["patch_panel_port"]) == socket_norm:
                return row
        return None

    def list_map_points(self, *, branch_id: int, map_id: Optional[int] = None) -> list[dict[str, Any]]:
        with self._lock, self._connect() as conn:
            params: list[Any] = [int(branch_id)]
            where = ["mp.branch_id=?"]
            if map_id is not None:
                where.append("mp.map_id=?")
                params.append(int(map_id))
            rows = conn.execute(
                f"""
                SELECT
                    mp.*,
                    m.title map_title,
                    m.file_name map_file_name,
                    d.device_code,
                    d.model device_model,
                    p.port_name,
                    COALESCE(s.socket_code, p.patch_panel_port) patch_panel_port,
                    s.socket_code,
                    p.location_code port_location_code,
                    p.vlan_raw port_vlan_raw,
                    p.endpoint_name_raw,
                    p.endpoint_ip_raw,
                    COALESCE(s.mac_address, p.endpoint_mac_raw) endpoint_mac_raw,
                    s.fio,
                    s.fio_source_db
                FROM network_map_points mp
                LEFT JOIN network_maps m ON m.id=mp.map_id
                LEFT JOIN network_devices d ON d.id=mp.device_id
                LEFT JOIN network_ports p ON p.id=mp.port_id
                LEFT JOIN network_sockets s ON s.id=mp.socket_id
                WHERE {' AND '.join(where)}
                ORDER BY mp.map_id, COALESCE(s.socket_code, p.patch_panel_port,''), COALESCE(p.port_name,''), mp.id
                """,
                params,
            ).fetchall()
            return [self._map_point_public_dict(r) for r in rows if r is not None]

    def create_map_point(self, *, payload: dict[str, Any], actor_user_id: Optional[int], actor_role: Optional[str]) -> dict[str, Any]:
        branch_id = int(payload.get("branch_id") or 0)
        map_id = int(payload.get("map_id") or 0)
        x_ratio = float(payload.get("x_ratio") or 0.0)
        y_ratio = float(payload.get("y_ratio") or 0.0)
        if branch_id <= 0 or map_id <= 0:
            raise ValueError("branch_id and map_id are required")
        if x_ratio < 0 or x_ratio > 1 or y_ratio < 0 or y_ratio > 1:
            raise ValueError("x_ratio and y_ratio must be in range [0..1]")
        device_id = int(payload.get("device_id") or 0) or None
        port_id = int(payload.get("port_id") or 0) or None
        socket_id = int(payload.get("socket_id") or 0) or None
        if port_id is None and socket_id is None:
            raise ValueError("Для точки на карте необходимо выбрать розетку или порт")
        now = _now()
        with self._lock, self._connect() as conn:
            map_row = conn.execute("SELECT id,branch_id,site_id FROM network_maps WHERE id=?", (map_id,)).fetchone()
            if map_row is None:
                raise ValueError("Map not found")
            if int(map_row["branch_id"]) != branch_id:
                raise ValueError("Map does not belong to branch")
            if conn.execute("SELECT 1 FROM network_branches WHERE id=?", (branch_id,)).fetchone() is None:
                raise ValueError("Branch not found")
            if device_id is not None:
                device_row = conn.execute("SELECT id,branch_id FROM network_devices WHERE id=?", (int(device_id),)).fetchone()
                if device_row is None:
                    raise ValueError("Device not found")
                if int(device_row["branch_id"]) != branch_id:
                    raise ValueError("Device does not belong to branch")

            port_row = None
            socket_row = None
            if socket_id is not None:
                socket_row = conn.execute(
                    "SELECT * FROM network_sockets WHERE id=? AND branch_id=?",
                    (int(socket_id), int(branch_id)),
                ).fetchone()
                if socket_row is None:
                    raise ValueError("Socket not found")
                if port_id is None and socket_row["port_id"] is not None:
                    port_id = int(socket_row["port_id"])
                if device_id is None and socket_row["device_id"] is not None:
                    device_id = int(socket_row["device_id"])

            if port_id is not None:
                port_row = self._validate_port_for_map_point(
                    conn,
                    branch_id=branch_id,
                    port_id=port_id,
                    expected_device_id=device_id,
                )
                if device_id is None:
                    device_id = int(port_row["device_id"])
                socket_id = self._sync_socket_for_port_in_conn(conn, port_id=int(port_id))
                if socket_id is not None:
                    socket_row = conn.execute("SELECT * FROM network_sockets WHERE id=?", (int(socket_id),)).fetchone()

            socket_value = _s(socket_row["socket_code"] if socket_row is not None else (port_row["patch_panel_port"] if port_row is not None else ""))
            label_default = _point_label_from_port(
                port_row["port_name"] if port_row is not None else "",
                socket_value,
            )
            label_payload = _s(payload.get("label"))
            note_payload = _s(payload.get("note")) or None
            color_payload = _s(payload.get("color")) or None

            conflict_row = self._find_point_conflict_by_socket(
                conn,
                branch_id=branch_id,
                socket_id=socket_id,
                socket_value=socket_value,
            )
            if conflict_row is not None:
                conflict_id = int(conflict_row["id"])
                before_conflict = conn.execute("SELECT * FROM network_map_points WHERE id=?", (conflict_id,)).fetchone()
                conn.execute(
                    """
                    UPDATE network_map_points
                    SET map_id=?, site_id=?, device_id=?, port_id=?, socket_id=?, x_ratio=?, y_ratio=?, label=?, note=?, color=?, updated_at=?
                    WHERE id=?
                    """,
                    (
                        map_id,
                        map_row["site_id"],
                        device_id,
                        int(port_id),
                        int(socket_id) if socket_id is not None else None,
                        x_ratio,
                        y_ratio,
                        label_payload or _s(before_conflict["label"]) or label_default,
                        note_payload if note_payload is not None else before_conflict["note"],
                        color_payload if color_payload is not None else before_conflict["color"],
                        now,
                        conflict_id,
                    ),
                )
                row = conn.execute("SELECT * FROM network_map_points WHERE id=?", (conflict_id,)).fetchone()
                self._audit(
                    conn,
                    branch_id=branch_id,
                    entity_type="map_point",
                    entity_id=str(conflict_id),
                    action="socket_rebind",
                    diff={
                        "before": self._d(before_conflict),
                        "after": self._d(row),
                        "socket": socket_value,
                        "socket_key": _socket_key(socket_value),
                        "reason": "branch_socket_rebind",
                    },
                    actor_user_id=actor_user_id,
                    actor_role=actor_role,
                )
                conn.commit()
                return self._map_point_public_dict(self._map_point_row_with_links(conn, point_id=conflict_id))

            conn.execute(
                """
                INSERT INTO network_map_points(
                    branch_id,map_id,site_id,device_id,port_id,socket_id,x_ratio,y_ratio,label,note,color,created_at,updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    branch_id,
                    map_id,
                    map_row["site_id"],
                    device_id,
                    int(port_id) if port_id is not None else None,
                    int(socket_id) if socket_id is not None else None,
                    x_ratio,
                    y_ratio,
                    label_payload or label_default,
                    note_payload,
                    color_payload,
                    now,
                    now,
                ),
            )
            point_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            row = conn.execute("SELECT * FROM network_map_points WHERE id=?", (point_id,)).fetchone()
            self._audit(
                conn,
                branch_id=branch_id,
                entity_type="map_point",
                entity_id=str(point_id),
                action="socket_bind",
                diff={"after": self._d(row), "socket": socket_value, "socket_key": _socket_key(socket_value)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._map_point_public_dict(self._map_point_row_with_links(conn, point_id=point_id))

    def update_map_point(
        self,
        *,
        point_id: int,
        payload: dict[str, Any],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        allowed_fields = {"map_id", "device_id", "port_id", "socket_id", "x_ratio", "y_ratio", "label", "note", "color"}
        patch = {k: v for k, v in payload.items() if k in allowed_fields}
        if not patch:
            raise ValueError("No fields to update")
        with self._lock, self._connect() as conn:
            before = conn.execute("SELECT * FROM network_map_points WHERE id=?", (int(point_id),)).fetchone()
            if before is None:
                raise ValueError("Map point not found")

            branch_id = int(before["branch_id"])
            map_id = int(patch.get("map_id") or before["map_id"])
            device_id = patch.get("device_id", before["device_id"])
            port_id = patch.get("port_id", before["port_id"])
            socket_id = patch.get("socket_id", before["socket_id"])

            if device_id in {"", 0, "0"}:
                device_id = None
            if "port_id" in patch and port_id in {"", 0, "0", None}:
                port_id = None
            if port_id in {"", 0, "0"}:
                port_id = None
            if socket_id in {"", 0, "0"}:
                socket_id = None
            if device_id is not None:
                device_id = int(device_id)
            if port_id is not None:
                port_id = int(port_id)
            if socket_id is not None:
                socket_id = int(socket_id)

            map_row = conn.execute("SELECT id,branch_id,site_id FROM network_maps WHERE id=?", (map_id,)).fetchone()
            if map_row is None:
                raise ValueError("Map not found")
            if int(map_row["branch_id"]) != branch_id:
                raise ValueError("Map does not belong to branch")
            if device_id is not None:
                device_row = conn.execute("SELECT id,branch_id FROM network_devices WHERE id=?", (device_id,)).fetchone()
                if device_row is None:
                    raise ValueError("Device not found")
                if int(device_row["branch_id"]) != branch_id:
                    raise ValueError("Device does not belong to branch")

            socket_row = None
            socket_value = ""
            if socket_id is not None:
                socket_row = conn.execute(
                    "SELECT * FROM network_sockets WHERE id=? AND branch_id=?",
                    (int(socket_id), int(branch_id)),
                ).fetchone()
                if socket_row is None:
                    raise ValueError("Socket not found")
                socket_value = _s(socket_row["socket_code"])
                if port_id is None and socket_row["port_id"] is not None:
                    port_id = int(socket_row["port_id"])
                if device_id is None and socket_row["device_id"] is not None:
                    device_id = int(socket_row["device_id"])

            if port_id is not None:
                port_row = self._validate_port_for_map_point(
                    conn,
                    branch_id=branch_id,
                    port_id=port_id,
                    expected_device_id=device_id,
                )
                if device_id is None:
                    device_id = int(port_row["device_id"])
                    patch["device_id"] = device_id
                socket_id = self._sync_socket_for_port_in_conn(conn, port_id=port_id)
                if socket_id is not None:
                    socket_row = conn.execute("SELECT * FROM network_sockets WHERE id=?", (int(socket_id),)).fetchone()
                    socket_value = _s(socket_row["socket_code"])
                else:
                    socket_value = _s(port_row["patch_panel_port"])
                conflict_row = self._find_point_conflict_by_socket(
                    conn,
                    branch_id=branch_id,
                    socket_id=socket_id,
                    socket_value=socket_value,
                    exclude_point_id=int(point_id),
                )
                if conflict_row is not None:
                    raise NetworkConflictError(
                        self._point_conflict_detail(
                            map_id=int(conflict_row["map_id"]),
                            point_id=int(conflict_row["id"]),
                            socket_value=socket_value,
                        )
                    )
            elif socket_id is not None:
                conflict_row = self._find_point_conflict_by_socket(
                    conn,
                    branch_id=branch_id,
                    socket_id=socket_id,
                    socket_value=socket_value,
                    exclude_point_id=int(point_id),
                )
                if conflict_row is not None:
                    raise NetworkConflictError(
                        self._point_conflict_detail(
                            map_id=int(conflict_row["map_id"]),
                            point_id=int(conflict_row["id"]),
                            socket_value=socket_value or _s(socket_row["socket_code"] if socket_row is not None else ""),
                        )
                    )

            fields: list[str] = []
            params: list[Any] = []
            if "map_id" in patch:
                fields.append("map_id=?")
                params.append(map_id)
                fields.append("site_id=?")
                params.append(map_row["site_id"])
            if "device_id" in patch:
                fields.append("device_id=?")
                params.append(device_id)
            if "port_id" in patch:
                fields.append("port_id=?")
                params.append(port_id)
            if "socket_id" in patch or port_id is not None:
                fields.append("socket_id=?")
                params.append(socket_id)
            if "x_ratio" in patch:
                x_ratio = float(patch.get("x_ratio") or 0.0)
                if x_ratio < 0 or x_ratio > 1:
                    raise ValueError("x_ratio must be in range [0..1]")
                fields.append("x_ratio=?")
                params.append(x_ratio)
            if "y_ratio" in patch:
                y_ratio = float(patch.get("y_ratio") or 0.0)
                if y_ratio < 0 or y_ratio > 1:
                    raise ValueError("y_ratio must be in range [0..1]")
                fields.append("y_ratio=?")
                params.append(y_ratio)
            if "label" in patch:
                fields.append("label=?")
                params.append(_s(patch.get("label")) or None)
            if "note" in patch:
                fields.append("note=?")
                params.append(_s(patch.get("note")) or None)
            if "color" in patch:
                fields.append("color=?")
                params.append(_s(patch.get("color")) or None)

            fields.append("updated_at=?")
            params.append(_now())
            params.append(int(point_id))
            conn.execute(f"UPDATE network_map_points SET {', '.join(fields)} WHERE id=?", params)
            after = conn.execute("SELECT * FROM network_map_points WHERE id=?", (int(point_id),)).fetchone()
            after_with_links = self._map_point_row_with_links(conn, point_id=int(point_id))
            socket_after = _s(after_with_links["patch_panel_port"]) if after_with_links is not None else ""
            audit_action = "socket_rebind" if ("port_id" in patch or "socket_id" in patch) else "update"
            self._audit(
                conn,
                branch_id=branch_id,
                entity_type="map_point",
                entity_id=str(point_id),
                action=audit_action,
                diff={
                    "before": self._d(before),
                    "after": self._d(after),
                    "socket": socket_after or None,
                    "socket_key": _socket_key(socket_after),
                },
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._map_point_public_dict(after_with_links)

    def delete_map_point(self, *, point_id: int, actor_user_id: Optional[int], actor_role: Optional[str]) -> bool:
        with self._lock, self._connect() as conn:
            before = conn.execute("SELECT * FROM network_map_points WHERE id=?", (int(point_id),)).fetchone()
            if before is None:
                return False
            conn.execute("DELETE FROM network_map_points WHERE id=?", (int(point_id),))
            self._audit(
                conn,
                branch_id=int(before["branch_id"]),
                entity_type="map_point",
                entity_id=str(point_id),
                action="delete",
                diff={"before": self._d(before)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return True

    def get_map_file(self, map_id: int) -> dict[str, Any] | None:
        with self._lock, self._connect() as conn:
            row = conn.execute("SELECT id,file_name,mime_type,file_blob,file_size FROM network_maps WHERE id=?", (int(map_id),)).fetchone()
            return self._d(row)

    @staticmethod
    def _is_pdf_map(file_name: str, mime_type: str) -> bool:
        fn = _s(file_name).lower()
        mt = _s(mime_type).lower()
        return fn.endswith(".pdf") or "application/pdf" in mt or mt == "pdf"

    @staticmethod
    def _render_pdf_first_page_to_png(file_blob: bytes) -> bytes | None:
        if not file_blob or fitz is None:
            return None
        doc = None
        try:
            doc = fitz.open(stream=file_blob, filetype="pdf")
            if doc.page_count < 1:
                return None
            page = doc.load_page(0)
            matrix = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            return pix.tobytes("png")
        except Exception:
            return None
        finally:
            if doc is not None:
                doc.close()

    def get_map_file_for_view(self, map_id: int, *, render_mode: str = "auto") -> dict[str, Any] | None:
        row = self.get_map_file(map_id)
        if not row:
            return None

        mode = _s(render_mode).lower() or "auto"
        if mode not in {"auto", "original", "image"}:
            mode = "auto"

        file_name = _s(row.get("file_name")) or f"map_{int(map_id)}.bin"
        mime_type = _s(row.get("mime_type")) or "application/octet-stream"
        raw_blob = row.get("file_blob") or b""
        file_blob = bytes(raw_blob) if isinstance(raw_blob, (bytes, bytearray, memoryview)) else b""

        if mode == "original":
            return {**row, "rendered_from": ""}

        is_pdf = self._is_pdf_map(file_name, mime_type)
        should_render_to_image = is_pdf and mode in {"auto", "image"}
        if should_render_to_image:
            png_bytes = self._render_pdf_first_page_to_png(file_blob)
            if png_bytes:
                png_name = f"{Path(file_name).stem}.png"
                return {
                    **row,
                    "file_name": png_name,
                    "mime_type": "image/png",
                    "file_blob": png_bytes,
                    "file_size": len(png_bytes),
                    "rendered_from": "pdf",
                }
            if mode == "image":
                raise ValueError("Не удалось конвертировать PDF в изображение. Установите PyMuPDF.")

        return {**row, "rendered_from": ""}

    def list_audit(self, *, branch_id: Optional[int], limit: int = 100) -> list[dict[str, Any]]:
        safe_limit = max(1, min(int(limit), 500))
        with self._lock, self._connect() as conn:
            if branch_id is None:
                rows = conn.execute(
                    "SELECT * FROM network_audit_log ORDER BY id DESC LIMIT ?",
                    (safe_limit,),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM network_audit_log WHERE branch_id=? ORDER BY id DESC LIMIT ?",
                    (int(branch_id), safe_limit),
                ).fetchall()
            out: list[dict[str, Any]] = []
            for row in rows:
                item = self._d(row) or {}
                try:
                    item["diff"] = json.loads(item.get("diff_json") or "{}")
                except Exception:
                    item["diff"] = {}
                out.append(item)
            return out

    def create_device(self, *, branch_id: int, payload: dict[str, Any], actor_user_id: Optional[int], actor_role: Optional[str]) -> dict[str, Any]:
        code = _s(payload.get("device_code"))
        if not code:
            raise ValueError("device_code is required")
        now = _now()
        with self._lock, self._connect() as conn:
            if conn.execute("SELECT 1 FROM network_branches WHERE id=?", (int(branch_id),)).fetchone() is None:
                raise ValueError("Branch not found")
            site_id = self._ensure_site(
                conn,
                branch_id=int(branch_id),
                site_code=payload.get("site_code"),
                site_name=payload.get("site_name") or payload.get("site_code") or "",
            )
            conn.execute(
                """
                INSERT INTO network_devices(branch_id,site_id,device_code,device_type,vendor,model,sheet_name,mgmt_ip,notes,created_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    int(branch_id),
                    site_id,
                    code,
                    _s(payload.get("device_type")) or "switch",
                    _s(payload.get("vendor")) or None,
                    _s(payload.get("model")) or None,
                    _s(payload.get("sheet_name")) or None,
                    _s(payload.get("mgmt_ip")) or None,
                    _s(payload.get("notes")) or None,
                    now,
                    now,
                ),
            )
            new_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            row = conn.execute("SELECT * FROM network_devices WHERE id=?", (new_id,)).fetchone()
            self._audit(
                conn,
                branch_id=int(branch_id),
                entity_type="device",
                entity_id=str(new_id),
                action="create",
                diff={"after": self._d(row)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._d(row) or {}

    def update_device(self, *, device_id: int, payload: dict[str, Any], actor_user_id: Optional[int], actor_role: Optional[str]) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            before = conn.execute("SELECT * FROM network_devices WHERE id=?", (int(device_id),)).fetchone()
            if before is None:
                raise ValueError("Device not found")
            patch = {k: v for k, v in payload.items() if k in {"device_code", "device_type", "vendor", "model", "sheet_name", "mgmt_ip", "notes", "site_code", "site_name"}}
            if not patch:
                raise ValueError("No fields to update")
            fields: list[str] = []
            params: list[Any] = []
            for key in ("device_code", "device_type", "vendor", "model", "sheet_name", "mgmt_ip", "notes"):
                if key in patch:
                    fields.append(f"{key}=?")
                    params.append(_s(patch.get(key)) or None)
            if "site_code" in patch:
                site_id = self._ensure_site(
                    conn,
                    branch_id=int(before["branch_id"]),
                    site_code=patch.get("site_code"),
                    site_name=patch.get("site_name") or patch.get("site_code") or "",
                )
                fields.append("site_id=?")
                params.append(site_id)
            fields.append("updated_at=?")
            params.append(_now())
            params.append(int(device_id))
            conn.execute(f"UPDATE network_devices SET {', '.join(fields)} WHERE id=?", params)
            after = conn.execute("SELECT * FROM network_devices WHERE id=?", (int(device_id),)).fetchone()
            self._audit(
                conn,
                branch_id=int(before["branch_id"]),
                entity_type="device",
                entity_id=str(device_id),
                action="update",
                diff={"before": self._d(before), "after": self._d(after)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._d(after) or {}

    def delete_device(self, *, device_id: int, actor_user_id: Optional[int], actor_role: Optional[str]) -> bool:
        with self._lock, self._connect() as conn:
            before = conn.execute("SELECT * FROM network_devices WHERE id=?", (int(device_id),)).fetchone()
            if before is None:
                return False
            conn.execute("DELETE FROM network_devices WHERE id=?", (int(device_id),))
            self._audit(
                conn,
                branch_id=int(before["branch_id"]),
                entity_type="device",
                entity_id=str(device_id),
                action="delete",
                diff={"before": self._d(before)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return True

    def create_port(self, *, device_id: int, payload: dict[str, Any], actor_user_id: Optional[int], actor_role: Optional[str]) -> dict[str, Any]:
        port_name = _s(payload.get("port_name"))
        if not port_name:
            raise ValueError("port_name is required")
        ename = _s(payload.get("endpoint_name_raw"))
        eip = _s(payload.get("endpoint_ip_raw"))
        emac = _normalize_mac_multiline(payload.get("endpoint_mac_raw"))
        cnt, occ = _occupied(ename, eip, emac)
        vlan_raw = _s(payload.get("vlan_raw"))
        with self._lock, self._connect() as conn:
            device = conn.execute("SELECT id,branch_id FROM network_devices WHERE id=?", (int(device_id),)).fetchone()
            if device is None:
                raise ValueError("Device not found")
            now = _now()
            conn.execute(
                """
                INSERT INTO network_ports(device_id,port_name,patch_panel_port,location_code,vlan_raw,vlan_normalized_json,endpoint_name_raw,endpoint_ip_raw,endpoint_mac_raw,endpoint_count,is_occupied,row_source_hash,created_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    int(device_id),
                    port_name,
                    _s(payload.get("patch_panel_port")) or None,
                    _s(payload.get("location_code")) or None,
                    vlan_raw or None,
                    json.dumps(_vlans(vlan_raw), ensure_ascii=False),
                    ename or None,
                    eip or None,
                    emac or None,
                    cnt,
                    occ,
                    _s(payload.get("row_source_hash")) or None,
                    now,
                    now,
                ),
            )
            new_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            self._sync_socket_for_port_in_conn(conn, port_id=new_id)
            row = conn.execute("SELECT * FROM network_ports WHERE id=?", (new_id,)).fetchone()
            self._audit(
                conn,
                branch_id=int(device["branch_id"]),
                entity_type="port",
                entity_id=str(new_id),
                action="create",
                diff={"after": self._d(row)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._d(row) or {}

    def update_port(self, *, port_id: int, payload: dict[str, Any], actor_user_id: Optional[int], actor_role: Optional[str]) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            before = conn.execute(
                "SELECT p.*, d.branch_id FROM network_ports p JOIN network_devices d ON d.id=p.device_id WHERE p.id=?",
                (int(port_id),),
            ).fetchone()
            if before is None:
                raise ValueError("Port not found")
            patch = {k: v for k, v in payload.items() if k in {"port_name", "patch_panel_port", "location_code", "vlan_raw", "endpoint_name_raw", "endpoint_ip_raw", "endpoint_mac_raw", "row_source_hash"}}
            if not patch:
                raise ValueError("No fields to update")
            ename = _s(patch.get("endpoint_name_raw", before["endpoint_name_raw"]))
            eip = _s(patch.get("endpoint_ip_raw", before["endpoint_ip_raw"]))
            emac = _normalize_mac_multiline(patch.get("endpoint_mac_raw", before["endpoint_mac_raw"]))
            cnt, occ = _occupied(ename, eip, emac)
            vlan_raw = _s(patch.get("vlan_raw", before["vlan_raw"]))

            fields: list[str] = []
            params: list[Any] = []
            for key in ("port_name", "patch_panel_port", "location_code", "vlan_raw", "endpoint_name_raw", "endpoint_ip_raw", "endpoint_mac_raw", "row_source_hash"):
                if key in patch:
                    fields.append(f"{key}=?")
                    if key == "endpoint_mac_raw":
                        params.append(_normalize_mac_multiline(patch.get(key)) or None)
                    else:
                        params.append(_s(patch.get(key)) or None)
            fields.extend(["vlan_normalized_json=?", "endpoint_count=?", "is_occupied=?", "updated_at=?"])
            params.extend([json.dumps(_vlans(vlan_raw), ensure_ascii=False), cnt, occ, _now(), int(port_id)])
            conn.execute(f"UPDATE network_ports SET {', '.join(fields)} WHERE id=?", params)
            socket_id_after = self._sync_socket_for_port_in_conn(conn, port_id=int(port_id))
            after = conn.execute("SELECT * FROM network_ports WHERE id=?", (int(port_id),)).fetchone()
            socket_after_row = conn.execute(
                "SELECT id, socket_code FROM network_sockets WHERE id=?",
                (int(socket_id_after),),
            ).fetchone() if socket_id_after is not None else None
            point_after_row = conn.execute(
                "SELECT id, map_id FROM network_map_points WHERE socket_id=? ORDER BY id LIMIT 1",
                (int(socket_id_after),),
            ).fetchone() if socket_id_after is not None else None
            self._audit(
                conn,
                branch_id=int(before["branch_id"]),
                entity_type="port",
                entity_id=str(port_id),
                action="update",
                diff={"before": self._d(before), "after": self._d(after)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            payload_out = self._d(after) or {}
            payload_out["socket_id"] = int(socket_after_row["id"]) if socket_after_row is not None else None
            payload_out["socket_code"] = _s(socket_after_row["socket_code"] if socket_after_row is not None else payload_out.get("patch_panel_port"))
            payload_out["map_point_id"] = int(point_after_row["id"]) if point_after_row is not None else None
            payload_out["map_id"] = int(point_after_row["map_id"]) if point_after_row is not None else None
            payload_out["requires_point_creation"] = bool(payload_out["socket_code"]) and point_after_row is None
            return payload_out

    def delete_port(self, *, port_id: int, actor_user_id: Optional[int], actor_role: Optional[str]) -> bool:
        with self._lock, self._connect() as conn:
            before = conn.execute(
                "SELECT p.*, d.branch_id FROM network_ports p JOIN network_devices d ON d.id=p.device_id WHERE p.id=?",
                (int(port_id),),
            ).fetchone()
            if before is None:
                return False
            conn.execute("DELETE FROM network_ports WHERE id=?", (int(port_id),))
            self._sync_socket_for_port_in_conn(conn, port_id=int(port_id))
            self._audit(
                conn,
                branch_id=int(before["branch_id"]),
                entity_type="port",
                entity_id=str(port_id),
                action="delete",
                diff={"before": self._d(before)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return True

    def bootstrap_device_ports(
        self,
        *,
        device_id: int,
        port_count: int,
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        """
        Bulk create ports for a device with names PORT 1, PORT 2, ..., PORT N.
        Skips ports that already exist (by port_name).
        """
        if port_count < 1:
            raise ValueError("port_count must be at least 1")
        if port_count > 512:
            raise ValueError("port_count cannot exceed 512")

        with self._lock, self._connect() as conn:
            device = conn.execute(
                "SELECT id, branch_id, device_code FROM network_devices WHERE id=?",
                (int(device_id),),
            ).fetchone()
            if device is None:
                raise ValueError("Device not found")

            branch_id = int(device["branch_id"])
            device_code = _s(device["device_code"])
            now = _now()

            # Get existing port names to avoid duplicates
            existing_rows = conn.execute(
                "SELECT port_name FROM network_ports WHERE device_id=?",
                (int(device_id),),
            ).fetchall()
            existing_names = {_s(r["port_name"]) for r in existing_rows}

            created_count = 0
            skipped_count = 0
            created_ids: list[int] = []

            for i in range(1, port_count + 1):
                port_name = str(i)
                if port_name in existing_names:
                    skipped_count += 1
                    continue

                conn.execute(
                    """
                    INSERT INTO network_ports(device_id,port_name,patch_panel_port,location_code,vlan_raw,vlan_normalized_json,endpoint_name_raw,endpoint_ip_raw,endpoint_mac_raw,endpoint_count,is_occupied,row_source_hash,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        int(device_id),
                        port_name,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        0,
                        0,
                        None,
                        now,
                        now,
                    ),
                )
                new_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
                created_ids.append(new_id)
                created_count += 1

            # Audit log for bulk operation
            self._audit(
                conn,
                branch_id=branch_id,
                entity_type="device",
                entity_id=str(device_id),
                action="bootstrap_ports",
                diff={
                    "port_count": port_count,
                    "created_count": created_count,
                    "skipped_count": skipped_count,
                    "port_ids": created_ids,
                },
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()

            return {
                "device_id": device_id,
                "device_code": device_code,
                "port_count": port_count,
                "created_count": created_count,
                "skipped_count": skipped_count,
                "port_ids": created_ids,
            }

    def upload_map(
        self,
        *,
        branch_id: int,
        file_name: str,
        mime_type: str,
        file_bytes: bytes,
        site_code: Optional[str],
        site_name: Optional[str],
        title: Optional[str],
        floor_label: Optional[str],
        source_path: Optional[str],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        if not file_bytes:
            raise ValueError("file is empty")
        checksum = hashlib.sha256(file_bytes).hexdigest()
        now = _now()
        with self._lock, self._connect() as conn:
            if conn.execute("SELECT 1 FROM network_branches WHERE id=?", (int(branch_id),)).fetchone() is None:
                raise ValueError("Branch not found")
            site_id = self._ensure_site(
                conn,
                branch_id=int(branch_id),
                site_code=site_code,
                site_name=site_name or site_code or "",
            )
            existing = conn.execute(
                "SELECT * FROM network_maps WHERE branch_id=? AND file_name=?",
                (int(branch_id), _s(file_name)),
            ).fetchone()
            if existing is None:
                conn.execute(
                    """
                    INSERT INTO network_maps(branch_id,site_id,title,floor_label,file_name,mime_type,file_blob,file_size,checksum_sha256,source_path,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        int(branch_id),
                        site_id,
                        _s(title) or None,
                        _s(floor_label) or None,
                        _s(file_name),
                        _s(mime_type) or "application/octet-stream",
                        sqlite3.Binary(file_bytes),
                        len(file_bytes),
                        checksum,
                        _s(source_path) or None,
                        now,
                        now,
                    ),
                )
                map_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
                action = "create"
            else:
                map_id = int(existing["id"])
                conn.execute(
                    """
                    UPDATE network_maps
                    SET site_id=?, title=?, floor_label=?, mime_type=?, file_blob=?, file_size=?, checksum_sha256=?, source_path=?, updated_at=?
                    WHERE id=?
                    """,
                    (
                        site_id,
                        _s(title) or existing["title"],
                        _s(floor_label) or existing["floor_label"],
                        _s(mime_type) or existing["mime_type"],
                        sqlite3.Binary(file_bytes),
                        len(file_bytes),
                        checksum,
                        _s(source_path) or existing["source_path"],
                        now,
                        map_id,
                    ),
                )
                action = "update"
            row = conn.execute("SELECT * FROM network_maps WHERE id=?", (map_id,)).fetchone()
            self._audit(
                conn,
                branch_id=int(branch_id),
                entity_type="map",
                entity_id=str(map_id),
                action=action,
                diff={"after": self._d(row), "file_name": _s(file_name)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._map_row_for_api(conn, map_id=map_id)

    def update_map_meta(self, *, map_id: int, payload: dict[str, Any], actor_user_id: Optional[int], actor_role: Optional[str]) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            before = conn.execute("SELECT * FROM network_maps WHERE id=?", (int(map_id),)).fetchone()
            if before is None:
                raise ValueError("Map not found")
            fields: list[str] = []
            params: list[Any] = []
            if "title" in payload:
                fields.append("title=?")
                params.append(_s(payload.get("title")) or None)
            if "floor_label" in payload:
                fields.append("floor_label=?")
                params.append(_s(payload.get("floor_label")) or None)
            if "site_code" in payload:
                site_id = self._ensure_site(
                    conn,
                    branch_id=int(before["branch_id"]),
                    site_code=payload.get("site_code"),
                    site_name=payload.get("site_name") or payload.get("site_code") or "",
                )
                fields.append("site_id=?")
                params.append(site_id)
            if not fields:
                raise ValueError("No fields to update")
            fields.append("updated_at=?")
            params.append(_now())
            params.append(int(map_id))
            conn.execute(f"UPDATE network_maps SET {', '.join(fields)} WHERE id=?", params)
            after = conn.execute("SELECT * FROM network_maps WHERE id=?", (int(map_id),)).fetchone()
            self._audit(
                conn,
                branch_id=int(before["branch_id"]),
                entity_type="map",
                entity_id=str(map_id),
                action="update",
                diff={"before": self._d(before), "after": self._d(after)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return self._map_row_for_api(conn, map_id=int(map_id))

    def delete_map(self, *, map_id: int, actor_user_id: Optional[int], actor_role: Optional[str]) -> bool:
        with self._lock, self._connect() as conn:
            before = conn.execute("SELECT * FROM network_maps WHERE id=?", (int(map_id),)).fetchone()
            if before is None:
                return False
            conn.execute("DELETE FROM network_maps WHERE id=?", (int(map_id),))
            self._audit(
                conn,
                branch_id=int(before["branch_id"]),
                entity_type="map",
                entity_id=str(map_id),
                action="delete",
                diff={"before": self._d(before)},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()
            return True

    def import_branch_data(
        self,
        *,
        city_code: str,
        branch_code: str,
        branch_name: str,
        excel_file_name: str,
        excel_file_bytes: bytes,
        map_files: Iterable[dict[str, Any]],
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        if openpyxl is None:
            raise RuntimeError("openpyxl is not available")
        if not excel_file_bytes:
            raise ValueError("excel file is empty")

        summary = ImportSummary()
        started = _now()
        with self._lock, self._connect() as conn:
            conn.execute(
                "INSERT INTO network_import_jobs(city_code,status,started_at) VALUES (?, 'running', ?)",
                (_s(city_code), started),
            )
            job_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            conn.commit()

        try:
            branch = self.ensure_branch(city_code=city_code, branch_code=branch_code, name=branch_name)
            branch_id = int(branch["id"])

            wb = openpyxl.load_workbook(BytesIO(excel_file_bytes), data_only=True)
            summary.sheets_total = len(wb.sheetnames)

            with self._lock, self._connect() as conn:
                site19 = self._ensure_site(conn, branch_id=branch_id, site_code="p19", site_name="Первомайская 19")
                site21 = self._ensure_site(conn, branch_id=branch_id, site_code="p21", site_name="Первомайская 21")
                site_ids = {"p19": site19, "p21": site21}
                now = _now()

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    sc = _site_code(sheet_name)
                    existing_device = conn.execute(
                        "SELECT id FROM network_devices WHERE branch_id=? AND device_code=?",
                        (branch_id, _s(sheet_name)),
                    ).fetchone()
                    if existing_device is None:
                        conn.execute(
                            """
                            INSERT INTO network_devices(branch_id,site_id,device_code,device_type,sheet_name,created_at,updated_at)
                            VALUES (?,?,?,?,?,?,?)
                            """,
                            (branch_id, site_ids.get(sc), _s(sheet_name), "switch", _s(sheet_name), now, now),
                        )
                        device_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
                        summary.devices_created += 1
                    else:
                        device_id = int(existing_device["id"])
                        conn.execute("UPDATE network_devices SET site_id=?, sheet_name=?, updated_at=? WHERE id=?", (site_ids.get(sc), _s(sheet_name), now, device_id))
                        summary.devices_updated += 1

                    headers = {_h(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
                    col_name = headers.get("name")
                    col_ip = headers.get("ip address")
                    col_mac = headers.get("mac address")
                    col_vlan = headers.get("vlan")
                    col_port = headers.get("port")
                    col_pp = headers.get("port p/p") or headers.get("port p p")
                    col_loc = headers.get("location")
                    if col_port is None:
                        continue

                    for row_no in range(2, ws.max_row + 1):
                        port_name = _s(ws.cell(row_no, col_port).value)
                        if not port_name:
                            continue
                        name_raw = _s(ws.cell(row_no, col_name).value) if col_name else ""
                        ip_raw = _s(ws.cell(row_no, col_ip).value) if col_ip else ""
                        mac_raw = _s(ws.cell(row_no, col_mac).value) if col_mac else ""
                        vlan_raw = _s(ws.cell(row_no, col_vlan).value) if col_vlan else ""
                        pp_raw = _s(ws.cell(row_no, col_pp).value) if col_pp else ""
                        loc_raw = _s(ws.cell(row_no, col_loc).value) if col_loc else ""
                        cnt, occ = _occupied(name_raw, ip_raw, mac_raw)
                        row_hash = hashlib.sha1(f"{sheet_name}|{row_no}|{port_name}|{name_raw}|{ip_raw}|{mac_raw}|{vlan_raw}|{pp_raw}|{loc_raw}".encode("utf-8", errors="ignore")).hexdigest()
                        ex_port = conn.execute("SELECT id FROM network_ports WHERE device_id=? AND port_name=?", (device_id, port_name)).fetchone()
                        if ex_port is None:
                            conn.execute(
                                """
                                INSERT INTO network_ports(device_id,port_name,patch_panel_port,location_code,vlan_raw,vlan_normalized_json,endpoint_name_raw,endpoint_ip_raw,endpoint_mac_raw,endpoint_count,is_occupied,row_source_hash,created_at,updated_at)
                                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                                """,
                                (
                                    device_id,
                                    port_name,
                                    pp_raw or None,
                                    loc_raw or None,
                                    vlan_raw or None,
                                    json.dumps(_vlans(vlan_raw), ensure_ascii=False),
                                    name_raw or None,
                                    ip_raw or None,
                                    mac_raw or None,
                                    cnt,
                                    occ,
                                    row_hash,
                                    now,
                                    now,
                                ),
                            )
                            summary.ports_created += 1
                        else:
                            conn.execute(
                                """
                                UPDATE network_ports SET
                                    patch_panel_port=?,location_code=?,vlan_raw=?,vlan_normalized_json=?,
                                    endpoint_name_raw=?,endpoint_ip_raw=?,endpoint_mac_raw=?,
                                    endpoint_count=?,is_occupied=?,row_source_hash=?,updated_at=?
                                WHERE device_id=? AND port_name=?
                                """,
                                (
                                    pp_raw or None,
                                    loc_raw or None,
                                    vlan_raw or None,
                                    json.dumps(_vlans(vlan_raw), ensure_ascii=False),
                                    name_raw or None,
                                    ip_raw or None,
                                    mac_raw or None,
                                    cnt,
                                    occ,
                                    row_hash,
                                    now,
                                    device_id,
                                    port_name,
                                ),
                            )
                            summary.ports_updated += 1
                        summary.ports_total += 1
                for mf in map_files:
                    fname = _s(mf.get("file_name"))
                    fbytes = mf.get("file_bytes") or b""
                    if not fname or not isinstance(fbytes, (bytes, bytearray)):
                        continue
                    site_code = _site_code(fname)
                    site_name = "Первомайская 21" if site_code == "p21" else "Первомайская 19"
                    site_id = self._ensure_site(conn, branch_id=branch_id, site_code=site_code, site_name=site_name)
                    mime_type = _s(mf.get("mime_type")) or "application/octet-stream"
                    checksum = hashlib.sha256(bytes(fbytes)).hexdigest()
                    exists_map = conn.execute(
                        "SELECT id FROM network_maps WHERE branch_id=? AND file_name=?",
                        (branch_id, fname),
                    ).fetchone()
                    if exists_map is None:
                        conn.execute(
                            """
                            INSERT INTO network_maps(branch_id,site_id,title,floor_label,file_name,mime_type,file_blob,file_size,checksum_sha256,source_path,created_at,updated_at)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                            """,
                            (
                                branch_id,
                                site_id,
                                _s(mf.get("title")) or fname,
                                _s(mf.get("floor_label")) or None,
                                fname,
                                mime_type,
                                sqlite3.Binary(bytes(fbytes)),
                                len(bytes(fbytes)),
                                checksum,
                                _s(mf.get("source_path")) or None,
                                now,
                                now,
                            ),
                        )
                    else:
                        conn.execute(
                            """
                            UPDATE network_maps
                            SET site_id=?, title=?, floor_label=?, mime_type=?, file_blob=?, file_size=?, checksum_sha256=?, source_path=?, updated_at=?
                            WHERE id=?
                            """,
                            (
                                site_id,
                                _s(mf.get("title")) or fname,
                                _s(mf.get("floor_label")) or None,
                                mime_type,
                                sqlite3.Binary(bytes(fbytes)),
                                len(bytes(fbytes)),
                                checksum,
                                _s(mf.get("source_path")) or None,
                                now,
                                int(exists_map["id"]),
                            ),
                        )
                    if exists_map is None:
                        summary.maps_created += 1
                    else:
                        summary.maps_updated += 1

                self._sync_all_sockets_in_conn(conn, branch_id=branch_id)
                self._link_map_points_to_sockets_in_conn(conn, branch_id=branch_id)

                self._audit(
                    conn,
                    branch_id=branch_id,
                    entity_type="import",
                    entity_id=str(job_id),
                    action="import",
                    diff={"excel_file_name": excel_file_name, "summary": summary.as_dict()},
                    actor_user_id=actor_user_id,
                    actor_role=actor_role,
                )
                conn.execute(
                    "UPDATE network_import_jobs SET branch_id=?, status='success', finished_at=?, summary_json=? WHERE id=?",
                    (branch_id, _now(), json.dumps(summary.as_dict(), ensure_ascii=False), job_id),
                )
                conn.commit()

            return {"job_id": job_id, "branch_id": branch_id, "branch_code": _s(branch_code), "summary": summary.as_dict()}
        except Exception as exc:
            with self._lock, self._connect() as conn:
                conn.execute(
                    "UPDATE network_import_jobs SET status='failed', finished_at=?, error_text=? WHERE id=?",
                    (_now(), str(exc), job_id),
                )
                conn.commit()
            raise


    def import_equipment_from_excel(
        self,
        *,
        branch_id: int,
        file_name: str,
        file_bytes: bytes,
        actor_user_id: Optional[int],
        actor_role: Optional[str],
    ) -> dict[str, Any]:
        """Import devices and ports from a multi-sheet Excel file into an existing branch.
        Each sheet represents a network device (switch). Rows are ports.
        The 'Port P/P' column links ports to existing sockets."""
        if openpyxl is None:
            raise RuntimeError("openpyxl is not available")
        if not file_bytes:
            raise ValueError("Excel file is empty")

        summary = ImportSummary()
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        summary.sheets_total = len(wb.sheetnames)

        with self._lock, self._connect() as conn:
            branch = conn.execute("SELECT * FROM network_branches WHERE id=?", (int(branch_id),)).fetchone()
            if branch is None:
                raise ValueError("Branch not found")

            now = _now()

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = {_h(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}

                col_port = headers.get("port")
                if col_port is None:
                    # Sheet has no "Port" column — skip
                    continue

                col_name = headers.get("name")
                col_ip = headers.get("ip address") or headers.get("ip")
                col_mac = headers.get("mac address") or headers.get("mac")
                col_vlan = headers.get("vlan")
                col_pp = headers.get("port p p")
                col_loc = headers.get("location")
                col_switch = headers.get("swich") or headers.get("switch") or headers.get("asw")

                # Determine device_code: prefer "Swich" column value from first data row, fallback to sheet_name
                device_code = _s(sheet_name)
                if col_switch:
                    for rn in range(2, ws.max_row + 1):
                        val = _s(ws.cell(rn, col_switch).value)
                        if val:
                            device_code = val
                            break

                existing_device = conn.execute(
                    "SELECT id FROM network_devices WHERE branch_id=? AND device_code=?",
                    (branch_id, device_code),
                ).fetchone()
                if existing_device is None:
                    conn.execute(
                        """
                        INSERT INTO network_devices(branch_id, device_code, device_type, sheet_name, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (branch_id, device_code, "switch", _s(sheet_name), now, now),
                    )
                    device_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
                    summary.devices_created += 1
                else:
                    device_id = int(existing_device["id"])
                    conn.execute(
                        "UPDATE network_devices SET sheet_name=?, updated_at=? WHERE id=?",
                        (_s(sheet_name), now, device_id),
                    )
                    summary.devices_updated += 1

                for row_no in range(2, ws.max_row + 1):
                    port_name = _s(ws.cell(row_no, col_port).value)
                    if not port_name:
                        continue

                    name_raw = _s(ws.cell(row_no, col_name).value) if col_name else ""
                    ip_raw = _s(ws.cell(row_no, col_ip).value) if col_ip else ""
                    mac_raw = _s(ws.cell(row_no, col_mac).value) if col_mac else ""
                    vlan_raw = _s(ws.cell(row_no, col_vlan).value) if col_vlan else ""
                    pp_raw = _s(ws.cell(row_no, col_pp).value) if col_pp else ""
                    loc_raw = _s(ws.cell(row_no, col_loc).value) if col_loc else ""

                    cnt, occ = _occupied(name_raw, ip_raw, mac_raw)
                    row_hash = hashlib.sha1(
                        f"{sheet_name}|{row_no}|{port_name}|{name_raw}|{ip_raw}|{mac_raw}|{vlan_raw}|{pp_raw}|{loc_raw}".encode("utf-8", errors="ignore")
                    ).hexdigest()

                    ex_port = conn.execute(
                        "SELECT id FROM network_ports WHERE device_id=? AND port_name=?",
                        (device_id, port_name),
                    ).fetchone()

                    if ex_port is None:
                        conn.execute(
                            """
                            INSERT INTO network_ports(
                                device_id, port_name, patch_panel_port, location_code,
                                vlan_raw, vlan_normalized_json,
                                endpoint_name_raw, endpoint_ip_raw, endpoint_mac_raw,
                                endpoint_count, is_occupied, row_source_hash, created_at, updated_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                device_id, port_name, pp_raw or None, loc_raw or None,
                                vlan_raw or None,
                                json.dumps(_vlans(vlan_raw), ensure_ascii=False),
                                name_raw or None, ip_raw or None, mac_raw or None,
                                cnt, occ, row_hash, now, now,
                            ),
                        )
                        summary.ports_created += 1
                    else:
                        conn.execute(
                            """
                            UPDATE network_ports SET
                                patch_panel_port=?, location_code=?, vlan_raw=?, vlan_normalized_json=?,
                                endpoint_name_raw=?, endpoint_ip_raw=?, endpoint_mac_raw=?,
                                endpoint_count=?, is_occupied=?, row_source_hash=?, updated_at=?
                            WHERE device_id=? AND port_name=?
                            """,
                            (
                                pp_raw or None, loc_raw or None,
                                vlan_raw or None,
                                json.dumps(_vlans(vlan_raw), ensure_ascii=False),
                                name_raw or None, ip_raw or None, mac_raw or None,
                                cnt, occ, row_hash, now,
                                device_id, port_name,
                            ),
                        )
                        summary.ports_updated += 1
                    summary.ports_total += 1

            # After importing all devices/ports, sync sockets ↔ ports via patch_panel_port
            self._sync_all_sockets_in_conn(conn, branch_id=branch_id)
            self._link_map_points_to_sockets_in_conn(conn, branch_id=branch_id)

            self._audit(
                conn,
                branch_id=branch_id,
                entity_type="equipment_import",
                entity_id=str(branch_id),
                action="import_equipment",
                diff={"file_name": _s(file_name), "summary": summary.as_dict()},
                actor_user_id=actor_user_id,
                actor_role=actor_role,
            )
            conn.commit()

        return {
            "branch_id": int(branch_id),
            "file_name": _s(file_name),
            "summary": summary.as_dict(),
        }


network_service = NetworkService()
