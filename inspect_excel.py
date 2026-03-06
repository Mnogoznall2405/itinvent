import openpyxl

wb = openpyxl.load_workbook(r"C:\Project\Image_scan\_Сети\!Сети\tmn\ip mac address_19-21.xlsx", data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
    # Print header row
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        headers.append(val)
    print(f"Header row: {headers}")
    # Print first 3 data rows
    for row in range(2, min(5, ws.max_row + 1)):
        row_data = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        print(f"Row {row}: {row_data}")
