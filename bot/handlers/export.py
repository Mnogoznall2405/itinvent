#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Обработчики экспорта данных
Экспорт ненайденного оборудования и перемещений, отправка на email.
"""
import logging
import os
import json
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ContextTypes, ConversationHandler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from bot.config import States, Messages
from bot.utils.decorators import require_user_access, handle_errors
from bot.utils.keyboards import create_main_menu_keyboard
from database_manager import database_manager
from equipment_data_manager import EquipmentDataManager
from email_sender import send_export_email

logger = logging.getLogger(__name__)

# Глобальный менеджер данных
equipment_manager = EquipmentDataManager()


@require_user_access
async def show_export_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Отображает меню экспорта данных

    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения

    Возвращает:
        int: Состояние для ConversationHandler
    """
    keyboard = [
        [InlineKeyboardButton("📦 Экспорт ненайденного оборудования", callback_data="export_type:unfound")],
        [InlineKeyboardButton("🔄 Экспорт перемещений", callback_data="export_type:transfers")],
        [InlineKeyboardButton("🔧 Экспорт замен комплектующих", callback_data="export_type:cartridges")],
        [InlineKeyboardButton("🔋 Экспорт замены батареи ИБП", callback_data="export_type:battery")],
        [InlineKeyboardButton("🔙 Назад в главное меню", callback_data="back_to_main")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        "📊 <b>Экспорт данных</b>\n\n"
        "Выберите тип данных для экспорта:",
        parse_mode='HTML',
        reply_markup=reply_markup
    )

    return States.DB_SELECTION_MENU  # Используем существующее состояние


@handle_errors
async def handle_export_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обрабатывает выбор типа экспорта
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: Следующее состояние
    """
    query = update.callback_query
    await query.answer()
    
    callback_data = query.data
    
    if callback_data.startswith("export_type:"):
        export_type = callback_data.split(":")[1]
        context.user_data['export_type'] = export_type
        
        # Показываем выбор периода
        return await show_export_period(update, context)
    
    elif callback_data == "back_to_main":
        await query.edit_message_text("✅ Возврат в главное меню")
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=Messages.MAIN_MENU,
            reply_markup=create_main_menu_keyboard()
        )
        return ConversationHandler.END
    
    return States.DB_SELECTION_MENU


async def show_export_period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Показывает выбор периода экспорта

    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения

    Возвращает:
        int: Следующее состояние
    """
    export_type = context.user_data.get('export_type', 'unfound')

    # Разные клавиатуры для разных типов экспорта
    if export_type in ('cartridges', 'battery'):
        # Для картриджей и батареи - выбор периода без выбора базы
        keyboard = [
            [InlineKeyboardButton("📅 За последний месяц", callback_data="export_period:1month")],
            [InlineKeyboardButton("📊 За последние 3 месяца", callback_data="export_period:3months")],
            [InlineKeyboardButton("📕 За весь период", callback_data="export_period:all")],
            [InlineKeyboardButton("🔙 Назад", callback_data="back_to_export_menu")]
        ]
    else:
        # Для остальных типов - стандартный выбор
        keyboard = [
            [InlineKeyboardButton("📅 Все данные", callback_data="export_period:full")],
            [InlineKeyboardButton("🆕 Только новые", callback_data="export_period:new")],
            [InlineKeyboardButton("🔙 Назад", callback_data="back_to_export_menu")]
        ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    type_names = {
        'unfound': 'ненайденного оборудования',
        'transfers': 'перемещений',
        'cartridges': 'замен комплектующих',
        'battery': 'замен батареи ИБП'
    }
    type_name = type_names.get(export_type, 'данных')

    period_text = "Выберите период для экспорта:"
    if export_type == 'cartridges':
        period_text = "Выберите период для анализа картриджей:"
    elif export_type == 'battery':
        period_text = "Выберите период для анализа замен батареи:"

    await update.callback_query.edit_message_text(
        f"📊 <b>Экспорт {type_name}</b>\n\n"
        f"{period_text}",
        parse_mode='HTML',
        reply_markup=reply_markup
    )

    return States.DB_SELECTION_MENU


@handle_errors
async def handle_export_period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обрабатывает выбор периода экспорта

    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения

    Возвращает:
        int: Следующее состояние
    """
    query = update.callback_query
    await query.answer()

    callback_data = query.data
    export_type = context.user_data.get('export_type', 'unfound')

    if callback_data.startswith("export_period:"):
        period = callback_data.split(":")[1]
        context.user_data['export_period'] = period

        # Для картриджей и батареи - прямой экспорт без выбора базы
        if export_type == 'cartridges':
            return await handle_cartridge_export_directly(update, context, period)
        elif export_type == 'battery':
            return await handle_battery_export_directly(update, context, period)
        else:
            # Для остальных типов - показываем выбор базы данных
            return await show_export_database(update, context)

    elif callback_data == "back_to_export_menu":
        # Возврат к выбору типа экспорта
        keyboard = [
            [InlineKeyboardButton("📦 Экспорт ненайденного оборудования", callback_data="export_type:unfound")],
            [InlineKeyboardButton("🔄 Экспорт перемещений", callback_data="export_type:transfers")],
            [InlineKeyboardButton("🔧 Экспорт замен комплектующих", callback_data="export_type:cartridges")],
            [InlineKeyboardButton("🔋 Экспорт замены батареи ИБП", callback_data="export_type:battery")],
            [InlineKeyboardButton("🔙 Назад в главное меню", callback_data="back_to_main")]
        ]

        reply_markup = InlineKeyboardMarkup(keyboard)

        await query.edit_message_text(
            "📊 <b>Экспорт данных</b>\n\n"
            "Выберите тип данных для экспорта:",
            parse_mode='HTML',
            reply_markup=reply_markup
        )

    return States.DB_SELECTION_MENU


@handle_errors
async def handle_cartridge_export_directly(update: Update, context: ContextTypes.DEFAULT_TYPE, period: str) -> int:
    """
    Обрабатывает прямой экспорт картриджей без выбора базы данных

    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        period: Выбранный период

    Возвращает:
        int: Следующее состояние
    """
    query = update.callback_query
    await query.edit_message_text("⏳ Анализ данных о заменах комплектующих...")

    try:
        # Выполняем экспорт с LLM-структурированием
        excel_file = await export_components_to_excel_structured(period=period, db_filter=None)

        if excel_file and os.path.exists(excel_file):
            context.user_data['export_file'] = excel_file
            return await show_delivery_options(update, context, excel_file)
        else:
            await query.edit_message_text(
                "❌ Нет данных для экспорта или ошибка создания файла."
            )
            return ConversationHandler.END

    except Exception as e:
        logger.error(f"Ошибка при экспорте картриджей: {e}")
        await query.edit_message_text(
            f"❌ Ошибка при экспорте: {str(e)}"
        )
        return ConversationHandler.END


async def show_export_database(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Показывает выбор базы данных для экспорта
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: Следующее состояние
    """
    # Получаем список доступных БД
    available_databases = database_manager.get_available_databases()
    
    keyboard = [[InlineKeyboardButton("📦 Все базы", callback_data="export_db:all")]]
    
    for db_name in available_databases:
        keyboard.append([InlineKeyboardButton(f"🏛 {db_name}", callback_data=f"export_db:{db_name}")])
    
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data="back_to_period")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.callback_query.edit_message_text(
        "📂 <b>Выбор базы данных</b>\n\n"
        "Выберите базу данных для экспорта:",
        parse_mode='HTML',
        reply_markup=reply_markup
    )
    
    return States.DB_SELECTION_MENU


@handle_errors
async def handle_export_database(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обрабатывает выбор БД и выполняет экспорт
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: Следующее состояние
    """
    query = update.callback_query
    await query.answer()
    
    callback_data = query.data
    
    if callback_data.startswith("export_db:"):
        db_name = callback_data.split(":")[1]
        
        export_type = context.user_data.get('export_type', 'unfound')
        period = context.user_data.get('export_period', 'full')
        
        # Выполняем экспорт
        await query.edit_message_text("⏳ Подготовка данных для экспорта...")
        
        try:
            only_new = (period == 'new')
            db_filter = None if db_name == 'all' else db_name
            
            if export_type == 'unfound':
                # Экспорт ненайденного оборудования
                exported_files = equipment_manager.export_to_csv(
                    date_filter=None,
                    db_filter=db_filter,
                    only_new=only_new
                )
                
                unfound_csv = exported_files.get('unfound')
                
                if unfound_csv and os.path.exists(unfound_csv):
                    # Сохраняем путь к файлу
                    context.user_data['export_file'] = unfound_csv
                    
                    # Показываем опции доставки
                    return await show_delivery_options(update, context, unfound_csv)
                else:
                    await query.edit_message_text(
                        "❌ Нет данных для экспорта или ошибка создания файла."
                    )
                    return ConversationHandler.END
            
            elif export_type == 'transfers':
                # Экспорт перемещений
                text_file = equipment_manager.export_transfers_to_text(
                    date_filter=None,
                    db_filter=db_filter,
                    only_new=only_new
                )
                
                if text_file and os.path.exists(text_file):
                    context.user_data['export_file'] = text_file
                    return await show_delivery_options(update, context, text_file)
                else:
                    await query.edit_message_text(
                        "❌ Нет данных для экспорта или ошибка создания файла."
                    )
                    return ConversationHandler.END
            
            elif export_type == 'cartridges':
                # Экспорт замен комплектующих
                excel_file = export_cartridges_to_excel(only_new=only_new, db_filter=db_filter)
                
                if excel_file and os.path.exists(excel_file):
                    context.user_data['export_file'] = excel_file
                    return await show_delivery_options(update, context, excel_file)
                else:
                    await query.edit_message_text(
                        "❌ Нет данных для экспорта или ошибка создания файла."
                    )
                    return ConversationHandler.END

            elif export_type == 'battery':
                # Экспорт замены батареи ИБП - показываем выбор периода
                return await handle_battery_export_directly(update, context)

        except Exception as e:
            logger.error(f"Ошибка при экспорте: {e}")
            await query.edit_message_text(
                f"❌ Ошибка при экспорте: {str(e)}"
            )
            return ConversationHandler.END
    
    elif callback_data == "back_to_period":
        return await show_export_period(update, context)
    
    return States.DB_SELECTION_MENU


async def show_delivery_options(update: Update, context: ContextTypes.DEFAULT_TYPE, file_path: str) -> int:
    """
    Показывает опции доставки файла
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        file_path: Путь к экспортированному файлу
        
    Возвращает:
        int: Следующее состояние
    """
    # Получаем размер файла
    file_size = os.path.getsize(file_path)
    size_kb = round(file_size / 1024, 1)
    filename = os.path.basename(file_path)
    
    keyboard = [
        [InlineKeyboardButton("💬 Отправить в чат", callback_data="delivery:chat")],
        [InlineKeyboardButton("📧 Отправить на email", callback_data="delivery:email")],
        [InlineKeyboardButton("🔙 Назад в главное меню", callback_data="back_to_main")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.callback_query.edit_message_text(
        f"✅ <b>Файл создан</b>\n\n"
        f"📄 Имя: {filename}\n"
        f"📊 Размер: {size_kb} КБ\n\n"
        f"Выберите способ доставки:",
        parse_mode='HTML',
        reply_markup=reply_markup
    )
    
    return States.DB_SELECTION_MENU


@handle_errors
async def handle_delivery(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обрабатывает выбор способа доставки
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: ConversationHandler.END
    """
    query = update.callback_query
    await query.answer()
    
    callback_data = query.data
    
    if callback_data == "delivery:chat":
        # Отправка в чат
        file_path = context.user_data.get('export_file')
        
        if file_path and os.path.exists(file_path):
            await query.edit_message_text("📤 Отправка файла...")
            
            try:
                with open(file_path, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=query.message.chat_id,
                        document=file,
                        filename=os.path.basename(file_path),
                        caption="✅ Экспортированные данные"
                    )
                
                await context.bot.send_message(
                    chat_id=query.message.chat_id,
                    text="✅ Файл успешно отправлен!"
                )
            
            except Exception as e:
                logger.error(f"Ошибка отправки файла: {e}")
                await context.bot.send_message(
                    chat_id=query.message.chat_id,
                    text=f"❌ Ошибка отправки файла: {str(e)}"
                )
        else:
            await query.edit_message_text("❌ Файл не найден.")
        
        return ConversationHandler.END
    
    elif callback_data == "delivery:email":
        # Запрос email адреса
        await query.edit_message_text(
            "📧 <b>Отправка на email</b>\n\n"
            "Введите email адрес для отправки файла:",
            parse_mode='HTML'
        )
        
        return States.UNFOUND_EMPLOYEE_INPUT  # Используем существующее состояние для ввода текста
    
    elif callback_data == "back_to_main":
        await query.edit_message_text("✅ Возврат в главное меню")
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=Messages.MAIN_MENU,
            reply_markup=create_main_menu_keyboard()
        )
        return ConversationHandler.END
    
    return ConversationHandler.END


@handle_errors
async def handle_email_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обрабатывает ввод email адреса и отправляет файл
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: ConversationHandler.END
    """
    email = update.message.text.strip()
    
    # Простая валидация email
    import re
    if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
        await update.message.reply_text(
            "❌ Некорректный email адрес. Попробуйте еще раз."
        )
        return States.UNFOUND_EMPLOYEE_INPUT
    
    file_path = context.user_data.get('export_file')
    export_type = context.user_data.get('export_type', 'export')
    
    if not file_path or not os.path.exists(file_path):
        await update.message.reply_text("❌ Файл не найден.")
        return ConversationHandler.END
    
    await update.message.reply_text("📤 Отправка на email...")
    
    try:
        # Отправляем email
        success = send_export_email(
            recipient=email,
            csv_files={export_type: file_path},
            subject="Экспорт данных IT-invent",
            body="Во вложении экспортированные данные из системы IT-invent."
        )
        
        if success:
            await update.message.reply_text(
                f"✅ Файл успешно отправлен на {email}!"
            )
        else:
            await update.message.reply_text(
                "❌ Ошибка отправки email. Проверьте настройки SMTP."
            )
    
    except Exception as e:
        logger.error(f"Ошибка отправки email: {e}")
        await update.message.reply_text(
            f"❌ Ошибка отправки email: {str(e)}"
        )
    
    return ConversationHandler.END



def export_cartridges_to_excel(only_new: bool = False, db_filter: str = None) -> str:
    """
    Экспортирует замены комплектующих МФУ в Excel

    Параметры:
        only_new: Экспортировать только новые записи
        db_filter: Фильтр по базе данных (None = все базы)

    Возвращает:
        str: Путь к созданному файлу
    """
    import json
    import pandas as pd
    from pathlib import Path
    from datetime import datetime
    
    try:
        file_path = Path("data/cartridge_replacements.json")
        
        if not file_path.exists():
            return None
        
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if not data:
            return None
        
        # Фильтруем по БД если указан фильтр
        if db_filter:
            data = [item for item in data if item.get('db_name') == db_filter]
        
        if not data:
            return None
        
        # Создаем DataFrame
        df = pd.DataFrame(data)
        
        # Добавляем db_name если отсутствует (для старых записей)
        if 'db_name' not in df.columns:
            df['db_name'] = 'ITINVENT'
        
        # Форматируем timestamp
        if 'timestamp' in df.columns:
            df['timestamp'] = pd.to_datetime(df['timestamp']).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Создаем отображаемые имена компонентов
        def get_component_name(component_type):
            names = {
                'cartridge': 'Картридж',
                'fuser': 'Фьюзер (печка)',
                'drum': 'Фотобарабан',  # Обратная совместимость
                'photoconductor': 'Фотобарабан',
                'waste_toner': 'Контейнер отраб. тонера',
                'transfer_belt': 'Трансферный ремень'
            }
            return names.get(component_type, component_type)

        # Если есть component_type, используем новые поля
        if 'component_type' in df.columns:
            # Создаем колонку с отображаемыми именами компонентов
            df['Компонент'] = df['component_type'].apply(get_component_name)

            # Переименовываем колонки
            column_names = {
                'branch': 'Филиал',
                'location': 'Локация',
                'printer_model': 'Модель принтера',
                'component_type': 'Тип компонента',
                'component_color': 'Цвет',
                'db_name': 'База данных',
                'timestamp': 'Дата и время'
            }
            df = df.rename(columns=column_names)

            # Упорядочиваем колонки для нового формата
            desired_order = ['Дата и время', 'База данных', 'Филиал', 'Локация', 'Модель принтера', 'Тип компонента', 'Компонент', 'Цвет']
        else:
            # Старый формат для обратной совместимости
            column_names = {
                'branch': 'Филиал',
                'location': 'Локация',
                'printer_model': 'Модель принтера',
                'cartridge_color': 'Цвет картриджа',
                'db_name': 'База данных',
                'timestamp': 'Дата и время'
            }
            df = df.rename(columns=column_names)

            # Упорядочиваем колонки для старого формата
            desired_order = ['Дата и время', 'База данных', 'Филиал', 'Локация', 'Модель принтера', 'Цвет картриджа']

        existing_cols = [col for col in desired_order if col in df.columns]
        df = df[existing_cols]
        
        # Создаем имя файла
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"exports/component_replacements_{timestamp}.xlsx"
        
        # Создаем директорию если не существует
        Path("exports").mkdir(exist_ok=True)
        
        # Сохраняем в Excel
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        logger.info(f"Экспорт замен комплектующих завершен: {output_file}")
        return output_file
        
    except Exception as e:
        logger.error(f"Ошибка экспорта замен комплектующих: {e}")
        return None


@handle_errors
async def handle_battery_export_directly(update: Update, context: ContextTypes.DEFAULT_TYPE, period: str) -> int:
    """
    Обрабатывает прямой экспорт замены батареи без выбора базы данных

    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        period: Выбранный период

    Возвращает:
        int: Следующее состояние
    """
    query = update.callback_query
    await query.edit_message_text("⏳ Анализ данных о заменах батареи ИБП...")

    try:
        # Выполняем экспорт с структурированием по филиалам
        excel_file = await export_battery_to_excel_structured(period=period, db_filter=None)

        if excel_file and os.path.exists(excel_file):
            context.user_data['export_file'] = excel_file
            return await show_delivery_options(update, context, excel_file)
        else:
            await query.edit_message_text(
                "❌ Нет данных для экспорта или ошибка создания файла."
            )
            return ConversationHandler.END

    except Exception as e:
        logger.error(f"Ошибка при экспорте замены батареи: {e}")
        await query.edit_message_text(
            f"❌ Ошибка при экспорте: {str(e)}"
        )
        return ConversationHandler.END


async def export_battery_to_excel_structured(period: str = "all", db_filter: str = None) -> str:
    """
    Экспорт замен батареи ИБП в Excel с разделением по филиалам

    Параметры:
        period: Период экспорта (1month, 3months, all)
        db_filter: Фильтр по базе данных (None = все базы)

    Возвращает:
        str: Путь к созданному файлу
    """
    import json
    import pandas as pd
    from pathlib import Path
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    try:
        file_path = Path("data/battery_replacements.json")

        if not file_path.exists():
            return None

        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if not data:
            return None

        # Фильтруем по БД если указан фильтр
        if db_filter:
            data = [item for item in data if item.get('db_name') == db_filter]

        if not data:
            return None

        # Фильтруем по периоду и получаем даты
        filtered_data, start_date, end_date = filter_data_by_period(data, period)

        if not filtered_data:
            return None

        # Создаем DataFrame с нужными полями
        rows = []
        logger.info(f"Обрабатываю {len(filtered_data)} записей о замене батареи")

        for item in filtered_data:
            row = {
                'Дата': item.get('timestamp', '').split('T')[0] if item.get('timestamp') else '',
                'Время': item.get('timestamp', '').split('T')[1].split('.')[0] if item.get('timestamp') else '',
                'Филиал': item.get('branch', ''),
                'Локация': item.get('location', ''),
                'Серийный номер': item.get('serial_no', ''),
                'Модель ИБП': item.get('model_name', ''),
                'Производитель': item.get('manufacturer', ''),
                'Сотрудник': item.get('employee', ''),
                'Инв. номер': item.get('inv_no', ''),
                'База данных': item.get('db_name', '')
            }
            rows.append(row)

        # Создаем DataFrame
        df = pd.DataFrame(rows)

        # Порядок колонок (с филиалом для группировки)
        df = df[['Дата', 'Время', 'Филиал', 'Локация', 'Серийный номер',
                 'Модель ИБП', 'Производитель', 'Сотрудник', 'Инв. номер', 'База данных']]

        # Сортируем по дате (новые сверху)
        df = df.sort_values('Дата', ascending=False)

        # Создаем директорию
        Path("exports").mkdir(exist_ok=True)

        # Имя файла с датами
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        if start_date and end_date:
            date_range = f"{start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}"
        else:
            date_range = "все_даты"

        output_file = f"exports/батареи_ибп_{date_range}_{timestamp}.xlsx"

        # Создаем Excel с разными листами по филиалам
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Получаем список уникальных филиалов
            branches = df['Филиал'].unique()

            # Создаем лист для каждого филиала
            for branch in branches:
                if pd.isna(branch):
                    continue

                branch_data = df[df['Филиал'] == branch].copy()
                # Удаляем колонку Филиал, т.к. она уже в названии листа
                branch_data = branch_data.drop('Филиал', axis=1)

                # Имя листа (ограничение 31 символ)
                sheet_name = str(branch)[:31]
                branch_data.to_excel(writer, sheet_name=sheet_name, index=False)

                # Форматируем лист
                worksheet = writer.sheets[sheet_name]

                # Заголовок листа
                title_cell = worksheet.cell(row=1, column=11, value=f'ФИЛИАЛ: {branch}')
                title_cell.font = Font(bold=True, size=12, color='FFFFFF')
                title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center')

                # Добавляем диапазон дат
                date_cell = worksheet.cell(row=2, column=11, value=f'Период: {date_range}')
                date_cell.font = Font(bold=True, size=10)
                date_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

                # Ширина колонок
                column_widths = {
                    'A': 12,  # Дата
                    'B': 10,  # Время
                    'C': 20,  # Локация
                    'D': 25,  # Серийный номер
                    'E': 30,  # Модель ИБП
                    'F': 15,  # Производитель
                    'G': 20,  # Сотрудник
                    'H': 12,  # Инв. номер
                    'I': 15,  # База данных
                    'J': 5,   # Резерв
                    'K': 30   # Для заголовка
                }

                for col, width in column_widths.items():
                    if col in [cell.column_letter for cell in worksheet[1]]:
                        worksheet.column_dimensions[col].width = width

            # Создаем сводный лист со всеми данными
            df_summary = df.drop('Филиал', axis=1)
            df_summary.to_excel(writer, sheet_name='Сводка', index=False)

            # Форматируем сводный лист
            worksheet = writer.sheets['Сводка']

            # Заголовок
            title_cell = worksheet.cell(row=1, column=11, value='СВОДНЫЙ ОТЧЕТ ПО ЗАМЕНЕ БАТАРЕИ ИБП')
            title_cell.font = Font(bold=True, size=12, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center')

            # Диапазон дат
            date_cell = worksheet.cell(row=2, column=11, value=f'Период: {date_range}')
            date_cell.font = Font(bold=True, size=10)
            date_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

            # Добавляем общую статистику
            stats_row = len(df_summary) + 5
            worksheet.cell(row=stats_row, column=1, value='СТАТИСТИКА').font = Font(bold=True)
            worksheet.cell(row=stats_row + 1, column=1, value=f'Всего записей: {len(df_summary)}')
            worksheet.cell(row=stats_row + 2, column=1, value=f'Филиалов: {len(branches)}')

            # Ширина колонок для сводки
            for col, width in column_widths.items():
                if col in [cell.column_letter for cell in worksheet[1]]:
                    worksheet.column_dimensions[col].width = width

        logger.info(f"Создан отчет по замене батареи ИБП с разделением по филиалам: {output_file}")
        return output_file

    except Exception as e:
        logger.error(f"Ошибка экспорта замены батареи ИБП: {e}")
        import traceback
        traceback.print_exc()
        return None


def get_period_name_ru(period: str) -> str:
    """
    Возвращает русское название периода

    Параметры:
        period: Период (1month, 3months, all)

    Возвращает:
        str: Русское название периода
    """
    period_names = {
        '1month': 'За последний месяц',
        '3months': 'За последние 3 месяца',
        'all': 'За весь период',
        'full': 'За весь период',
        'new': 'Только новые'
    }
    return period_names.get(period, period)


def filter_data_by_period(data: list, period: str) -> tuple:
    """
    Фильтрует данные по указанному периоду

    Параметры:
        data: Список записей
        period: Период (1month, 3months, all)

    Возвращает:
        tuple: (отфильтрованные данные, начальная дата, конечная дата)
    """
    from datetime import datetime, timedelta

    try:
        now = datetime.now()

        if period == "all":
            # Для всех данных ищем минимальную и максимальную даты
            if not data:
                return [], None, None

            dates = []
            for item in data:
                if 'timestamp' in item:
                    try:
                        item_date = datetime.fromisoformat(item['timestamp'].replace('Z', '+00:00'))
                        dates.append(item_date)
                    except:
                        continue

            if dates:
                return data, min(dates).date(), max(dates).date()
            else:
                return data, None, None

        # Вычисляем начальную дату
        if period == "1month":
            start_date = now - timedelta(days=30)
        elif period == "3months":
            start_date = now - timedelta(days=90)
        else:
            return data, None, None

        # Фильтруем данные
        filtered_data = []
        for item in data:
            if 'timestamp' in item:
                try:
                    item_date = datetime.fromisoformat(item['timestamp'].replace('Z', '+00:00'))
                    if item_date >= start_date:
                        filtered_data.append(item)
                except:
                    continue

        return filtered_data, start_date.date(), now.date()
    except Exception as e:
        logger.error(f"Ошибка фильтрации по периоду: {e}")
        return data, None, None


async def export_components_to_excel_structured(period: str = "all", db_filter: str = None) -> str:
    """
    Экспорт замен комплектующих в Excel с разделением по филиалам

    Параметры:
        period: Период экспорта (1month, 3months, all)
        db_filter: Фильтр по базе данных (None = все базы)

    Возвращает:
        str: Путь к созданному файлу
    """
    import json
    import pandas as pd
    from pathlib import Path
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    try:
        file_path = Path("data/cartridge_replacements.json")

        if not file_path.exists():
            return None

        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if not data:
            return None

        # Фильтруем по БД если указан фильтр
        if db_filter:
            data = [item for item in data if item.get('db_name') == db_filter]

        if not data:
            return None

        # Фильтруем по периоду и получаем даты
        filtered_data, start_date, end_date = filter_data_by_period(data, period)

        if not filtered_data:
            return None

        # Создаем DataFrame с нужными полями
        rows = []
        logger.info(f"Обрабатываю {len(filtered_data)} записей")
        for i, item in enumerate(filtered_data):
            if i < 5:  # Логируем первые 5 записей для отладки
                logger.info(f"Запись {i}: {item.get('printer_model')} - {item.get('component_type')}")
            # Базовые поля
            row = {
                'Дата': item.get('timestamp', '').split('T')[0] if item.get('timestamp') else '',
                'Время': item.get('timestamp', '').split('T')[1].split('.')[0] if item.get('timestamp') else '',
                'Филиал': item.get('branch', ''),
                'Локация': item.get('location', ''),
                'Модель принтера': item.get('printer_model', ''),
                'База данных': item.get('db_name', '')
            }

            # Определяем тип компонента и цвет
            component_model = ''
            if item.get('component_type'):
                # Новый формат
                component_type = item.get('component_type', '')
                color = item.get('component_color', '')

                # Русские названия
                type_names = {
                    'cartridge': 'Картридж',
                    'fuser': 'Фьюзер',
                    'photoconductor': 'Фотобарабан',
                    'drum': 'Фотобарабан',
                    'waste_toner': 'Контейнер',
                    'transfer_belt': 'Ремень'
                }
                row['Компонент'] = type_names.get(component_type, component_type)

                # Определяем модель компонента
                # Сначала проверяем, есть ли уже сохраненная модель
                if item_model := item.get('cartridge_model'):
                    component_model = item_model
                else:
                    # Если нет модели, ищем в базе данных
                    printer_model = item.get('printer_model', '')
                    try:
                        from bot.services.cartridge_database import cartridge_database
                        # Получаем полную информацию о принтере, а не только картриджи
                        compatibility = cartridge_database.find_printer_compatibility(printer_model)

                        # Доп. проверка для отладки
                        if component_type in ['fuser', 'photoconductor', 'drum', 'waste_toner'] and compatibility:
                            logger.debug(f"Для {printer_model} (тип: {component_type}) найдено: fuser={len(compatibility.fuser_models or [])}, drum={len(compatibility.photoconductor_models or [])}, waste={len(compatibility.waste_toner_models or [])}")

                        if compatibility:
                            logger.info(f"Найдена совместимость для {printer_model}: {component_type}")
                            # Ищем нужный тип компонента
                            if component_type == 'cartridge':
                                # Для картриджей ищем по цвету
                                color_cartridges = []
                                color_variants = [color]
                                # Пробуем разные варианты названий цветов
                                if color == 'Синий (Cyan)':
                                    color_variants.extend(['Синий', 'Cyan', 'Blue'])
                                elif color == 'Желтый (Yellow)':
                                    color_variants.extend(['Желтый', 'Yellow'])
                                elif color == 'Пурпурный (Magenta)':
                                    color_variants.extend(['Пурпурный', 'Magenta'])
                                elif color == 'Черный':
                                    color_variants.extend(['Black', 'Black (K)'])

                                for color_variant in color_variants:
                                    found = [cart for cart in compatibility.compatible_models if cart.color == color_variant]
                                    if found:
                                        color_cartridges.extend(found)
                                        break

                                if color_cartridges:
                                    # Берем первую совместимую модель
                                    component_model = color_cartridges[0].model
                                else:
                                    # Если нет нужного цвета, берем любую модель
                                    if compatibility.compatible_models:
                                        component_model = compatibility.compatible_models[0].model
                                    else:
                                        component_model = 'Картридж'
                            elif component_type == 'fuser':
                                # Используем fuser_models из базы данных
                                if compatibility.fuser_models and len(compatibility.fuser_models) > 0:
                                    component_model = compatibility.fuser_models[0]
                                    logger.info(f"Найден фьюзер для {printer_model}: {component_model}")
                                else:
                                    # Если нет в базе, используем базовые модели
                                    logger.warning(f"Не найдены фьюзеры для {printer_model}")
                                    if 'Xerox' in printer_model.upper():
                                        component_model = 'RM1-6405'
                                    elif 'HP' in printer_model.upper():
                                        component_model = 'RM1-4353'
                                    elif 'Kyocera' in printer_model.upper():
                                        component_model = 'FK-580'
                                    else:
                                        component_model = 'Фьюзер'
                            elif component_type in ['photoconductor', 'drum']:
                                # Используем photoconductor_models из базы данных
                                if compatibility.photoconductor_models and len(compatibility.photoconductor_models) > 0:
                                    component_model = compatibility.photoconductor_models[0]
                                    logger.info(f"Найден фотобарабан для {printer_model}: {component_model}")
                                else:
                                    # Если нет в базе, используем базовые модели
                                    logger.warning(f"Не найдены фотобарабаны для {printer_model}")
                                    if 'Xerox' in printer_model.upper():
                                        component_model = '115R00090'
                                    elif 'HP' in printer_model.upper():
                                        component_model = 'CE390A'
                                    elif 'Kyocera' in printer_model.upper():
                                        component_model = 'DK-580'
                                    else:
                                        component_model = 'Фотобарабан'
                            elif component_type == 'waste_toner':
                                # Используем waste_toner_models из базы данных
                                if hasattr(compatibility, 'waste_toner_models') and compatibility.waste_toner_models:
                                    component_model = compatibility.waste_toner_models[0]
                                else:
                                    component_model = 'Контейнер отраб. тонера'
                            elif component_type == 'transfer_belt':
                                # Ищем трансферные ремни (если есть)
                                if hasattr(compatibility, 'transfer_belt_models') and compatibility.transfer_belt_models:
                                    component_model = compatibility.transfer_belt_models[0]
                                else:
                                    component_model = 'Трансферный ремень'
                            else:
                                # Для других типов
                                component_model = item.get('component_type', '')
                    except Exception as e:
                        logger.error(f"Error getting component model: {e}")
                        logger.error(f"Printer model: {printer_model}, Component type: {component_type}")
                        import traceback
                        traceback.print_exc()
                        component_model = 'Ошибка поиска модели'

            elif item.get('cartridge_color'):
                # Старый формат (только картриджи)
                row['Компонент'] = 'Картридж'
                color = item.get('cartridge_color', '')
                # Для старого формата тоже используем базу данных
                # Сначала проверяем, есть ли сохраненная модель
                if item_model := item.get('cartridge_model'):
                    component_model = item_model
                else:
                    printer_model = item.get('printer_model', '')
                    try:
                        from bot.services.cartridge_database import cartridge_database
                        # Для старого формата тоже используем полную информацию
                        compatibility = cartridge_database.find_printer_compatibility(printer_model)

                        if compatibility and compatibility.compatible_models:
                            # Берем первую совместимую модель (обычно картридж)
                            component_model = compatibility.compatible_models[0].model
                        else:
                            # Если нет в базе, пробуем определить по названию
                            if 'Xerox Versalink' in printer_model:
                                component_model = 'Xerox 106R02773'
                            elif 'Kyocera' in printer_model:
                                component_model = 'Kyocera TK-3172'
                            else:
                                component_model = 'Картридж'
                    except:
                        component_model = 'Картридж'
            else:
                row['Компонент'] = 'Неизвестно'
                color = ''
                component_model = ''

            row['Цвет'] = color
            row['Модель'] = component_model

            # Добавляем строчку
            rows.append(row)

        # Создаем DataFrame
        df = pd.DataFrame(rows)

        # Порядок колонок (с филиалом для группировки)
        df = df[['Дата', 'Время', 'Филиал', 'Локация', 'Модель принтера', 'Компонент', 'Модель', 'Цвет', 'База данных']]

        # Сортируем по дате (новые сверху)
        df = df.sort_values('Дата', ascending=False)

        # Создаем директорию
        Path("exports").mkdir(exist_ok=True)

        # Имя файла с датами
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        if start_date and end_date:
            date_range = f"{start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}"
        else:
            date_range = "все_даты"

        output_file = f"exports/комплектующие_{date_range}_{timestamp}.xlsx"

        # Создаем Excel с разными листами по филиалам
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Получаем список уникальных филиалов
            branches = df['Филиал'].unique()

            # Создаем лист для каждого филиала
            for branch in branches:
                if pd.isna(branch):
                    continue

                branch_data = df[df['Филиал'] == branch].copy()
                # Удаляем колонку Филиал, т.к. она уже в названии листа
                branch_data = branch_data.drop('Филиал', axis=1)

                # Имя листа (ограничение 31 символ)
                sheet_name = str(branch)[:31]
                branch_data.to_excel(writer, sheet_name=sheet_name, index=False)

                # Форматируем лист
                worksheet = writer.sheets[sheet_name]

                # Заголовок листа
                title_cell = worksheet.cell(row=1, column=9, value=f'ФИЛИАЛ: {branch}')
                title_cell.font = Font(bold=True, size=12, color='FFFFFF')
                title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center')

                # Добавляем диапазон дат
                date_cell = worksheet.cell(row=2, column=9, value=f'Период: {date_range}')
                date_cell.font = Font(bold=True, size=10)
                date_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

                # Ширина колонок
                column_widths = {
                    'A': 12,  # Дата
                    'B': 10,  # Время
                    'C': 35,  # Филиал (увеличили для полного текста)
                    'D': 15,  # Локация
                    'E': 30,  # Модель принтера
                    'F': 15,  # Компонент
                    'G': 20,  # Модель (новая колонка)
                    'H': 15,  # Цвет
                    'I': 15   # База данных
                }

                for col, width in column_widths.items():
                    if col in [cell.column_letter for cell in worksheet[1]]:
                        worksheet.column_dimensions[col].width = width

            # Создаем сводный лист со всеми данными
            df_summary = df.drop('Филиал', axis=1)
            df_summary.to_excel(writer, sheet_name='Сводка', index=False)

            # Форматируем сводный лист
            worksheet = writer.sheets['Сводка']

            # Заголовок
            title_cell = worksheet.cell(row=1, column=9, value='СВОДНЫЙ ОТЧЕТ')
            title_cell.font = Font(bold=True, size=12, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center')

            # Диапазон дат
            date_cell = worksheet.cell(row=2, column=9, value=f'Период: {date_range}')
            date_cell.font = Font(bold=True, size=10)
            date_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

            # Добавляем общую статистику
            stats_row = len(df_summary) + 5
            worksheet.cell(row=stats_row, column=1, value='СТАТИСТИКА').font = Font(bold=True)
            worksheet.cell(row=stats_row, column=1, value=f'Всего записей: {len(df_summary)}')
            worksheet.cell(row=stats_row + 1, column=1, value=f'Филиалов: {len(branches)}')

            # Ширина колонок для сводки
            for col, width in column_widths.items():
                if col in [cell.column_letter for cell in worksheet[1]]:
                    worksheet.column_dimensions[col].width = width

        logger.info(f"Создан отчет с разделением по филиалам: {output_file}")
        return output_file

    except Exception as e:
        logger.error(f"Ошибка экспорта замен комплектующих: {e}")
        import traceback
        traceback.print_exc()
        return None


