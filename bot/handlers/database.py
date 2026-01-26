#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Обработчики управления базами данных
Выбор БД, просмотр по типам оборудования, статистика.
"""
import logging
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ContextTypes, ConversationHandler

from bot.config import States, Messages
from bot.utils.decorators import require_user_access, handle_errors
from bot.utils.keyboards import create_main_menu_keyboard
from bot.utils.pagination import paginate_results
from bot.utils.formatters import format_equipment_info
from database_manager import database_manager
from universal_database import UniversalInventoryDB

logger = logging.getLogger(__name__)


@require_user_access
async def show_database_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Отображает меню управления базами данных
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: Состояние DB_SELECTION_MENU
    """
    # Получаем текущую БД пользователя
    user_id = update.effective_user.id
    current_db = database_manager.get_user_database(user_id)
    
    # Получаем список доступных баз данных
    available_databases = database_manager.get_available_databases()
    
    # Создаем клавиатуру
    keyboard = []
    
    # Кнопки выбора БД
    for db_name in available_databases:
        is_selected = '✅' if current_db == db_name else ''
        keyboard.append([InlineKeyboardButton(
            f"📊 {db_name} {is_selected}",
            callback_data=f"select_db:{db_name}"
        )])
    
    # Дополнительные опции
    keyboard.extend([
        [InlineKeyboardButton("🔧 Просмотр по типу оборудования", callback_data="equipment_types_menu")],
        [InlineKeyboardButton("📤 Экспорт базы в CSV", callback_data="export_db_menu")],
        [InlineKeyboardButton("🔙 Назад в главное меню", callback_data="back_to_main")]
    ])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        f"🗄️ <b>Управление базами данных</b>\n\n"
        f"📋 <b>Текущая база:</b> {current_db}\n\n"
        f"Выберите действие:",
        parse_mode='HTML',
        reply_markup=reply_markup
    )
    
    return States.DB_SELECTION_MENU


@handle_errors
async def handle_database_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обрабатывает callback'и от меню управления БД
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: Следующее состояние
    """
    query = update.callback_query
    await query.answer()
    
    callback_data = query.data
    logger.info(f"[DATABASE] Получен callback: {callback_data}")
    
    if callback_data.startswith("select_db:"):
        # Выбор базы данных
        db_name = callback_data.split(":")[1]
        user_id = update.effective_user.id
        
        # Сохраняем выбор
        database_manager.set_user_database(user_id, db_name)
        context.user_data['selected_database'] = db_name
        
        await query.edit_message_text(
            f"✅ <b>База данных изменена</b>\n\n"
            f"📋 <b>Выбрана база:</b> {db_name}\n\n"
            f"Теперь все операции будут выполняться с этой базой данных.",
            parse_mode='HTML'
        )
        
        # Возвращаемся в главное меню
        import asyncio
        await asyncio.sleep(2)
        
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=Messages.MAIN_MENU,
            reply_markup=create_main_menu_keyboard()
        )
        
        return ConversationHandler.END
    
    elif callback_data == "equipment_types_menu":
        # Просмотр по типу оборудования
        return await show_equipment_types_menu(update, context)
    
    elif callback_data == "export_db_menu":
        # Экспорт базы данных
        return await show_export_database_menu(update, context)

    elif callback_data.startswith("export_db:"):
        # Обработка выбора базы для экспорта
        return await handle_export_database_callback(update, context)

    elif callback_data == "back_to_main":
        # Возврат в главное меню
        await query.edit_message_text("✅ Возврат в главное меню")
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=Messages.MAIN_MENU,
            reply_markup=create_main_menu_keyboard()
        )
        return ConversationHandler.END
    
    elif callback_data.startswith("eqtype:"):
        # Выбор типа оборудования по индексу
        try:
            type_idx = int(callback_data.split(":", 1)[1])
            equipment_types = context.user_data.get('equipment_types_list', [])
            
            if 0 <= type_idx < len(equipment_types):
                equipment_type = equipment_types[type_idx]
                return await show_equipment_by_type(update, context, equipment_type)
            else:
                await query.answer("Тип оборудования не найден", show_alert=True)
                return States.DB_VIEW_PAGINATION
        except (ValueError, IndexError) as e:
            logger.error(f"Ошибка обработки выбора типа: {e}")
            await query.answer("Ошибка обработки", show_alert=True)
            return States.DB_VIEW_PAGINATION
    
    elif callback_data == "branch_all":
        # Все филиалы
        equipment_type = context.user_data.get('equipment_type_filter')
        if equipment_type:
            return await show_equipment_by_type_and_branch(update, context, equipment_type, None)
    
    elif callback_data.startswith("branch:"):
        # Выбор конкретного филиала
        try:
            branch_idx = int(callback_data.split(":", 1)[1])
            equipment_type = context.user_data.get('equipment_type_filter')
            branches = context.user_data.get('branches_list', [])
            
            if equipment_type and 0 <= branch_idx < len(branches):
                branch_name = branches[branch_idx].get('BRANCH_NAME')
                return await show_equipment_by_type_and_branch(update, context, equipment_type, branch_name)
        except (ValueError, IndexError) as e:
            logger.error(f"Ошибка обработки выбора филиала: {e}")
            await query.answer("Ошибка обработки", show_alert=True)
            return States.DB_VIEW_PAGINATION
    
    elif callback_data == "types_prev":
        # Предыдущая страница типов
        current_page = context.user_data.get('equipment_types_page', 0)
        if current_page > 0:
            context.user_data['equipment_types_page'] = current_page - 1
        return await show_equipment_types_menu(update, context)
    
    elif callback_data == "types_next":
        # Следующая страница типов
        equipment_types = context.user_data.get('equipment_types_list', [])
        current_page = context.user_data.get('equipment_types_page', 0)
        items_per_page = 8
        total_pages = (len(equipment_types) + items_per_page - 1) // items_per_page
        
        if current_page < total_pages - 1:
            context.user_data['equipment_types_page'] = current_page + 1
        return await show_equipment_types_menu(update, context)
    
    elif callback_data == "back_to_db_menu":
        # Возврат в меню БД
        return await show_database_menu_from_callback(update, context)
    
    return States.DB_SELECTION_MENU


async def show_equipment_types_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Отображает меню выбора типа оборудования с пагинацией
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: Состояние DB_VIEW_PAGINATION
    """
    try:
        user_id = update.effective_user.id
        current_db = database_manager.get_user_database(user_id)
        
        # Создаем подключение к БД
        db = database_manager.create_database_connection(user_id)
        if not db:
            await update.callback_query.edit_message_text(
                "❌ Ошибка подключения к базе данных",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Назад", callback_data="back_to_db_menu")
                ]])
            )
            return States.DB_SELECTION_MENU
        
        # Получаем типы оборудования
        all_equipment_types = db.get_equipment_types()
        db.close_connection()
        
        if not all_equipment_types:
            await update.callback_query.edit_message_text(
                "❌ Типы оборудования не найдены",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Назад", callback_data="back_to_db_menu")
                ]])
            )
            return States.DB_SELECTION_MENU
        
        # Приоритетные типы
        priority_equipment_types = ['МФУ', 'Монитор', 'Системный блок', 'ИБП']
        
        # Сортируем: сначала приоритетные, потом остальные
        priority_types = []
        other_types = []
        
        for priority_type in priority_equipment_types:
            for db_type in all_equipment_types:
                if priority_type.upper() in db_type.upper():
                    if db_type not in priority_types:
                        priority_types.append(db_type)
        
        for db_type in all_equipment_types:
            if db_type not in priority_types and db_type not in other_types:
                other_types.append(db_type)
        
        equipment_types = priority_types + other_types
        
        # Сохраняем в контексте
        context.user_data['equipment_types_list'] = equipment_types
        current_page = context.user_data.get('equipment_types_page', 0)
        
        # Создаем клавиатуру с пагинацией
        reply_markup = generate_equipment_types_keyboard(equipment_types, current_page)
        
        await update.callback_query.edit_message_text(
            f"🔧 <b>Выбор типа оборудования</b>\n\n"
            f"📊 <b>База данных:</b> {current_db}\n"
            f"📋 <b>Найдено типов:</b> {len(equipment_types)}\n\n"
            f"Выберите тип оборудования для просмотра:",
            parse_mode='HTML',
            reply_markup=reply_markup
        )
        
        return States.DB_VIEW_PAGINATION
        
    except Exception as e:
        logger.error(f"Ошибка при отображении типов оборудования: {e}")
        await update.callback_query.edit_message_text(
            f"❌ Ошибка: {str(e)}",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("🔙 Назад", callback_data="back_to_db_menu")
            ]])
        )
        return States.DB_SELECTION_MENU


def generate_equipment_types_keyboard(equipment_types: list, current_page: int = 0) -> InlineKeyboardMarkup:
    """
    Генерирует клавиатуру с типами оборудования и пагинацией
    
    Параметры:
        equipment_types: Список типов оборудования
        current_page: Текущая страница
        
    Возвращает:
        InlineKeyboardMarkup: Клавиатура с кнопками
    """
    items_per_page = 8
    total_pages = (len(equipment_types) + items_per_page - 1) // items_per_page
    
    start_idx = current_page * items_per_page
    end_idx = start_idx + items_per_page
    page_types = equipment_types[start_idx:end_idx]
    
    keyboard = []
    
    # Кнопки типов оборудования (используем индексы для коротких callback)
    for idx, eq_type in enumerate(page_types):
        global_idx = start_idx + idx
        # Ограничиваем длину callback_data используя индекс
        keyboard.append([InlineKeyboardButton(
            f"🔧 {eq_type}",
            callback_data=f"eqtype:{global_idx}"
        )])
    
    # Навигация
    nav_buttons = []
    if current_page > 0:
        nav_buttons.append(InlineKeyboardButton("◀️ Назад", callback_data="types_prev"))
    
    if total_pages > 1:
        nav_buttons.append(InlineKeyboardButton(
            f"📄 {current_page + 1}/{total_pages}",
            callback_data="types_page_info"
        ))
    
    if current_page < total_pages - 1:
        nav_buttons.append(InlineKeyboardButton("Вперед ▶️", callback_data="types_next"))
    
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    # Кнопка возврата
    keyboard.append([InlineKeyboardButton("🔙 Назад в меню БД", callback_data="back_to_db_menu")])
    
    return InlineKeyboardMarkup(keyboard)


async def show_equipment_by_type(update: Update, context: ContextTypes.DEFAULT_TYPE, equipment_type: str) -> int:
    """
    Отображает выбор филиала для оборудования выбранного типа
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        equipment_type: Тип оборудования
        
    Возвращает:
        int: Состояние DB_VIEW_PAGINATION
    """
    try:
        # Сохраняем выбранный тип
        context.user_data['equipment_type_filter'] = equipment_type
        
        user_id = update.effective_user.id
        current_db = database_manager.get_user_database(user_id)
        
        # Создаем подключение к БД
        db = database_manager.create_database_connection(user_id)
        if not db:
            await update.callback_query.edit_message_text(
                "❌ Ошибка подключения к базе данных",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Назад", callback_data="equipment_types_menu")
                ]])
            )
            return States.DB_VIEW_PAGINATION
        
        # Получаем список филиалов
        branches = db.get_branches()
        db.close_connection()
        
        # Сохраняем список филиалов
        context.user_data['branches_list'] = branches
        
        if not branches:
            # Если филиалов нет, показываем все оборудование
            return await show_equipment_by_type_and_branch(update, context, equipment_type, None)
        
        # Создаем клавиатуру с филиалами
        keyboard = []
        
        # Кнопка "Все филиалы"
        keyboard.append([InlineKeyboardButton(
            "🌐 Все филиалы",
            callback_data="branch_all"
        )])
        
        # Кнопки для каждого филиала (используем короткие индексы)
        for idx, branch in enumerate(branches):
            branch_name = branch.get('BRANCH_NAME', 'Неизвестен')
            keyboard.append([InlineKeyboardButton(
                f"🏢 {branch_name}",
                callback_data=f"branch:{idx}"
            )])
        
        # Кнопки возврата
        keyboard.extend([
            [InlineKeyboardButton("🔙 К типам оборудования", callback_data="equipment_types_menu")],
            [InlineKeyboardButton("🏠 В меню БД", callback_data="back_to_db_menu")]
        ])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.callback_query.edit_message_text(
            f"🏢 <b>Выберите филиал для типа оборудования: {equipment_type}</b>\n\n"
            f"Выберите филиал, чтобы увидеть оборудование только этого филиала:",
            parse_mode='HTML',
            reply_markup=reply_markup
        )
        
        return States.DB_VIEW_PAGINATION
        
    except Exception as e:
        logger.error(f"Ошибка при получении филиалов для типа '{equipment_type}': {e}")
        await update.callback_query.edit_message_text(
            f"❌ Ошибка: {str(e)}",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("🔙 Назад", callback_data="equipment_types_menu")
            ]])
        )
        return States.DB_VIEW_PAGINATION


async def show_equipment_by_type_and_branch(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                                            equipment_type: str, branch_name: str = None) -> int:
    """
    Отображает оборудование выбранного типа и филиала
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        equipment_type: Тип оборудования
        branch_name: Название филиала (None для всех)
        
    Возвращает:
        int: Состояние DB_VIEW_PAGINATION
    """
    try:
        user_id = update.effective_user.id
        
        # Создаем подключение к БД
        db = database_manager.create_database_connection(user_id)
        if not db:
            await update.callback_query.edit_message_text(
                "❌ Ошибка подключения к базе данных",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Назад", callback_data="equipment_types_menu")
                ]])
            )
            return States.DB_VIEW_PAGINATION
        
        # Получаем оборудование
        equipment_list = db.get_equipment_by_type(equipment_type, branch_name=branch_name)
        db.close_connection()
        
        if not equipment_list:
            await update.callback_query.edit_message_text(
                f"❌ Оборудование типа <b>{equipment_type}</b> не найдено",
                parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Назад", callback_data="equipment_types_menu")
                ]])
            )
            return States.DB_VIEW_PAGINATION
        
        # Сортируем весь список по локации ДО пагинации
        def equipment_sort_key(equipment):
            """Сортировка оборудования по локации"""
            location = equipment.get('LOCATION', 'Не указана')
            if not location or location.strip() == '':
                location = 'Не указана'
            
            # "Не указана" в конце, остальные по алфавиту
            if location == 'Не указана':
                return (1, '')
            return (0, location.lower())
        
        equipment_list = sorted(equipment_list, key=equipment_sort_key)
        
        # Сохраняем отсортированный список для пагинации
        context.user_data['equipment_view_list'] = equipment_list
        context.user_data['equipment_view_page'] = 0
        context.user_data['equipment_type_filter'] = equipment_type
        context.user_data['equipment_branch_filter'] = branch_name
        
        # Показываем первую страницу
        await show_equipment_page(update, context)
        
        return States.DB_VIEW_PAGINATION
        
    except Exception as e:
        logger.error(f"Ошибка при отображении оборудования: {e}")
        await update.callback_query.edit_message_text(
            f"❌ Ошибка: {str(e)}",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("🔙 Назад", callback_data="equipment_types_menu")
            ]])
        )
        return States.DB_VIEW_PAGINATION


async def show_equipment_page(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Отображает страницу с оборудованием
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
    """
    equipment_list = context.user_data.get('equipment_view_list', [])
    current_page = context.user_data.get('equipment_view_page', 0)
    equipment_type = context.user_data.get('equipment_type_filter', 'Неизвестен')
    branch_filter = context.user_data.get('equipment_branch_filter')
    
    from bot.config import PaginationConfig
    config = PaginationConfig()
    
    # Пагинация
    page_items, total_pages, has_prev, has_next = paginate_results(
        equipment_list,
        current_page,
        config.items_per_page
    )
    
    # Формируем сообщение
    message_lines = [
        f"🔧 <b>Тип:</b> {equipment_type}",
    ]
    
    if branch_filter:
        message_lines.append(f"🏢 <b>Филиал:</b> {branch_filter}")
    else:
        message_lines.append(f"🏢 <b>Филиал:</b> Все")
    
    message_lines.extend([
        f"📋 <b>Найдено единиц:</b> {len(equipment_list)}",
        f"📄 <b>Страница:</b> {current_page + 1} из {total_pages}\n"
    ])
    
    # Выводим оборудование (уже отсортировано при загрузке)
    for i, equipment in enumerate(page_items, 1):
        item_num = current_page * config.items_per_page + i
        message_lines.append(f"<b>{item_num}.</b>")
        message_lines.append(format_equipment_info(equipment))
        message_lines.append("")
    
    message_text = "\n".join(message_lines)
    
    # Создаем клавиатуру навигации
    keyboard = []
    nav_buttons = []
    
    if has_prev:
        nav_buttons.append(InlineKeyboardButton("◀️ Назад", callback_data="eq_prev"))
    
    nav_buttons.append(InlineKeyboardButton(f"📄 {current_page+1}/{total_pages}", callback_data="page_info"))
    
    if has_next:
        nav_buttons.append(InlineKeyboardButton("Вперед ▶️", callback_data="eq_next"))
    
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    keyboard.append([InlineKeyboardButton("🔙 К типам оборудования", callback_data="equipment_types_menu")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.callback_query.edit_message_text(
        message_text,
        parse_mode='HTML',
        reply_markup=reply_markup
    )

    return States.DB_VIEW_PAGINATION


@handle_errors
async def handle_equipment_pagination(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обрабатывает навигацию по страницам оборудования
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: Состояние DB_VIEW_PAGINATION
    """
    query = update.callback_query
    await query.answer()
    
    callback_data = query.data
    
    if callback_data == "eq_prev":
        current_page = context.user_data.get('equipment_view_page', 0)
        if current_page > 0:
            context.user_data['equipment_view_page'] = current_page - 1
        return await show_equipment_page(update, context)
    
    elif callback_data == "eq_next":
        equipment_list = context.user_data.get('equipment_view_list', [])
        current_page = context.user_data.get('equipment_view_page', 0)

        from bot.config import PaginationConfig
        config = PaginationConfig()
        total_pages = (len(equipment_list) + config.items_per_page - 1) // config.items_per_page

        if current_page < total_pages - 1:
            context.user_data['equipment_view_page'] = current_page + 1
        return await show_equipment_page(update, context)

    else:
        # Обработка неизвестных callback_data
        logger.warning(f"Неизвестный callback в handle_equipment_pagination: {callback_data}")
        await query.answer("Неизвестная команда", show_alert=True)

    return States.DB_VIEW_PAGINATION


async def show_database_statistics(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Отображает статистику баз данных
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: Состояние DB_SELECTION_MENU
    """
    try:
        # Получаем статистику всех БД
        all_stats = database_manager.get_all_statistics()
        
        if not all_stats:
            await update.callback_query.edit_message_text(
                "❌ Не удалось получить статистику баз данных",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Назад", callback_data="back_to_db_menu")
                ]])
            )
            return States.DB_SELECTION_MENU
        
        # Формируем сообщение
        message_lines = ["📈 <b>Статистика баз данных</b>\n"]
        
        for stat in all_stats:
            db_name = stat.get('database', 'Неизвестно')
            total_items = stat.get('total_items', 0)
            total_employees = stat.get('total_employees', 0)
            
            message_lines.append(f"📊 <b>{db_name}</b>")
            message_lines.append(f"  • Единиц оборудования: {total_items}")
            message_lines.append(f"  • Сотрудников: {total_employees}\n")
        
        message_text = "\n".join(message_lines)
        
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="back_to_db_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.callback_query.edit_message_text(
            message_text,
            parse_mode='HTML',
            reply_markup=reply_markup
        )
        
        return States.DB_SELECTION_MENU
        
    except Exception as e:
        logger.error(f"Ошибка при получении статистики: {e}")
        await update.callback_query.edit_message_text(
            f"❌ Ошибка: {str(e)}",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("🔙 Назад", callback_data="back_to_db_menu")
            ]])
        )
        return States.DB_SELECTION_MENU


async def show_database_menu_from_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Отображает меню БД из callback'а
    
    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения
        
    Возвращает:
        int: Состояние DB_SELECTION_MENU
    """
    user_id = update.effective_user.id
    current_db = database_manager.get_user_database(user_id)
    
    available_databases = database_manager.get_available_databases()
    
    keyboard = []
    
    for db_name in available_databases:
        is_selected = '✅' if current_db == db_name else ''
        keyboard.append([InlineKeyboardButton(
            f"📊 {db_name} {is_selected}",
            callback_data=f"select_db:{db_name}"
        )])
    
    keyboard.extend([
        [InlineKeyboardButton("🔧 Просмотр по типу оборудования", callback_data="equipment_types_menu")],
        [InlineKeyboardButton("📤 Экспорт базы в CSV", callback_data="export_db_menu")],
        [InlineKeyboardButton("🔙 Назад в главное меню", callback_data="back_to_main")]
    ])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.callback_query.edit_message_text(
        f"🗄️ <b>Управление базами данных</b>\n\n"
        f"📋 <b>Текущая база:</b> {current_db}\n\n"
        f"Выберите действие:",
        parse_mode='HTML',
        reply_markup=reply_markup
    )

    return States.DB_SELECTION_MENU


@handle_errors
async def show_export_database_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Отображает меню выбора базы данных для экспорта в CSV

    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения

    Возвращает:
        int: Состояние DB_SELECTION_MENU
    """
    # Получаем список доступных баз данных
    available_databases = database_manager.get_available_databases()

    # Создаем клавиатуру с кнопками для каждой базы
    keyboard = []

    for db_name in available_databases:
        keyboard.append([InlineKeyboardButton(
            f"📊 {db_name}",
            callback_data=f"export_db:{db_name}"
        )])

    # Кнопка назад
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data="back_to_db_menu")])

    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.callback_query.edit_message_text(
        "📤 <b>Экспорт базы данных в CSV</b>\n\n"
        "Выберите базу данных для экспорта:",
        parse_mode='HTML',
        reply_markup=reply_markup
    )

    return States.DB_SELECTION_MENU


@handle_errors
async def handle_export_database_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обрабатывает выбор базы данных для экспорта

    Параметры:
        update: Объект обновления от Telegram API
        context: Контекст выполнения

    Возвращает:
        int: Состояние DB_SELECTION_MENU
    """
    query = update.callback_query
    await query.answer()

    callback_data = query.data

    if callback_data.startswith("export_db:"):
        # Экспорт выбранной базы
        db_name = callback_data.split(":")[1]

        await query.edit_message_text(
            f"⏳ Экспортирую базу <b>{db_name}</b> в Excel...\n\n"
            f"Это может занять некоторое время.",
            parse_mode='HTML'
        )

        # Экспортируем базу
        excel_path = await export_database_to_csv(db_name)

        if excel_path:
            # Отправляем файл
            try:
                import os
                record_count = count_records_in_excel(excel_path)
                filename = os.path.basename(excel_path)

                with open(excel_path, 'rb') as excel_file:
                    await context.bot.send_document(
                        chat_id=query.message.chat_id,
                        document=excel_file,
                        filename=filename,
                        caption=f"✅ Экспорт базы <b>{db_name}</b>\n\nВсего записей: {record_count}",
                        parse_mode='HTML'
                    )
            except Exception as e:
                logger.error(f"Ошибка отправки Excel файла: {e}")
                await context.bot.send_message(
                    chat_id=query.message.chat_id,
                    text=f"❌ Ошибка при отправке файла: {e}"
                )
        else:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="❌ Ошибка при экспорте базы данных"
            )

        # Возвращаемся в меню БД
        return await show_database_menu_from_callback(update, context)

    return States.DB_SELECTION_MENU


async def export_database_to_csv(db_name: str) -> str:
    """
    Экспортирует все данные из базы данных в Excel файл с группировкой

    Параметры:
        db_name: Имя базы данных для экспорта

    Возвращает:
        str: Путь к созданному Excel файлу или None в случае ошибки
    """
    import os
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        # Получаем конфиг базы данных
        config = database_manager.get_database_config(db_name)
        if not config:
            logger.error(f"Конфиг базы данных {db_name} не найден")
            return None

        # Создаем подключение к базе
        db = UniversalInventoryDB(config)
        conn = db._get_connection()
        cursor = conn.cursor()

        # SQL запрос для получения всех данных (с LOCATIONS и BRANCHES)
        query_with_location = """
        SELECT
            i.INV_NO,
            o.OWNER_DISPLAY_NAME as EMPLOYEE_NAME,
            t.TYPE_NAME,
            i.SERIAL_NO,
            i.HW_SERIAL_NO,
            i.PART_NO,
            m.MODEL_NAME,
            v.VENDOR_NAME as MANUFACTURER,
            l.DESCR as LOCATION,
            i.EMPL_NO,
            o.OWNER_DEPT as EMPLOYEE_DEPT,
            b.BRANCH_NAME as BRANCH,
            s.DESCR as STATUS,
            i.DESCR as DESCRIPTION
        FROM ITEMS i
        LEFT JOIN CI_TYPES t ON i.CI_TYPE = t.CI_TYPE AND i.TYPE_NO = t.TYPE_NO
        LEFT JOIN CI_MODELS m ON i.MODEL_NO = m.MODEL_NO AND i.CI_TYPE = m.CI_TYPE
        LEFT JOIN VENDORS v ON m.VENDOR_NO = v.VENDOR_NO
        LEFT JOIN LOCATIONS l ON i.LOC_NO = l.LOC_NO
        LEFT JOIN OWNERS o ON i.EMPL_NO = o.OWNER_NO
        LEFT JOIN BRANCHES b ON i.BRANCH_NO = b.BRANCH_NO
        LEFT JOIN STATUS s ON i.STATUS_NO = s.STATUS_NO
        ORDER BY b.BRANCH_NAME, l.DESCR, i.SERIAL_NO
        """

        # Fallback запрос без BRANCHES и LOCATIONS
        query_without_location = """
        SELECT
            i.INV_NO,
            o.OWNER_DISPLAY_NAME as EMPLOYEE_NAME,
            t.TYPE_NAME,
            i.SERIAL_NO,
            i.HW_SERIAL_NO,
            i.PART_NO,
            m.MODEL_NAME,
            v.VENDOR_NAME as MANUFACTURER,
            'Не указано' as LOCATION,
            i.EMPL_NO,
            o.OWNER_DEPT as EMPLOYEE_DEPT,
            'Не указан' as BRANCH,
            s.DESCR as STATUS,
            i.DESCR as DESCRIPTION
        FROM ITEMS i
        LEFT JOIN CI_TYPES t ON i.CI_TYPE = t.CI_TYPE AND i.TYPE_NO = t.TYPE_NO
        LEFT JOIN CI_MODELS m ON i.MODEL_NO = m.MODEL_NO AND i.CI_TYPE = m.CI_TYPE
        LEFT JOIN VENDORS v ON m.VENDOR_NO = v.VENDOR_NO
        LEFT JOIN OWNERS o ON i.EMPL_NO = o.OWNER_NO
        LEFT JOIN STATUS s ON i.STATUS_NO = s.STATUS_NO
        ORDER BY i.SERIAL_NO
        """

        # Пробуем выполнить запрос с fallback
        try:
            cursor.execute(query_with_location)
            rows = cursor.fetchall()
        except Exception as e:
            error_msg = str(e).lower()
            if 'branches' in error_msg or 'locations' in error_msg or 'permission' in error_msg or 'запрещено' in error_msg:
                logger.warning(f"Нет доступа к BRANCHES/LOCATIONS, используем fallback: {e}")
                cursor.execute(query_without_location)
                rows = cursor.fetchall()
            else:
                raise e

        cursor.close()

        if not rows:
            logger.warning(f"Нет данных для экспорта в базе {db_name}")
            return None

        # Создаем директорию для экспорта
        export_dir = "exports"
        os.makedirs(export_dir, exist_ok=True)

        # Формируем имя файла с датой
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        excel_file = os.path.join(export_dir, f"{db_name}_export_{timestamp}.xlsx")

        # Создаем Excel книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Экспорт оборудования"

        # Стили
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9")
        header_alignment = Alignment(horizontal="center", vertical="center")
        group_alignment = Alignment(horizontal="left", vertical="center")
        branch_font = Font(bold=True, size=13, color="000000")  # Чёрный текст для лучшей видимости
        branch_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7")  # Светло-синий фон
        location_font = Font(bold=True, size=11)
        location_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Добавляем заголовок с технической информацией (первые 3 строки)
        ws['A1'] = f"База данных: {db_name}"
        ws['A2'] = f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        ws['A3'] = f"Всего записей: {len(rows)}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A1:N1')
        ws.merge_cells(f'A2:N2')
        ws.merge_cells(f'A3:N3')

        # Заголовки колонок (строка 5)
        headers = [
            'Инв. №',
            'Сотрудник',
            'Тип',
            'Серийный №',
            'Апп. серийный №',
            'Партийный №',
            'Модель',
            'Производитель',
            'Местоположение',
            'Таб. №',
            'Отдел',
            'Филиал',
            'Статус',
            'Описание'
        ]

        header_row = 5
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Группируем данные по филиалу и местоположению
        current_row = header_row + 1
        _FIRST_BRANCH_SENTINEL = object()  # Уникальный объект-маркер
        current_branch = _FIRST_BRANCH_SENTINEL
        current_location = None

        for row in rows:
            # Распаковываем строку по индексам (как в SQL запросе)
            inv_no = row[0]
            employee_name = row[1]
            equipment_type = row[2]
            serial_no = row[3]
            hw_serial_no = row[4]
            part_no = row[5]
            model = row[6]
            manufacturer = row[7]
            location = row[8] or 'Не указано'
            empl_no = row[9]
            dept = row[10]
            branch = row[11] or 'Не указан'
            status = row[12]
            description = row[13]

            # Логирование для отладки
            logger.debug(f"Row: branch='{branch}', location='{location}'")

            # Новый филиал - добавляем заголовок группы
            if branch != current_branch:
                if current_branch is not _FIRST_BRANCH_SENTINEL:  # Если это НЕ первый филиал
                    current_row += 1  # Пустая строка между филиалами

                current_branch = branch
                current_location = None

                logger.info(f"Creating branch header at row {current_row}: '{branch}'")

                # Заголовок филиала (на всю ширину)
                ws.merge_cells(f'A{current_row}:N{current_row}')
                cell = ws.cell(row=current_row, column=1, value=f"🏢 {branch}")
                cell.font = branch_font
                cell.fill = branch_fill
                cell.alignment = group_alignment
                current_row += 1

            # Новое местоположение внутри филиала
            if location != current_location:
                current_location = location

                # Заголовок местоположения
                ws.merge_cells(f'A{current_row}:N{current_row}')
                cell = ws.cell(row=current_row, column=1, value=f"📍 {location}")
                cell.font = location_font
                cell.fill = location_fill
                cell.alignment = group_alignment
                current_row += 1

            # Данные оборудования - ВСЕ 14 полей в правильном порядке
            data = [
                inv_no or '',                      # 1. Инв. №
                employee_name or 'Не назначен',    # 2. Сотрудник
                equipment_type or '',              # 3. Тип
                serial_no or '',                   # 4. Серийный №
                hw_serial_no or '',                # 5. Апп. серийный №
                part_no or '',                     # 6. Партийный №
                model or '',                       # 7. Модель
                manufacturer or '',                 # 8. Производитель
                location,                          # 9. Местоположение
                empl_no or '',                     # 10. Таб. №
                dept or '',                        # 11. Отдел
                branch,                            # 12. Филиал
                status or '',                      # 13. Статус
                description or ''                  # 14. Описание
            ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=False)

                # Форматирование для важных полей
                if col_idx == 4:  # Серийный номер - жирный
                    cell.font = Font(bold=True)
                elif col_idx == 2 and employee_name and employee_name != 'Не назначен':  # Сотрудник - жёлтый фон
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC")

            current_row += 1

        # Автоматическая ширина колонок
        column_widths = [12, 20, 15, 18, 18, 12, 20, 15, 25, 10, 18, 15, 12, 35]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # Сохраняем файл
        wb.save(excel_file)

        logger.info(f"Экспорт базы {db_name} завершен: {excel_file} ({len(rows)} записей)")
        return excel_file

    except Exception as e:
        logger.error(f"Ошибка при экспорте базы {db_name}: {e}", exc_info=True)
        return None


def count_records_in_excel(excel_path: str) -> int:
    """
    Подсчитывает количество записей в Excel файле

    Параметры:
        excel_path: Путь к Excel файлу

    Возвращает:
        int: Количество записей оборудования
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active

        # Количество записей = все строки минус заголовки (5 строк технической информации)
        # Формула: общее количество строк - 5 (заголовки)
        record_count = ws.max_row - 5
        wb.close()

        return max(0, record_count)
    except Exception:
        return 0
