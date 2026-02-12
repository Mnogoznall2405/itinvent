#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Унифицированный сервис для экспорта данных в Excel.
Предоставляет общие методы для форматирования и создания Excel файлов.
"""
import logging
from abc import ABC, abstractmethod
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)


# ============================ КОНСТАНТЫ СТИЛЕЙ ============================

@dataclass(frozen=True)
class ExcelStyles:
    """Константы стилей для Excel файлов"""

    # Шрифты
    HEADER_FONT = Font(bold=True, size=11)
    TITLE_FONT = Font(bold=True, size=12, color='FFFFFF')
    BRANCH_FONT = Font(bold=True, size=13, color="000000")
    LOCATION_FONT = Font(bold=True, size=11)
    STATS_FONT = Font(bold=True)
    BOLD_FONT = Font(bold=True)

    # Заливки
    HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9")
    TITLE_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    DATE_RANGE_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    BRANCH_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7")
    LOCATION_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6")
    EMPLOYEE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC")

    # Выравнивание
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    TOP_ALIGNMENT = Alignment(vertical="top", wrap_text=False)

    # Границы
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


@dataclass(frozen=True)
class ColumnWidth:
    """Ширина колонок по умолчанию"""

    DATE = 12
    TIME = 10
    LOCATION = 20
    SERIAL = 25
    MODEL = 30
    MANUFACTURER = 15
    EMPLOYEE = 20
    INV_NO = 12
    DATABASE = 15
    COLOR = 15
    COMPONENT = 15
    RESERVE = 5
    TITLE = 30


# ============================ БАЗОВЫЙ КЛАСС ============================

class BaseExcelExporter(ABC):
    """
    Базовый класс для экспорта данных в Excel.
    Содержит общие методы для всех типов экспорта.
    """

    def __init__(self, output_dir: str = "exports"):
        """
        Инициализация экспортера

        Параметры:
            output_dir: Директория для сохранения файлов
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.styles = ExcelStyles()

    def generate_filename(self, prefix: str, date_range: str = None) -> str:
        """
        Генерирует имя файла с временной меткой

        Параметры:
            prefix: Префикс имени файла
            date_range: Диапазон дат (опционально)

        Возвращает:
            str: Полный путь к файлу
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        if date_range:
            filename = f"{prefix}_{date_range}_{timestamp}.xlsx"
        else:
            filename = f"{prefix}_{timestamp}.xlsx"

        return str(self.output_dir / filename)

    def format_date_range(self, start_date: datetime, end_date: datetime) -> str:
        """
        Форматирует диапазон дат в строку

        Параметры:
            start_date: Начальная дата
            end_date: Конечная дата

        Возвращает:
            str: Отформатированный диапазон дат
        """
        if start_date and end_date:
            return f"{start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}"
        return "все_даты"

    def apply_header_style(self, cell, title: str = None) -> None:
        """
        Применяет стиль заголовка к ячейке

        Параметры:
            cell: Ячейка Excel
            title: Заголовок (если есть, устанавливает значение)
        """
        if title:
            cell.value = title
        cell.font = self.styles.HEADER_FONT
        cell.fill = self.styles.HEADER_FILL
        cell.alignment = self.styles.CENTER_ALIGNMENT
        cell.border = self.styles.THIN_BORDER

    def apply_title_style(self, cell, title: str, column: int = 11) -> None:
        """
        Применяет стиль заголовка раздела к ячейке

        Параметры:
            cell: Ячейка Excel
            title: Заголовок
            column: Номер колонки для ячейки
        """
        cell.value = title
        cell.font = self.styles.TITLE_FONT
        cell.fill = self.styles.TITLE_FILL
        cell.alignment = self.styles.CENTER_ALIGNMENT

    def apply_date_range_style(self, cell, date_range: str, column: int = 11) -> None:
        """
        Применяет стиль диапазона дат к ячейке

        Параметры:
            cell: Ячейка Excel
            date_range: Диапазон дат
            column: Номер колонки для ячейки
        """
        cell.value = f'Период: {date_range}'
        cell.font = self.styles.BOLD_FONT
        cell.fill = self.styles.DATE_RANGE_FILL


# ============================ ЭКСПОРТ С ГРУППИРОВКОЙ ============================

class GroupedExcelExporter(BaseExcelExporter):
    """
    Экспортер данных с группировкой по филиалам.
    Создает отдельные листы для каждого филиала + сводный лист.
    """

    def export_by_branches(
        self,
        df: pd.DataFrame,
        output_file: str,
        sheet_title_prefix: str = "ФИЛИАЛ",
        summary_title: str = "СВОДНЫЙ ОТЧЕТ",
        date_range: str = None,
        branch_column: str = 'Филиал',
        column_widths: Dict[str, int] = None
    ) -> str:
        """
        Экспортирует DataFrame с группировкой по филиалам

        Параметры:
            df: DataFrame с данными (должен содержать колонку филиала)
            output_file: Путь к выходному файлу
            sheet_title_prefix: Префикс заголовка листа филиала
            summary_title: Заголовок сводного листа
            date_range: Диапазон дат для заголовка
            branch_column: Название колонки с филиалом
            column_widths: Словарь ширины колонок {'A': 10, 'B': 15, ...}

        Возвращает:
            str: Путь к созданному файлу
        """
        if df.empty:
            logger.warning("DataFrame пуст, нет данных для экспорта")
            return None

        branches = df[branch_column].unique()
        title_column = self._get_title_column(df)

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Создаем лист для каждого филиала
            for branch in branches:
                if pd.isna(branch):
                    continue

                branch_data = df[df[branch_column] == branch].copy()
                branch_data = branch_data.drop(branch_column, axis=1)

                sheet_name = str(branch)[:31]
                branch_data.to_excel(writer, sheet_name=sheet_name, index=False)

                # Форматируем лист филиала
                self._format_branch_sheet(
                    writer.sheets[sheet_name],
                    branch=branch,
                    title_prefix=sheet_title_prefix,
                    date_range=date_range,
                    title_column=title_column,
                    column_widths=column_widths
                )

            # Создаем сводный лист
            df_summary = df.drop(branch_column, axis=1)
            df_summary.to_excel(writer, sheet_name='Сводка', index=False)

            # Форматируем сводный лист
            self._format_summary_sheet(
                writer.sheets['Сводка'],
                title=summary_title,
                date_range=date_range,
                title_column=title_column,
                total_records=len(df_summary),
                total_branches=len(branches),
                column_widths=column_widths
            )

        logger.info(f"Создан файл с группировкой по филиалам: {output_file}")
        return output_file

    def _get_title_column(self, df: pd.DataFrame) -> int:
        """Определяет колонку для заголовка"""
        return len(df.columns) + 2

    def _format_branch_sheet(
        self,
        worksheet,
        branch: str,
        title_prefix: str,
        date_range: str,
        title_column: int,
        column_widths: Dict[str, int] = None
    ) -> None:
        """Форматирует лист филиала"""
        # Заголовок филиала
        title_cell = worksheet.cell(row=1, column=title_column, value=f'{title_prefix}: {branch}')
        self.apply_title_style(title_cell, f'{title_prefix}: {branch}')

        # Диапазон дат
        if date_range:
            date_cell = worksheet.cell(row=2, column=title_column)
            self.apply_date_range_style(date_cell, date_range)

        # Ширина колонок
        self._apply_column_widths(worksheet, column_widths)

    def _format_summary_sheet(
        self,
        worksheet,
        title: str,
        date_range: str,
        title_column: int,
        total_records: int,
        total_branches: int,
        column_widths: Dict[str, int] = None
    ) -> None:
        """Форматирует сводный лист"""
        # Заголовок
        title_cell = worksheet.cell(row=1, column=title_column, value=title)
        self.apply_title_style(title_cell, title)

        # Диапазон дат
        if date_range:
            date_cell = worksheet.cell(row=2, column=title_column)
            self.apply_date_range_style(date_cell, date_range)

        # Статистика
        stats_row = len(list(worksheet.rows)) + 5
        stats_cell = worksheet.cell(row=stats_row, column=1, value='СТАТИСТИКА')
        stats_cell.font = self.styles.STATS_FONT

        worksheet.cell(row=stats_row + 1, column=1, value=f'Всего записей: {total_records}')
        worksheet.cell(row=stats_row + 2, column=1, value=f'Филиалов: {total_branches}')

        # Ширина колонок
        self._apply_column_widths(worksheet, column_widths)

    def _apply_column_widths(
        self,
        worksheet,
        column_widths: Dict[str, int] = None
    ) -> None:
        """Применяет ширину колонок к листу"""
        if not column_widths:
            return

        for col, width in column_widths.items():
            if col in [cell.column_letter for cell in worksheet[1]]:
                worksheet.column_dimensions[col].width = width


# ============================ ЭКСПОРТ БАЗЫ ДАННЫХ ============================

class DatabaseExcelExporter(BaseExcelExporter):
    """
    Экспортер базы данных оборудования с иерархической группировкой
    по филиалам и локациям.
    """

    def export_database(
        self,
        rows: List[tuple],
        db_name: str,
        output_file: str = None
    ) -> str:
        """
        Экспортирует данные базы данных в Excel с группировкой

        Параметры:
            rows: Список кортежей с данными (из SQL запроса)
            db_name: Имя базы данных
            output_file: Путь к выходному файлу (если None, генерируется автоматически)

        Возвращает:
            str: Путь к созданному файлу
        """
        if not rows:
            logger.warning(f"Нет данных для экспорта базы {db_name}")
            return None

        if not output_file:
            output_file = self.generate_filename(f"{db_name}_export")

        wb = Workbook()
        ws = wb.active
        ws.title = "Экспорт оборудования"

        # Техническая информация
        self._add_header_info(ws, db_name, len(rows))

        # Заголовки колонок
        headers = self._get_database_headers()
        header_row = 5
        self._add_column_headers(ws, headers, header_row)

        # Данные с группировкой
        self._add_grouped_data(ws, rows, header_row + 1)

        # Ширина колонок
        self._set_database_column_widths(ws)

        wb.save(output_file)
        logger.info(f"Экспорт базы {db_name} завершен: {output_file} ({len(rows)} записей)")

        return output_file

    def _add_header_info(self, worksheet, db_name: str, record_count: int) -> None:
        """Добавляет техническую информацию в начало файла"""
        worksheet['A1'] = f"База данных: {db_name}"
        worksheet['A2'] = f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        worksheet['A3'] = f"Всего записей: {record_count}"

        worksheet['A1'].font = self.styles.BOLD_FONT

        # Объединяем ячейки
        worksheet.merge_cells('A1:N1')
        worksheet.merge_cells('A2:N2')
        worksheet.merge_cells('A3:N3')

    def _get_database_headers(self) -> List[str]:
        """Возвращает заголовки колонок для базы данных"""
        return [
            'Инв. №', 'Сотрудник', 'Тип', 'Серийный №', 'Апп. серийный №',
            'Партийный №', 'Модель', 'Производитель', 'Местоположение',
            'Таб. №', 'Отдел', 'Филиал', 'Статус', 'Описание'
        ]

    def _add_column_headers(self, worksheet, headers: List[str], row: int) -> None:
        """Добавляет заголовки колонок"""
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row, column=col_idx)
            self.apply_header_style(cell, header)

    def _add_grouped_data(self, worksheet, rows: List[tuple], start_row: int) -> None:
        """Добавляет данные с группировкой по филиалам и локациям"""
        current_row = start_row
        current_branch = object()  # Уникальный маркер
        current_location = None

        for row in rows:
            # Распаковка данных
            inv_no = row[0]
            employee_name = row[1]
            branch = row[11] or 'Не указан'
            location = row[8] or 'Не указано'

            # Новый филиал
            if branch != current_branch:
                if current_branch is not object():
                    current_row += 1  # Пустая строка между филиалами

                current_branch = branch
                current_location = None

                # Заголовок филиала
                worksheet.merge_cells(f'A{current_row}:N{current_row}')
                cell = worksheet.cell(row=current_row, column=1, value=f"🏢 {branch}")
                cell.font = self.styles.BRANCH_FONT
                cell.fill = self.styles.BRANCH_FILL
                cell.alignment = self.styles.LEFT_ALIGNMENT
                current_row += 1

            # Новая локация
            if location != current_location:
                current_location = location

                worksheet.merge_cells(f'A{current_row}:N{current_row}')
                cell = worksheet.cell(row=current_row, column=1, value=f"📍 {location}")
                cell.font = self.styles.LOCATION_FONT
                cell.fill = self.styles.LOCATION_FILL
                cell.alignment = self.styles.LEFT_ALIGNMENT
                current_row += 1

            # Данные оборудования
            self._add_equipment_row(worksheet, row, current_row)
            current_row += 1

    def _add_equipment_row(self, worksheet, row_data: tuple, row_num: int) -> None:
        """Добавляет строку с данными оборудования"""
        data = [
            row_data[0] or '',                      # Инв. №
            row_data[1] or 'Не назначен',           # Сотрудник
            row_data[2] or '',                      # Тип
            row_data[3] or '',                      # Серийный №
            row_data[4] or '',                      # Апп. серийный №
            row_data[5] or '',                      # Партийный №
            row_data[6] or '',                      # Модель
            row_data[7] or '',                      # Производитель
            row_data[8] or 'Не указано',            # Местоположение
            row_data[9] or '',                      # Таб. №
            row_data[10] or '',                     # Отдел
            row_data[11] or 'Не указан',            # Филиал
            row_data[12] or '',                     # Статус
            row_data[13] or ''                      # Описание
        ]

        for col_idx, value in enumerate(data, start=1):
            cell = worksheet.cell(row=row_num, column=col_idx, value=value)
            cell.border = self.styles.THIN_BORDER
            cell.alignment = self.styles.TOP_ALIGNMENT

            # Форматирование для важных полей
            if col_idx == 4:  # Серийный номер
                cell.font = self.styles.BOLD_FONT
            elif col_idx == 2 and data[1] != 'Не назначен':  # Сотрудник
                cell.fill = self.styles.EMPLOYEE_FILL

    def _set_database_column_widths(self, worksheet) -> None:
        """Устанавливает ширину колонок для базы данных"""
        widths = [12, 20, 15, 18, 18, 12, 20, 15, 25, 10, 18, 15, 12, 35]
        for col_idx, width in enumerate(widths, start=1):
            worksheet.column_dimensions[chr(64 + col_idx)].width = width


# ============================ ПРОСТОЙ ЭКСПОРТЕР ============================

class SimpleExcelExporter(BaseExcelExporter):
    """
    Простой экспортер для базовых случаев без группировки.
    Просто сохраняет DataFrame в Excel с базовым форматированием.
    """

    def export_dataframe(
        self,
        df: pd.DataFrame,
        output_file: str,
        title: str = None
    ) -> str:
        """
        Экспортирует DataFrame в Excel с базовым форматированием

        Параметры:
            df: DataFrame для экспорта
            output_file: Путь к выходному файлу
            title: Заголовок (опционально)

        Возвращает:
            str: Путь к созданному файлу
        """
        if df.empty:
            logger.warning("DataFrame пуст, нет данных для экспорта")
            return None

        # Сохраняем в Excel
        df.to_excel(output_file, index=False, engine='openpyxl')

        # Применяем базовое форматирование
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active

        # Форматируем заголовки
        for cell in ws[1]:
            self.apply_header_style(cell)

        # Добавляем заголовок если указан
        if title:
            ws.insert_rows(1)
            ws['A1'] = title
            ws['A1'].font = self.styles.BOLD_FONT
            ws.merge_cells(f'A1:{chr(64 + len(df.columns))}1')

        wb.save(output_file)
        logger.info(f"Создан простой Excel файл: {output_file}")

        return output_file


# ============================ ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ============================

def filter_data_by_period(data: list, period: str) -> tuple:
    """
    Фильтрует данные по указанному периоду

    Параметры:
        data: Список записей
        period: Период (1month, 3months, all)

    Возвращает:
        tuple: (отфильтрованные данные, начальная дата, конечная дата)
    """
    from datetime import datetime, timedelta

    try:
        now = datetime.now()

        if period == "all":
            if not data:
                return [], None, None

            dates = []
            for item in data:
                if 'timestamp' in item:
                    try:
                        item_date = datetime.fromisoformat(item['timestamp'].replace('Z', '+00:00'))
                        dates.append(item_date)
                    except:
                        continue

            if dates:
                return data, min(dates).date(), max(dates).date()
            return data, None, None

        # Вычисляем начальную дату
        if period == "1month":
            start_date = now - timedelta(days=30)
        elif period == "3months":
            start_date = now - timedelta(days=90)
        else:
            return data, None, None

        # Фильтруем данные
        filtered_data = []
        for item in data:
            if 'timestamp' in item:
                try:
                    item_date = datetime.fromisoformat(item['timestamp'].replace('Z', '+00:00'))
                    if item_date >= start_date:
                        filtered_data.append(item)
                except:
                    continue

        return filtered_data, start_date.date(), now.date()

    except Exception as e:
        logger.error(f"Ошибка фильтрации по периоду: {e}")
        return data, None, None


def count_excel_records(excel_path: str) -> int:
    """
    Подсчитывает количество записей в Excel файле

    Параметры:
        excel_path: Путь к Excel файлу

    Возвращает:
        int: Количество записей
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        record_count = ws.max_row - 5  # Вычитаем заголовки
        wb.close()
        return max(0, record_count)
    except Exception:
        return 0
